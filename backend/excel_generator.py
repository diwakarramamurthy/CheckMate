"""
RERA Reports Excel Generator - Goa State Official Format
Generates Excel (.xlsx) versions of Form-1, Form-3, Form-4, and Annexure-A
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


# ─── Helpers ────────────────────────────────────────────────────────────────

def format_indian_number(num):
    if num is None or num == 0:
        return "0"
    num = int(num)
    is_negative = num < 0
    num = abs(num)
    s = str(num)
    if len(s) <= 3:
        result = s
    else:
        result = s[-3:]
        s = s[:-3]
        while s:
            result = s[-2:] + "," + result
            s = s[:-2]
    return ("-" + result) if is_negative else result


def get_cat_avg(cat_data):
    if not cat_data:
        return 0
    vals = [v.get("completion", 0) for v in cat_data.values() if isinstance(v, dict)]
    return sum(vals) / len(vals) if vals else 0


# ─── Shared Styles ───────────────────────────────────────────────────────────

TITLE_FILL    = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FILL   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SECTION_FILL  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FILL    = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def title_cell(cell, text):
    cell.value = text
    cell.font = Font(bold=True, size=13, color="FFFFFF")
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def header_cell(cell, text):
    cell.value = text
    cell.font = Font(bold=True, size=9, color="FFFFFF")
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN


def section_cell(cell, text):
    cell.value = text
    cell.font = Font(bold=True, size=10)
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN


def total_cell(cell, text, align="right"):
    cell.value = text
    cell.font = Font(bold=True, size=9)
    cell.fill = TOTAL_FILL
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = THIN


def data_cell(cell, text, align="left", bold=False):
    cell.value = text
    cell.font = Font(size=9, bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = THIN


def label_cell(cell, text):
    cell.value = text
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def value_cell(cell, text):
    cell.value = str(text) if text else ""
    cell.font = Font(size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")


# ─── FORM 1 ──────────────────────────────────────────────────────────────────

def generate_form1_excel(project, buildings, construction_progress, infrastructure_progress, quarter, year):
    wb = Workbook()

    # ── Sheet 1: Project Info ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Project Info"

    ws.merge_cells("A1:F1")
    title_cell(ws["A1"], "The Goa Real Estate (Regulation and Development) Rules 2017")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    title_cell(ws["A2"], "FORM 1 — ARCHITECT'S CERTIFICATE")
    ws.row_dimensions[2].height = 24

    ws.merge_cells("A3:F3")
    ws["A3"].value = "(Percentage of Completion of Construction)"
    ws["A3"].font = Font(italic=True, size=10, color="FFFFFF")
    ws["A3"].fill = TITLE_FILL
    ws["A3"].alignment = Alignment(horizontal="center")

    info_rows = [
        ("Project Name:", project.get("project_name", "")),
        ("RERA Registration No.:", project.get("rera_number", "")),
        ("Report Period:", f"{quarter} {year}"),
        ("Report Date:", datetime.now().strftime("%d/%m/%Y")),
        ("Promoter Name:", project.get("promoter_name", "")),
        ("Architect Name:", project.get("architect_name", "")),
        ("Architect License No.:", project.get("architect_license", "")),
        ("Village / Panchayat:", project.get("village", "")),
        ("Taluka:", project.get("taluka", "")),
        ("District:", project.get("district", "North Goa")),
        ("Survey No. / Plot No.:", project.get("survey_number", "")),
        ("Total Area (sq.m.):", str(project.get("total_area", ""))),
    ]
    for i, (lbl, val) in enumerate(info_rows, 5):
        ws.merge_cells(f"A{i}:B{i}")
        label_cell(ws[f"A{i}"], lbl)
        ws.merge_cells(f"C{i}:F{i}")
        value_cell(ws[f"C{i}"], val)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    for col in ["C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 18

    # ── Sheet 2: Table A — Building Progress ──────────────────────────────
    progress_lookup = {p.get("building_id"): p for p in (construction_progress or [])}

    ws_a = wb.create_sheet("Table A - Buildings")
    ws_a.merge_cells("A1:E1")
    title_cell(ws_a["A1"], "TABLE A: Percentage of Completion per Building/Wing")
    ws_a.row_dimensions[1].height = 28

    ws_a.merge_cells("A2:E2")
    ws_a["A2"].value = (
        f"Project: {project.get('project_name', '')}  |  "
        f"RERA No: {project.get('rera_number', '')}  |  Period: {quarter} {year}"
    )
    ws_a["A2"].font = Font(size=10)
    ws_a.row_dimensions[2].height = 18

    for col, hdr in enumerate(["Sr. No.", "Building / Wing", "Activity", "% Work Done", "Remarks"], 1):
        header_cell(ws_a.cell(row=4, column=col), hdr)
    ws_a.row_dimensions[4].height = 28

    row = 5
    sr = 1
    for building in (buildings or []):
        bname = building.get("building_name", "Building")
        progress = progress_lookup.get(building.get("building_id"), {})
        ta = progress.get("tower_activities", {})

        plinth        = ta.get("plinth_completion", {})
        basement_slab = ta.get("basement_slab_completion", {})
        slab          = ta.get("slab_completion", {})
        brickwk  = ta.get("brickwork_plastering", {})
        plumbing = ta.get("plumbing", {})
        elec     = ta.get("electrical_works", {})
        windows  = ta.get("window_works", {})
        tiling   = ta.get("tiling_flooring", {})
        doors    = ta.get("door_shutter_fixing", {})
        waterp   = ta.get("water_proofing", {})
        painting = ta.get("painting", {})
        carpark  = ta.get("carpark", {})

        # Building header
        ws_a.merge_cells(f"A{row}:E{row}")
        section_cell(
            ws_a[f"A{row}"],
            f"Building: {bname}  —  Overall Completion: {progress.get('overall_completion', 0):.1f}%",
        )
        ws_a.row_dimensions[row].height = 20
        row += 1

        activities = [
            ("Excavation",
             f"{plinth.get('excavation', {}).get('completion', 0):.0f}%"),
            ("Basements and Plinth",
             f"{((get_cat_avg(plinth) + get_cat_avg(basement_slab)) / 2 if any(isinstance(v, dict) and v.get('completion', 0) > 0 for v in basement_slab.values()) else get_cat_avg(plinth)):.0f}%"),
            (f"{building.get('residential_floors', 0)} Slabs of Super Structure",
             f"{get_cat_avg(slab):.0f}%"),
            ("Internal Walls, Plaster, Flooring (Brickwork & Plastering)",
             f"{get_cat_avg(brickwk):.0f}%"),
            ("Doors and Windows",
             f"{(get_cat_avg(doors) + get_cat_avg(windows)) / 2:.0f}%"),
            ("Sanitary and Electrical Fittings",
             f"{(get_cat_avg(plumbing) + get_cat_avg(elec)) / 2:.0f}%"),
            ("Tiling / Flooring",
             f"{get_cat_avg(tiling):.0f}%"),
            ("Water Proofing",
             f"{get_cat_avg(waterp):.0f}%"),
            ("Painting",
             f"{get_cat_avg(painting):.0f}%"),
            ("Carpark",
             f"{get_cat_avg(carpark):.0f}%"),
        ]
        for act_name, pct in activities:
            data_cell(ws_a.cell(row=row, column=1), str(sr), align="center")
            ws_a.merge_cells(f"B{row}:C{row}")
            data_cell(ws_a.cell(row=row, column=2), act_name)
            ws_a.cell(row=row, column=3).border = THIN
            data_cell(ws_a.cell(row=row, column=4), pct, align="center")
            ws_a.cell(row=row, column=5).border = THIN
            sr += 1
            row += 1
        row += 1  # blank spacer

    ws_a.column_dimensions["A"].width = 8
    ws_a.column_dimensions["B"].width = 35
    ws_a.column_dimensions["C"].width = 12
    ws_a.column_dimensions["D"].width = 15
    ws_a.column_dimensions["E"].width = 22

    # ── Sheet 3: Table B — Infrastructure ─────────────────────────────────
    ws_b = wb.create_sheet("Table B - Infrastructure")
    ws_b.merge_cells("A1:F1")
    title_cell(ws_b["A1"], "TABLE B: Internal & External Development Works")
    ws_b.row_dimensions[1].height = 28

    for col, hdr in enumerate(
        ["Sr. No.", "Common Areas / Facilities", "Proposed (Yes/No)", "% Work Done", "Details"], 1
    ):
        header_cell(ws_b.cell(row=3, column=col), hdr)
    ws_b.row_dimensions[3].height = 28

    infra = infrastructure_progress.get("activities", {}) if infrastructure_progress else {}

    def ic(key):
        return infra.get(key, {}).get("completion", 0)

    infra_rows = [
        ("1.", "Internal Roads & Footpaths", "Yes", f"{ic('road_footpath_storm_drain'):.0f}%"),
        ("2.", "Water Supply", "Yes", f"{ic('underground_water_distribution'):.0f}%"),
        ("3.", "Sewerage (STP)", "Yes", f"{ic('underground_sewage_network'):.0f}%"),
        ("4.", "Storm Water Drains", "Yes", f"{ic('road_footpath_storm_drain'):.0f}%"),
        ("5.", "Landscaping & Tree Planting", "Yes", f"{ic('gardens_playground'):.0f}%"),
        ("6.", "Street Lighting", "Yes", f"{ic('street_lights'):.0f}%"),
        ("7.", "Community Buildings (Club House)", "Yes", f"{ic('club_house'):.0f}%"),
        ("8.", "Treatment & Disposal of Sewage", "Yes", f"{ic('sewage_treatment_plant'):.0f}%"),
        ("9.", "Solid Waste Management", "Yes", "0%"),
        ("10.", "Water Conservation / Rain Water Harvesting", "Yes", f"{ic('overhead_sump_reservoir'):.0f}%"),
        ("11.", "Energy Management", "Yes", f"{ic('electric_substation_cables'):.0f}%"),
        ("12.", "Fire Protection & Safety", "Yes", "0%"),
        ("13.", "Electrical Meter Room / Substation", "Yes", f"{ic('electric_substation_cables'):.0f}%"),
        ("14.", "Swimming Pool", "Yes", f"{ic('swimming_pool'):.0f}%"),
        ("15.", "Amphitheatre", "Yes", f"{ic('amphitheatre'):.0f}%"),
        ("16.", "Boundary Wall & Entry Gate", "Yes",
         f"{(ic('boundary_wall') + ic('entry_gate')) / 2:.0f}%"),
    ]
    for r, (sr_no, item, proposed, pct) in enumerate(infra_rows, 4):
        data_cell(ws_b.cell(row=r, column=1), sr_no, align="center")
        data_cell(ws_b.cell(row=r, column=2), item)
        data_cell(ws_b.cell(row=r, column=3), proposed, align="center")
        data_cell(ws_b.cell(row=r, column=4), pct, align="center")
        ws_b.cell(row=r, column=5).border = THIN

    ws_b.column_dimensions["A"].width = 8
    ws_b.column_dimensions["B"].width = 42
    ws_b.column_dimensions["C"].width = 16
    ws_b.column_dimensions["D"].width = 14
    ws_b.column_dimensions["E"].width = 22

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── FORM 3 ──────────────────────────────────────────────────────────────────

def generate_form3_excel(project, buildings, construction_progress, infrastructure_progress, estimated_dev_cost, quarter, year):
    """Form-3: Cost Incurred = (% Work Completed from Construction Progress) × Estimated Cost"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Form 3 - Cost Incurred"

    ws.merge_cells("A1:G1")
    title_cell(ws["A1"], "The Goa Real Estate (Regulation and Development) Rules 2017")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:G2")
    title_cell(ws["A2"], "FORM 3 — ENGINEER'S CERTIFICATE")
    ws.row_dimensions[2].height = 24
    ws.merge_cells("A3:G3")
    ws["A3"].value = "(Cost Incurred Statement for Development)"
    ws["A3"].font = Font(italic=True, size=10, color="FFFFFF")
    ws["A3"].fill = TITLE_FILL
    ws["A3"].alignment = Alignment(horizontal="center")

    info = [
        ("Project Name:", project.get("project_name", "")),
        ("RERA No.:", project.get("rera_number", "")),
        ("Report Period:", f"{quarter} {year}"),
        ("Report Date:", datetime.now().strftime("%d/%m/%Y")),
        ("Engineer Name:", project.get("engineer_name", "")),
        ("Engineer License No.:", project.get("engineer_license", "")),
    ]
    for i, (lbl, val) in enumerate(info, 5):
        ws.merge_cells(f"A{i}:B{i}")
        label_cell(ws[f"A{i}"], lbl)
        ws.merge_cells(f"C{i}:G{i}")
        value_cell(ws[f"C{i}"], val)

    # TABLE A — Cost Incurred = completion% × estimated_cost
    row = 12
    ws.merge_cells(f"A{row}:G{row}")
    section_cell(ws[f"A{row}"], "TABLE A: Cost Incurred for Building Construction")
    ws.row_dimensions[row].height = 22
    row += 1

    hdrs = ["Sr.", "Building / Wing", "% Complete", "Estimated Cost (₹)", "Cost Incurred (₹)", "Balance (₹)", ""]
    for c, h in enumerate(hdrs, 1):
        header_cell(ws.cell(row=row, column=c), h)
    ws.row_dimensions[row].height = 28
    row += 1

    cp_lookup = {cp.get("building_id"): cp for cp in (construction_progress or [])}
    total_est = total_inc = 0

    for idx, b in enumerate(buildings or [], 1):
        progress = cp_lookup.get(b.get("building_id"), {})
        completion_pct = progress.get("overall_completion", 0)
        est = b.get("estimated_cost", 0)
        inc = round((completion_pct / 100) * est)
        bal = est - inc
        total_est += est
        total_inc += inc

        data_cell(ws.cell(row=row, column=1), str(idx), align="center")
        data_cell(ws.cell(row=row, column=2), b.get("building_name", ""))
        data_cell(ws.cell(row=row, column=3), f"{completion_pct:.1f}%", align="center")
        data_cell(ws.cell(row=row, column=4), format_indian_number(est), align="right")
        data_cell(ws.cell(row=row, column=5), format_indian_number(inc), align="right")
        data_cell(ws.cell(row=row, column=6), format_indian_number(bal), align="right")
        row += 1

    for c in range(1, 8):
        total_cell(ws.cell(row=row, column=c), "")
    total_cell(ws.cell(row=row, column=2), "TOTAL", align="center")
    total_cell(ws.cell(row=row, column=4), format_indian_number(total_est), align="right")
    total_cell(ws.cell(row=row, column=5), format_indian_number(total_inc), align="right")
    total_cell(ws.cell(row=row, column=6), format_indian_number(total_est - total_inc), align="right")
    row += 2

    # TABLE B — Infra Cost Incurred = completion% × estimated_infra_cost
    ws.merge_cells(f"A{row}:G{row}")
    section_cell(ws[f"A{row}"], "TABLE B: Cost of Internal/External Development Works")
    ws.row_dimensions[row].height = 22
    row += 1

    est_c = estimated_dev_cost or {}
    infra_cost = est_c.get("infrastructure_cost", 0)
    infra_completion = (infrastructure_progress or {}).get("overall_completion", 0)
    infra_incurred = round((infra_completion / 100) * infra_cost)
    infra_balance = infra_cost - infra_incurred

    b_hdrs = ["Sr.", "Development Work", "% Complete", "Estimated (₹)", "Incurred (₹)", "Balance (₹)", ""]
    for c, h in enumerate(b_hdrs, 1):
        header_cell(ws.cell(row=row, column=c), h)
    ws.row_dimensions[row].height = 28
    row += 1

    b_rows = [
        ("1", "Internal Roads & Footpaths"), ("2", "Water Supply & Distribution"),
        ("3", "Sewerage & STP"), ("4", "Storm Water Drains"),
        ("5", "Landscaping"), ("6", "Street Lighting"),
        ("7", "Club House"), ("8", "Swimming Pool"),
        ("9", "Electrical Infrastructure"), ("10", "Boundary Wall & Gate"),
    ]
    for sr_no, work in b_rows:
        data_cell(ws.cell(row=row, column=1), sr_no, align="center")
        data_cell(ws.cell(row=row, column=2), work)
        for c in range(3, 8):
            data_cell(ws.cell(row=row, column=c), "—", align="center")
        row += 1

    for c in range(1, 8):
        total_cell(ws.cell(row=row, column=c), "")
    total_cell(ws.cell(row=row, column=2), "TOTAL INFRASTRUCTURE", align="center")
    total_cell(ws.cell(row=row, column=3), f"{infra_completion:.1f}%", align="center")
    total_cell(ws.cell(row=row, column=4), format_indian_number(infra_cost), align="right")
    total_cell(ws.cell(row=row, column=5), format_indian_number(infra_incurred), align="right")
    total_cell(ws.cell(row=row, column=6), format_indian_number(infra_balance), align="right")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    for col in ["D", "E", "F"]:
        ws.column_dimensions[col].width = 20
    ws.column_dimensions["G"].width = 4

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── FORM 4 ──────────────────────────────────────────────────────────────────
# Helper styles specific to Form-4 (CA Certificate official format)

F4_TITLE_FILL   = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
F4_SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
F4_SUBTOT_FILL  = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
F4_SUMMARY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
F4_NET_FILL     = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
F4_NOTE_FILL    = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

THIN_ALL = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)


def _f4_cell(ws, addr, value, bold=False, size=9, color="000000",
             h="left", v="center", wrap=True, number_fmt=None, fill=None):
    """Write a single cell for Form-4 with consistent styling."""
    c = ws[addr] if isinstance(addr, str) else addr
    c.value = value
    c.font = Font(name="Arial", size=size, bold=bold, color=color)
    c.alignment = Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    c.border = THIN_ALL
    if number_fmt:
        c.number_format = number_fmt
    if fill:
        c.fill = fill


def _f4_merge(ws, rng, value, bold=False, size=9, color="000000",
              h="left", v="center", wrap=True, fill=None):
    """Merge a range and write a single styled cell."""
    ws.merge_cells(rng)
    top_left = rng.split(":")[0]
    _f4_cell(ws, top_left, value, bold=bold, size=size, color=color,
             h=h, v=v, wrap=wrap, fill=fill)
    # Apply border to all cells in the merged range
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(rng)
    for r in range(min_row, max_row + 1):
        for cl in range(min_col, max_col + 1):
            ws.cell(row=r, column=cl).border = THIN_ALL


def _na_or_num(val):
    """Return 'NA' if value is 0/None, else the numeric value (int if whole)."""
    if not val:
        return "NA"
    return int(val) if val == int(val) else val


def _fmt_num(val):
    """Return formatted Indian number string, or 'NA' if zero."""
    if not val:
        return "NA"
    return format_indian_number(int(val))


# Indian rupee formats for openpyxl cells
_RUPEE    = '[>=10000000]##\\,##\\,##\\,##0;[>=100000]##\\,##\\,##0;##\\,##0'
_RUPEE_D  = '[>=10000000]##\\,##\\,##\\,##0.00;[>=100000]##\\,##\\,##0.00;##\\,##0.00'
_PCT_FMT  = '0.00%'


def generate_form4_excel(project, form4_data, quarter, year):
    """
    Form-4: CA Certificate – official format matching RERA CA Certificate template.
    Columns: A=Sr No, B=Particulars, C=Estimated Amount (Rs.), D=Incurred Amount (Rs.)
    form4_data is a pre-computed dict from server._build_form4_data().
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Form-4 CA Certificate"

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 78
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22

    fd = form4_data or {}

    # Unpack all pre-computed values
    lc_a_est      = fd.get("lc_a_est", 0);    lc_a_inc      = fd.get("lc_a_inc", 0)
    lc_b_est      = fd.get("lc_b_est", 0);    lc_b_inc      = fd.get("lc_b_inc", 0)
    lc_c_est      = fd.get("lc_c_est", 0);    lc_c_inc      = fd.get("lc_c_inc", 0)
    lc_d_est      = fd.get("lc_d_est", 0);    lc_d_inc      = fd.get("lc_d_inc", 0)
    lc_e_est      = fd.get("lc_e_est", 0);    lc_e_inc      = fd.get("lc_e_inc", 0)
    rehab_i_est   = fd.get("rehab_i_est", 0); rehab_i_inc   = fd.get("rehab_i_inc", 0)
    rehab_ii_inc  = fd.get("rehab_ii_inc", 0)
    rehab_iii_inc = fd.get("rehab_iii_inc", 0)
    rehab_iv_inc  = fd.get("rehab_iv_inc", 0)
    rehab_any     = fd.get("rehab_any", False)
    land_sub_est  = fd.get("land_sub_est", 0); land_sub_inc  = fd.get("land_sub_inc", 0)

    dev_a1_est    = fd.get("dev_a1_est", 0)
    dev_a2_inc    = fd.get("dev_a2_inc", 0)
    dev_a3_est    = fd.get("dev_a3_est", 0);   dev_a3_inc    = fd.get("dev_a3_inc", 0)
    dev_b_est     = fd.get("dev_b_est", 0);    dev_b_inc     = fd.get("dev_b_inc", 0)
    dev_c_est     = fd.get("dev_c_est", 0);    dev_c_inc     = fd.get("dev_c_inc", 0)
    dev_sub_est   = fd.get("dev_sub_est", 0);  dev_sub_inc   = fd.get("dev_sub_inc", 0)

    total_est     = fd.get("total_est", 0);    total_inc     = fd.get("total_inc", 0)
    arch_pct      = fd.get("arch_pct", 0)
    proportion    = fd.get("proportion", 0)
    withdraw_allow = fd.get("withdraw_allow", 0)
    total_amount_received_sold = fd.get("total_amount_received_sold", 0)
    net_withdraw  = fd.get("net_withdraw", 0)

    bal_cost      = fd.get("bal_cost", 0)
    bal_recv_sold = fd.get("bal_recv_sold", 0)
    unsold_area   = fd.get("unsold_area", 0)
    asr_rate      = fd.get("asr_rate", 0)
    unsold_val    = fd.get("unsold_val", 0)
    total_recv    = fd.get("total_recv", 0)
    deposit_pct      = fd.get("deposit_pct", 0.70)
    deposit_amt      = fd.get("deposit_amt", 0)
    avg_sale_price   = fd.get("avg_sale_price", 0)
    total_sale_val_sold = fd.get("total_sale_val_sold", 0)
    sales            = fd.get("sales", [])
    buildings        = fd.get("buildings", [])
    building_map     = fd.get("building_map", {})

    # ─────────────────────────────────────────────────────────────────────────
    # BUILD WORKSHEET
    # ─────────────────────────────────────────────────────────────────────────

    def _row_ht(r, h):
        ws.row_dimensions[r].height = h

    R = 1   # current row pointer

    # ── TITLE ROWS ──────────────────────────────────────────────────────
    _row_ht(R, 30)
    _f4_merge(ws, f"A{R}:D{R}", "Form-4: CA Certificate",
              bold=True, size=13, h="center", fill=F4_TITLE_FILL, color="FFFFFF")
    R += 1

    _row_ht(R, 28)
    _f4_merge(ws, f"A{R}:D{R}",
              "(FOR REGISTRATION OF A PROJECT AND SUBSEQUENT WITHDRAWAL OF MONEY)",
              bold=True, size=10, h="center", fill=F4_TITLE_FILL, color="FFFFFF")
    R += 1

    # Project info strip
    proj_info = (
        f"Project: {project.get('project_name', '—')}   |   "
        f"RERA No.: {project.get('rera_number', '—')}   |   "
        f"Period: {quarter} {year}   |   Date: {datetime.now().strftime('%d/%m/%Y')}"
    )
    _row_ht(R, 20)
    _f4_merge(ws, f"A{R}:D{R}", proj_info, size=9, h="center", fill=F4_SECTION_FILL)
    R += 1

    # ── COLUMN HEADER ROW ───────────────────────────────────────────────
    _row_ht(R, 35)
    _f4_cell(ws, f"A{R}", "Sr No",  bold=True, size=10, h="center", fill=F4_SECTION_FILL)
    _f4_cell(ws, f"B{R}", "Particulars", bold=True, size=10, h="center", fill=F4_SECTION_FILL)
    _f4_cell(ws, f"C{R}", "Estimated Amount in Rs.", bold=True, size=10, h="center", fill=F4_SECTION_FILL)
    _f4_cell(ws, f"D{R}", "Incurred Amount in Rs.", bold=True, size=10, h="center", fill=F4_SECTION_FILL)
    R += 1

    # ─────────────────────────────────────────────────────────────────────────
    # SECTION 1 — LAND COST
    # ─────────────────────────────────────────────────────────────────────────
    _row_ht(R, 20)
    _f4_cell(ws, f"A{R}", None, fill=F4_SECTION_FILL)
    _f4_merge(ws, f"B{R}:D{R}", "i.  Land Cost :", bold=True, size=10, fill=F4_SECTION_FILL)
    R += 1

    # Helper: write a line-item row
    def _item(sr, text, c_val, d_val, ht=35, note=False):
        nonlocal R
        _row_ht(R, ht)
        if sr:
            _f4_cell(ws, f"A{R}", 1, bold=False, size=9, h="center")  # Sr 1 spans land section
        else:
            _f4_cell(ws, f"A{R}", None)
        _f4_cell(ws, f"B{R}", text, size=9, fill=F4_NOTE_FILL if note else None)
        # C
        if c_val == "NA":
            _f4_cell(ws, f"C{R}", "NA", bold=True, h="center", size=9)
        elif isinstance(c_val, (int, float)) and c_val:
            _f4_cell(ws, f"C{R}", c_val, bold=True, h="right", size=9, number_fmt=_RUPEE)
        else:
            _f4_cell(ws, f"C{R}", "NA", bold=True, h="center", size=9)
        # D
        if d_val == "NA":
            _f4_cell(ws, f"D{R}", "NA", bold=True, h="center", size=9)
        elif isinstance(d_val, (int, float)) and d_val:
            _f4_cell(ws, f"D{R}", d_val, bold=True, h="right", size=9, number_fmt=_RUPEE_D)
        else:
            _f4_cell(ws, f"D{R}", "NA", bold=True, h="center", size=9)
        R += 1

    _item(True,
          "a. Acquisition Cost of Land or Development Rights, lease Premium, lease rent, "
          "interest cost incurred or payable on Land Cost and legal cost.",
          lc_a_est, lc_a_inc, ht=45)

    _item(None,
          "b. Amount of Premium payable to obtain development rights, FSI, additional FSI, "
          "fungible area, and any other incentive under DCR from Local Authority or State "
          "Government or any Statutory Authority.",
          lc_b_est, lc_b_inc, ht=45)

    _item(None, "c. Acquisition cost of TDR (if any)", lc_c_est, lc_c_inc, ht=28)

    _item(None,
          "d. Amounts payable to State Government or competent authority or any other statutory "
          "authority of the State or Central Government, towards stamp duty, transfer charges, "
          "registration fees etc; and",
          lc_d_est, lc_d_inc, ht=45)

    _item(None,
          "e. Land Premium payable as per annual statement of rates (ASR) for redevelopment of "
          "land owned by public authorities.",
          lc_e_est, lc_e_inc, ht=35)

    # f. Rehabilitation scheme
    _row_ht(R, 18)
    _f4_cell(ws, f"A{R}", None)
    _f4_cell(ws, f"B{R}", "f. Under Rehabilitation Scheme :", bold=True, size=9)
    _f4_cell(ws, f"C{R}", None); _f4_cell(ws, f"D{R}", None)
    R += 1

    _item(None,
          "   (i) Estimated construction cost of rehab building including site development "
          "and infrastructure for the same as certified by Engineer.",
          rehab_i_est if rehab_any else "NA",
          rehab_i_inc if rehab_any else "NA", ht=40)

    _item(None,
          "   (ii) Actual Cost of construction of rehab building incurred as per the books "
          "of accounts as verified by the CA.",
          "NA",
          rehab_ii_inc if rehab_any else "NA", ht=35)

    # Note row (merged B:D)
    _row_ht(R, 25)
    _f4_cell(ws, f"A{R}", None)
    _f4_merge(ws, f"B{R}:D{R}",
              "Note : (for total cost of construction incurred, Minimum of (i) or (ii) is to be considered).",
              size=9, fill=F4_NOTE_FILL)
    R += 1

    _item(None,
          "   (iii) Cost towards clearance of land of all or any encumbrances including cost of "
          "removal of legal/illegal occupants, cost for providing temporary transit accommodation "
          "or rent in lieu of Transit Accommodation, overhead cost,",
          "NA", rehab_iii_inc if rehab_any else "NA", ht=55)

    _item(None,
          "   (iv) Cost of ASR linked premium, fees, charges and security deposits or maintenance "
          "deposit, or any amount whatsoever payable to any authorities towards and in project of "
          "rehabilitation.",
          "NA", rehab_iv_inc if rehab_any else "NA", ht=45)

    # Sub-total Land Cost
    _row_ht(R, 22)
    _f4_cell(ws, f"A{R}", None, fill=F4_SUBTOT_FILL)
    _f4_cell(ws, f"B{R}", "Sub-Total of LAND COST",
             bold=True, size=10, fill=F4_SUBTOT_FILL)
    _f4_cell(ws, f"C{R}", land_sub_est if land_sub_est else "NA",
             bold=True, h="right", size=10,
             number_fmt=_RUPEE if land_sub_est else None,
             fill=F4_SUBTOT_FILL)
    _f4_cell(ws, f"D{R}", land_sub_inc if land_sub_inc else "NA",
             bold=True, h="right", size=10,
             number_fmt=_RUPEE_D if land_sub_inc else None,
             fill=F4_SUBTOT_FILL)
    LAND_SUBTOT_ROW = R
    R += 1

    # ─────────────────────────────────────────────────────────────────────────
    # SECTION 1.ii — DEVELOPMENT / CONSTRUCTION COST
    # ─────────────────────────────────────────────────────────────────────────
    _row_ht(R, 20)
    _f4_cell(ws, f"A{R}", None, fill=F4_SECTION_FILL)
    _f4_merge(ws, f"B{R}:D{R}", "ii.  Development Cost / Cost of Construction :",
              bold=True, size=10, fill=F4_SECTION_FILL)
    R += 1

    _row_ht(R, 30)
    _f4_cell(ws, f"A{R}", None)
    _f4_cell(ws, f"B{R}",
             "a. (i) Estimated Cost of Construction as certified by Engineer.", size=9)
    _f4_cell(ws, f"C{R}", dev_a1_est if dev_a1_est else "NA",
             bold=True, h="right", size=9,
             number_fmt=_RUPEE if dev_a1_est else None)
    _f4_cell(ws, f"D{R}", "Refer Note", size=9, h="center")
    CONST_EST_ROW = R
    R += 1

    _row_ht(R, 30)
    _f4_cell(ws, f"A{R}", None)
    _f4_cell(ws, f"B{R}",
             "   (ii) Actual Cost of construction incurred as per the books of accounts "
             "as verified by the CA.", size=9)
    _f4_cell(ws, f"C{R}", "Refer Note", size=9, h="center")
    _f4_cell(ws, f"D{R}", dev_a2_inc if dev_a2_inc else "NA",
             bold=True, h="right", size=9,
             number_fmt=_RUPEE_D if dev_a2_inc else None)
    CONST_ACT_ROW = R
    R += 1

    # Note: MIN of (i) or (ii)
    _row_ht(R, 25)
    _f4_cell(ws, f"A{R}", None)
    _f4_merge(ws, f"B{R}:D{R}",
              "Note : (for adding to total cost of construction incurred, Minimum of (i) or (ii) "
              "is to be considered).",
              size=9, fill=F4_NOTE_FILL)
    R += 1

    # a(iii) On-site expenditure
    _row_ht(R, 80)
    _f4_cell(ws, f"A{R}", None)
    _f4_cell(ws, f"B{R}",
             "(iii) On-site expenditure for development of entire project excluding cost of "
             "construction as per (i) or (ii) above, i.e. salaries, consultants fees, site "
             "overheads, development works, cost of services (including water, electricity, "
             "sewerage, drainage, layout roads etc.), cost of machineries and equipment "
             "including its hire and maintenance costs, consumables etc. All costs directly "
             "incurred to complete the construction of the entire phase of the project registered.",
             size=9)
    _f4_cell(ws, f"C{R}", dev_a3_est if dev_a3_est else "NA",
             bold=True, h="right", size=9,
             number_fmt=_RUPEE if dev_a3_est else None)
    _f4_cell(ws, f"D{R}", dev_a3_inc if dev_a3_inc else "NA",
             bold=True, h="right", size=9,
             number_fmt=_RUPEE_D if dev_a3_inc else None)
    R += 1

    # b. Taxes
    _row_ht(R, 35)
    _f4_cell(ws, f"A{R}", None)
    _f4_cell(ws, f"B{R}",
             "b. Payment of Taxes, cess, fees, charges, premiums, interest etc. "
             "to any statutory Authority.", size=9)
    _f4_cell(ws, f"C{R}", dev_b_est if dev_b_est else "NA",
             bold=True, h="right", size=9,
             number_fmt=_RUPEE if dev_b_est else None)
    _f4_cell(ws, f"D{R}", dev_b_inc if dev_b_inc else "NA",
             bold=True, h="right", size=9,
             number_fmt=_RUPEE_D if dev_b_inc else None)
    R += 1

    # c. Finance cost
    _row_ht(R, 35)
    _f4_cell(ws, f"A{R}", None)
    _f4_cell(ws, f"B{R}",
             "c. Principal sum and interest payable to financial institutions, scheduled banks, "
             "non-banking financial institution (NBFC) or money lenders on construction funding "
             "or money borrowed for construction;", size=9)
    _f4_cell(ws, f"C{R}", dev_c_est if dev_c_est else "NA",
             bold=True, h="right", size=9,
             number_fmt=_RUPEE if dev_c_est else None)
    _f4_cell(ws, f"D{R}", dev_c_inc if dev_c_inc else "NA",
             bold=True, h="right", size=9,
             number_fmt=_RUPEE_D if dev_c_inc else None)
    R += 1

    # Sub-total Development Cost
    _row_ht(R, 22)
    _f4_cell(ws, f"A{R}", None, fill=F4_SUBTOT_FILL)
    _f4_cell(ws, f"B{R}", "Sub-Total of Development Cost",
             bold=True, size=10, fill=F4_SUBTOT_FILL)
    _f4_cell(ws, f"C{R}", dev_sub_est if dev_sub_est else "NA",
             bold=True, h="right", size=10,
             number_fmt=_RUPEE if dev_sub_est else None,
             fill=F4_SUBTOT_FILL)
    _f4_cell(ws, f"D{R}", dev_sub_inc if dev_sub_inc else "NA",
             bold=True, h="right", size=10,
             number_fmt=_RUPEE_D if dev_sub_inc else None,
             fill=F4_SUBTOT_FILL)
    DEV_SUBTOT_ROW = R
    R += 1

    # ─────────────────────────────────────────────────────────────────────────
    # SUMMARY ROWS Sr. 2 – 8
    # ─────────────────────────────────────────────────────────────────────────
    def _summary_row(sr, text, c_val=None, d_val=None,
                     c_fmt=_RUPEE, d_fmt=_RUPEE_D, ht=28, net=False):
        nonlocal R
        _row_ht(R, ht)
        bg = F4_NET_FILL if net else F4_SUMMARY_FILL
        _f4_cell(ws, f"A{R}", sr, bold=net, size=9, h="center", fill=bg)
        _f4_cell(ws, f"B{R}", text, bold=net, size=9, fill=bg)
        # C
        if c_val is not None:
            _f4_cell(ws, f"C{R}", c_val, bold=True, h="right", size=9,
                     number_fmt=c_fmt if isinstance(c_val, (int, float)) else None,
                     fill=bg)
        else:
            _f4_cell(ws, f"C{R}", None, fill=bg)
        # D
        if d_val is not None:
            _f4_cell(ws, f"D{R}", d_val, bold=True, h="right", size=9,
                     number_fmt=d_fmt if isinstance(d_val, (int, float)) else None,
                     fill=bg)
        else:
            _f4_cell(ws, f"D{R}", None, fill=bg)
        R += 1

    _summary_row(2,
                 "Total Estimated Cost of the Real Estate Project "
                 "[1(i) + 1(ii)] of Estimated Column.",
                 c_val=round(total_est, 2), ht=30)

    _summary_row(3,
                 "Total Cost Incurred of the Real Estate Project "
                 "[1(i) + 1(ii)] of Incurred Column.",
                 d_val=round(total_inc, 2), ht=30)

    _summary_row(4,
                 "% Completion of Construction Work "
                 "(as per Project Architect's Certificate)",
                 d_val=arch_pct, d_fmt=_PCT_FMT, ht=30)

    _summary_row(5,
                 f"Proportion of the Cost incurred on Land Cost and "
                 f"{arch_pct*100:.2f}% Construction Cost to the Total Estimated Cost.  ( Sr.3 / Sr.2 %)",
                 d_val=round(proportion, 10), d_fmt=_PCT_FMT, ht=30)

    _summary_row(6,
                 "Amount Which can be Withdrawn from the Designated Account.\n"
                 "(Total Estimated Cost × Proportion of cost incurred  =  Sr.2 × Sr.5)",
                 d_val=round(withdraw_allow, 2), ht=40)

    _summary_row(7,
                 "Less: Total Sale Amount Received from Sold Units "
                 "(as per Annexure A – sum of amounts received from all sold unit allottees).",
                 d_val=round(total_amount_received_sold, 2), ht=40)

    _summary_row(8,
                 "Net Amount which can be Withdrawn from the Designated Bank Account "
                 "under this Certificate.  (Sr.6 – Sr.7)",
                 d_val=round(net_withdraw, 2), ht=35, net=True)

    # ── Signature block 1 ─────────────────────────────────────────────
    _row_ht(R, 55)
    _f4_merge(ws, f"A{R}:D{R}",
              "This certificate is being issued for RERA compliance for the Company "
              f"[{project.get('promoter_name', 'Promoter')}] and is based on the records "
              "and documents produced before me and explanations provided to me by the "
              "management of the Company.",
              size=9, h="left", v="top")
    R += 1

    for lbl in ["Yours Faithfully,",
                f"Signature of Chartered Accountant – {project.get('ca_name', '')}",
                f"(Membership No.: {project.get('ca_membership', '…………')})",
                "______________________",
                "Name"]:
        _row_ht(R, 18)
        _f4_cell(ws, f"A{R}", None); _f4_cell(ws, f"B{R}", None)
        _f4_merge(ws, f"C{R}:D{R}", lbl,
                  bold=(lbl in ["Yours Faithfully,", "Name"]), size=9, h="center")
        R += 1

    # Spacer
    _row_ht(R, 6); ws[f"A{R}"].border = THIN_ALL
    for col in ["B", "C", "D"]:
        ws[f"{col}{R}"].border = THIN_ALL
    R += 1

    # ─────────────────────────────────────────────────────────────────────────
    # ADDITIONAL INFORMATION FOR ONGOING PROJECTS
    # ─────────────────────────────────────────────────────────────────────────
    _row_ht(R, 26)
    _f4_merge(ws, f"A{R}:D{R}",
              "(ADDITIONAL INFORMATION FOR ONGOING PROJECTS)",
              bold=True, size=11, h="center", fill=F4_TITLE_FILL, color="FFFFFF")
    R += 1

    def _add_row(sr, text, c_val=None, d_val=None,
                 c_fmt=_RUPEE, d_fmt=_RUPEE, ht=35):
        nonlocal R
        _row_ht(R, ht)
        _f4_cell(ws, f"A{R}", sr, size=9, h="center")
        _f4_cell(ws, f"B{R}", text, size=9)
        if c_val is not None:
            cv = c_val
            _f4_cell(ws, f"C{R}", cv, bold=isinstance(cv,(int,float)),
                     h="right" if isinstance(cv,(int,float)) else "center",
                     size=9, number_fmt=c_fmt if isinstance(cv,(int,float)) else None)
        else:
            _f4_cell(ws, f"C{R}", None)
        if d_val is not None:
            dv = d_val
            _f4_cell(ws, f"D{R}", dv, bold=isinstance(dv,(int,float)),
                     h="right" if isinstance(dv,(int,float)) else "center",
                     size=9, number_fmt=d_fmt if isinstance(dv,(int,float)) else None)
        else:
            _f4_cell(ws, f"D{R}", None)
        R += 1

    _add_row(1,
             "Estimated Balance Cost to Complete the Real Estate Project\n"
             "(Difference of Total Estimated Project cost less Cost incurred)\n"
             "(calculated as per Form IV)",
             d_val=round(bal_cost, 2), d_fmt=_RUPEE_D, ht=50)

    _add_row(2,
             "Balance amount of receivables from sold apartments as per Annexure A to this "
             "certificate (as certified by Chartered Accountant as verified from the records "
             "and books of Accounts)",
             d_val=round(bal_recv_sold, 2) if bal_recv_sold else "NIL",
             d_fmt=_RUPEE_D if isinstance(bal_recv_sold, (int, float)) and bal_recv_sold else None,
             ht=45)

    _add_row(3,
             "(i) Balance Unsold area (to be certified by Management and to be verified "
             "by CA from the records and books of accounts)",
             d_val=f"{unsold_area:,.2f} sq.m." if unsold_area else "NIL", ht=30)

    _add_row(None,
             f"(ii) Estimated amount of sales proceeds in respect of unsold apartments "
             f"(Average Sale Price per sq.m. of sold units: "
             f"Rs. {format_indian_number(int(avg_sale_price)) if avg_sale_price else '—'} × "
             f"Total Unsold Area: {unsold_area:,.2f} sq.m.) as per Annexure A to this certificate",
             d_val=round(unsold_val, 2) if unsold_val else "NIL",
             d_fmt=_RUPEE_D if unsold_val else None, ht=50)

    _add_row(4, "Estimated receivables of ongoing project.  "
             "Sum of Total Sale Value of Sold Units + 3(ii)\n"
             f"(Total Sale Value of Sold Units: Rs. {format_indian_number(int(total_sale_val_sold)) if total_sale_val_sold else '0'})",
             d_val=round(total_recv, 2) if total_recv else "NIL",
             d_fmt=_RUPEE_D if total_recv else None, ht=40)

    _add_row(5, "Amount to be deposited in Designated Account – 70% or 100%", ht=20)

    _row_ht(R, 30)
    _f4_cell(ws, f"A{R}", None)
    _f4_merge(ws, f"B{R}:C{R}",
              "IF Sr.4 is GREATER THAN Sr.1 : 70% of the balance receivables of ongoing "
              "project will be deposited in Designated Account", size=9)
    _f4_cell(ws, f"D{R}",
             f"70% × ₹{format_indian_number(int(total_recv))} = ₹{format_indian_number(int(total_recv * 0.70))}"
             if total_recv > bal_cost else "—",
             bold=True, h="right", size=9)
    R += 1

    _row_ht(R, 30)
    _f4_cell(ws, f"A{R}", None)
    _f4_merge(ws, f"B{R}:C{R}",
              "IF Sr.4 is LESSER THAN or EQUAL TO Sr.1 : 100% of the balance receivables "
              "of ongoing project will be deposited in Designated Account", size=9)
    _f4_cell(ws, f"D{R}",
             f"100% × ₹{format_indian_number(int(total_recv))} = ₹{format_indian_number(int(total_recv))}"
             if total_recv <= bal_cost else "—",
             bold=True, h="right", size=9)
    R += 1

    # ── Signature block 2 ─────────────────────────────────────────────
    _row_ht(R, 6)
    for col in ["A","B","C","D"]: ws[f"{col}{R}"].border = THIN_ALL
    R += 1

    _row_ht(R, 55)
    _f4_merge(ws, f"A{R}:D{R}",
              "This certificate is being issued for RERA compliance for the Company "
              f"[{project.get('promoter_name', 'Promoter')}] and is based on the records "
              "and documents produced before me and explanations provided to me by the "
              "management of the Company.",
              size=9, h="left", v="top")
    R += 1

    for lbl in ["Yours Faithfully,",
                "Signature of Chartered Accountant,",
                f"(Membership No.: {project.get('ca_membership', '…………')})",
                "______________________",
                "Name"]:
        _row_ht(R, 18)
        _f4_cell(ws, f"A{R}", None); _f4_cell(ws, f"B{R}", None)
        _f4_merge(ws, f"C{R}:D{R}", lbl,
                  bold=(lbl in ["Yours Faithfully,", "Name"]), size=9, h="center")
        R += 1

    # Spacer
    _row_ht(R, 6)
    for col in ["A","B","C","D"]: ws[f"{col}{R}"].border = THIN_ALL
    R += 1

    # ─────────────────────────────────────────────────────────────────────────
    # ANNEXURE A — Receivables from sold inventory
    # ─────────────────────────────────────────────────────────────────────────
    _row_ht(R, 22)
    _f4_merge(ws, f"A{R}:D{R}", "Annexure A",
              bold=True, size=12, h="center", fill=F4_TITLE_FILL, color="FFFFFF")
    R += 1

    _row_ht(R, 18)
    _f4_merge(ws, f"A{R}:D{R}",
              "Statement for calculation of Receivables from the Sales of the Ongoing "
              "Real Estate Project",
              bold=True, size=10, h="center")
    R += 1

    # Sold Inventory section
    _row_ht(R, 18)
    _f4_merge(ws, f"A{R}:D{R}", "Sold Inventory",
              bold=True, size=10, fill=F4_SECTION_FILL)
    R += 1

    # Header
    _row_ht(R, 28)
    for col, hdr in [("A","Sr No"), ("B","Flat No."),
                     ("C","Received Amount (Rs.)"), ("D","Balance Receivable (Rs.)")]:
        _f4_cell(ws, f"{col}{R}", hdr, bold=True, size=9, h="center",
                 fill=F4_SECTION_FILL)
    R += 1

    # Sold rows
    building_map = {b.get("building_id"): b.get("building_name","") for b in (buildings or [])}
    sold_sales = [s for s in (sales or []) if s.get("buyer_name")]
    total_recv_amt  = 0
    total_bal_recv  = 0

    if sold_sales:
        for idx, s in enumerate(sold_sales, 1):
            _row_ht(R, 18)
            bname = building_map.get(s.get("building_id"), s.get("building_name",""))
            unit  = f"{bname} – {s.get('unit_number','')}" if bname else s.get("unit_number","")
            recv  = s.get("amount_received", 0)
            bal   = s.get("sale_value", 0) - recv
            total_recv_amt += recv
            total_bal_recv += bal
            _f4_cell(ws, f"A{R}", idx, h="center", size=9)
            _f4_cell(ws, f"B{R}", unit, size=9)
            _f4_cell(ws, f"C{R}", recv, h="right", size=9, number_fmt=_RUPEE_D)
            _f4_cell(ws, f"D{R}", bal, h="right", size=9, number_fmt=_RUPEE_D)
            R += 1
    else:
        # placeholder rows
        for i in range(1, 5):
            _row_ht(R, 18)
            _f4_cell(ws, f"A{R}", i, h="center", size=9)
            for col in ["B","C","D"]: _f4_cell(ws, f"{col}{R}", None)
            R += 1

    # Total row
    _row_ht(R, 20)
    _f4_cell(ws, f"A{R}", None, fill=F4_SUBTOT_FILL)
    _f4_cell(ws, f"B{R}", "Total", bold=True, size=9, fill=F4_SUBTOT_FILL)
    _f4_cell(ws, f"C{R}", total_recv_amt if total_recv_amt else None,
             bold=True, h="right", size=9,
             number_fmt=_RUPEE_D if total_recv_amt else None,
             fill=F4_SUBTOT_FILL)
    _f4_cell(ws, f"D{R}", total_bal_recv if total_bal_recv else None,
             bold=True, h="right", size=9,
             number_fmt=_RUPEE_D if total_bal_recv else None,
             fill=F4_SUBTOT_FILL)
    R += 1

    # Spacer
    _row_ht(R, 6)
    for col in ["A","B","C","D"]: ws[f"{col}{R}"].border = THIN_ALL
    R += 1

    # Unsold Inventory Valuation
    _row_ht(R, 100)
    _f4_merge(ws, f"A{R}:D{R}",
              f"(Unsold Inventory Valuation)\n"
              f"Average Sale Price per sq.m. of Sold Units: "
              f"Rs. {format_indian_number(int(avg_sale_price)) if avg_sale_price else '—'} per sq.m.\n"
              f"Total Unsold Area: {unsold_area:,.2f} sq.m.\n"
              f"Estimated Unsold Inventory Value (Avg Price × Unsold Area): "
              f"Rs. {format_indian_number(int(unsold_val)) if unsold_val else 'NIL'}",
              size=9, v="top")
    R += 1

    # Unsold unit table header
    _row_ht(R, 28)
    for col, hdr in [("A","Sr No"), ("B","Flat Number"),
                     ("C","Area (sq.m.)"), ("D","Estimated Value (Rs.)")]:
        _f4_cell(ws, f"{col}{R}", hdr, bold=True, size=9, h="center",
                 fill=F4_SECTION_FILL)
    R += 1

    unsold_sales = [s for s in (sales or []) if not s.get("buyer_name")]
    total_unsold_area = 0
    total_unsold_val  = 0
    if unsold_sales:
        for idx, s in enumerate(unsold_sales, 1):
            _row_ht(R, 18)
            bname = building_map.get(s.get("building_id"), s.get("building_name",""))
            unit  = f"{bname} – {s.get('unit_number','')}" if bname else s.get("unit_number","")
            area  = s.get("carpet_area", 0) or 0
            val   = area * avg_sale_price if avg_sale_price else 0
            total_unsold_area += area
            total_unsold_val  += val
            _f4_cell(ws, f"A{R}", idx, h="center", size=9)
            _f4_cell(ws, f"B{R}", unit, size=9)
            _f4_cell(ws, f"C{R}", area, h="right", size=9)
            _f4_cell(ws, f"D{R}", val if val else None, h="right", size=9,
                     number_fmt=_RUPEE_D if val else None)
            R += 1
    else:
        for i in range(1, 5):
            _row_ht(R, 18)
            _f4_cell(ws, f"A{R}", i, h="center", size=9)
            for col in ["B","C","D"]: _f4_cell(ws, f"{col}{R}", None)
            R += 1

    # Unsold units total row
    _row_ht(R, 20)
    _f4_cell(ws, f"A{R}", None, fill=F4_SUBTOT_FILL)
    _f4_cell(ws, f"B{R}", "Total", bold=True, size=9, fill=F4_SUBTOT_FILL)
    _f4_cell(ws, f"C{R}", total_unsold_area if total_unsold_area else None,
             bold=True, h="right", size=9, fill=F4_SUBTOT_FILL)
    _f4_cell(ws, f"D{R}", total_unsold_val if total_unsold_val else None,
             bold=True, h="right", size=9,
             number_fmt=_RUPEE_D if total_unsold_val else None,
             fill=F4_SUBTOT_FILL)
    R += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── ANNEXURE-A ──────────────────────────────────────────────────────────────

def generate_annexure_a_excel(project, sales, buildings, quarter, year):
    wb = Workbook()
    ws = wb.active
    ws.title = "Annexure A - Receivables"

    ws.merge_cells("A1:J1")
    title_cell(ws["A1"], "ANNEXURE - A: Statement of Receivables from Allottees")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"].value = (
        f"Project: {project.get('project_name', '')}  |  "
        f"RERA No: {project.get('rera_number', '')}  |  "
        f"Period: {quarter} {year}  |  Date: {datetime.now().strftime('%d/%m/%Y')}"
    )
    ws["A2"].font = Font(size=10)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    headers = [
        "Sr.", "Unit No.", "Building / Wing", "Type", "Area (sq.m.)",
        "Agreement Value (₹)", "Amount Received (₹)", "Balance Due (₹)",
        "Due Date", "Allottee Name",
    ]
    for c, h in enumerate(headers, 1):
        header_cell(ws.cell(row=4, column=c), h)
    ws.row_dimensions[4].height = 35

    bl = {b.get("building_id"): b.get("building_name", "") for b in (buildings or [])}

    total_val = total_recv = total_bal = 0

    for idx, sale in enumerate(sales or [], 1):
        r = idx + 4
        agr  = sale.get("sale_value", sale.get("agreement_value", 0))
        recv = sale.get("amount_received", 0)
        bal  = agr - recv
        total_val  += agr
        total_recv += recv
        total_bal  += bal

        data_cell(ws.cell(row=r, column=1),  str(idx), align="center")
        data_cell(ws.cell(row=r, column=2),  sale.get("unit_number", ""), align="center")
        data_cell(ws.cell(row=r, column=3),  bl.get(sale.get("building_id"), sale.get("building_name", "")))
        data_cell(ws.cell(row=r, column=4),  sale.get("apartment_classification", sale.get("unit_type", sale.get("flat_type", "NA"))), align="center")
        data_cell(ws.cell(row=r, column=5),  str(sale.get("carpet_area", "")), align="center")
        data_cell(ws.cell(row=r, column=6),  format_indian_number(agr), align="right")
        data_cell(ws.cell(row=r, column=7),  format_indian_number(recv), align="right")
        data_cell(ws.cell(row=r, column=8),  format_indian_number(bal), align="right")
        data_cell(ws.cell(row=r, column=9),  sale.get("agreement_date", sale.get("due_date", "")), align="center")
        data_cell(ws.cell(row=r, column=10), sale.get("buyer_name", sale.get("allottee_name", "")))

    total_row = len(sales or []) + 5
    for c in range(1, 11):
        total_cell(ws.cell(row=total_row, column=c), "")
    total_cell(ws.cell(row=total_row, column=5),  "TOTAL",                          align="center")
    total_cell(ws.cell(row=total_row, column=6),  format_indian_number(total_val))
    total_cell(ws.cell(row=total_row, column=7),  format_indian_number(total_recv))
    total_cell(ws.cell(row=total_row, column=8),  format_indian_number(total_bal))

    # Summary
    sr = total_row + 2
    ws.merge_cells(f"A{sr}:J{sr}")
    ws[f"A{sr}"].value = (
        f"Summary: Total Units: {len(sales or [])}  |  "
        f"Total Agreement Value: ₹{format_indian_number(total_val)}  |  "
        f"Total Received: ₹{format_indian_number(total_recv)}  |  "
        f"Total Balance Due: ₹{format_indian_number(total_bal)}"
    )
    ws[f"A{sr}"].font = Font(bold=True, size=10)
    ws[f"A{sr}"].fill = SECTION_FILL
    ws[f"A{sr}"].alignment = Alignment(horizontal="left", vertical="center")

    col_widths = [6, 10, 16, 12, 12, 22, 22, 20, 14, 26]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
