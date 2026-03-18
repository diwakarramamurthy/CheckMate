"""Sales.Py routes."""

from fastapi import APIRouter, HTTPException, Depends, Query, UploadFile, File, status, Request
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional
import uuid
import json
from io import BytesIO

from database import db
from models import (
    UnitSaleBase, UnitSaleCreate, UnitSaleResponse,
    CommonDevelopmentWorksCreate, CommonDevelopmentWorksResponse,
    FinancialSummaryCreate, FinancialSummaryResponse
)
from auth import get_current_user

router = APIRouter()

@router.post("/unit-sales", response_model=UnitSaleResponse)
async def create_unit_sale(sale: UnitSaleCreate, current_user: dict = Depends(get_current_user)):
    sale_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    balance = sale.sale_value - sale.amount_received
    
    sale_doc = {
        "sale_id": sale_id,
        **sale.model_dump(),
        "balance_receivable": balance,
        "created_at": now
    }
    await db.unit_sales.insert_one(sale_doc)
    return UnitSaleResponse(**{k: v for k, v in sale_doc.items() if k != "_id"})

@router.get("/unit-sales", response_model=List[UnitSaleResponse])
async def get_unit_sales(project_id: str = Query(...), current_user: dict = Depends(get_current_user)):
    sales = await db.unit_sales.find({"project_id": project_id}, {"_id": 0}).to_list(10000)
    return [UnitSaleResponse(**s) for s in sales]

@router.put("/unit-sales/{sale_id}", response_model=UnitSaleResponse)
async def update_unit_sale(sale_id: str, sale: UnitSaleBase, current_user: dict = Depends(get_current_user)):
    existing = await db.unit_sales.find_one({"sale_id": sale_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Unit sale not found")
    
    balance = sale.sale_value - sale.amount_received
    update_data = sale.model_dump()
    update_data["balance_receivable"] = balance
    
    await db.unit_sales.update_one({"sale_id": sale_id}, {"$set": update_data})
    updated = await db.unit_sales.find_one({"sale_id": sale_id}, {"_id": 0})
    return UnitSaleResponse(**updated)

@router.delete("/unit-sales/{sale_id}")
async def delete_unit_sale(sale_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.unit_sales.delete_one({"sale_id": sale_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Unit sale not found")
    return {"message": "Unit sale deleted"}

# COMMON DEVELOPMENT WORKS (FORM-1 Table B)

@router.post("/common-development-works", response_model=CommonDevelopmentWorksResponse)
async def create_common_development_works(works: CommonDevelopmentWorksCreate, current_user: dict = Depends(get_current_user)):
    works_data = works.works.model_dump()
    
    # Calculate overall completion based on proposed works
    total_weight = 0
    weighted_completion = 0
    
    work_items = [
        ("internal_roads_footpaths", "internal_roads_proposed"),
        ("water_supply", "water_supply_proposed"),
        ("sewerage", "sewerage_proposed"),
        ("storm_water_drains", "storm_water_proposed"),
        ("landscaping_trees", "landscaping_proposed"),
        ("street_lighting", "street_lighting_proposed"),
        ("community_buildings", "community_buildings_proposed"),
        ("sewage_treatment", "sewage_treatment_proposed"),
        ("solid_waste_management", "solid_waste_proposed"),
        ("rainwater_harvesting", "rainwater_proposed"),
        ("energy_management", "energy_management_proposed"),
        ("fire_safety", "fire_safety_proposed"),
        ("electrical_infrastructure", "electrical_proposed"),
    ]
    
    for completion_field, proposed_field in work_items:
        if works_data.get(proposed_field, False):
            total_weight += 1
            weighted_completion += works_data.get(completion_field, 0)
    
    overall = (weighted_completion / total_weight) if total_weight > 0 else 0
    
    works_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    works_doc = {
        "works_id": works_id,
        "project_id": works.project_id,
        "quarter": works.quarter,
        "year": works.year,
        "works": works_data,
        "overall_completion": round(overall, 2),
        "created_at": now
    }
    
    await db.common_development_works.update_one(
        {"project_id": works.project_id, "quarter": works.quarter, "year": works.year},
        {"$set": works_doc},
        upsert=True
    )
    
    return CommonDevelopmentWorksResponse(**works_doc)

@router.get("/common-development-works", response_model=List[CommonDevelopmentWorksResponse])
async def get_common_development_works(
    project_id: str = Query(...),
    quarter: Optional[str] = None,
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"project_id": project_id}
    if quarter:
        query["quarter"] = quarter
    if year:
        query["year"] = year
    
    works = await db.common_development_works.find(query, {"_id": 0}).to_list(100)
    return [CommonDevelopmentWorksResponse(**w) for w in works]

@router.get("/common-development-works/latest/{project_id}", response_model=CommonDevelopmentWorksResponse)
async def get_latest_common_development_works(project_id: str, current_user: dict = Depends(get_current_user)):
    works = await db.common_development_works.find_one(
        {"project_id": project_id},
        {"_id": 0},
        sort=[("year", -1), ("quarter", -1)]
    )
    if not works:
        raise HTTPException(status_code=404, detail="No common development works data found")
    return CommonDevelopmentWorksResponse(**works)

# FINANCIAL SUMMARY (FORM-5)

@router.post("/financial-summary", response_model=FinancialSummaryResponse)
async def create_financial_summary(summary: FinancialSummaryCreate, current_user: dict = Depends(get_current_user)):
    # Calculate unsold inventory value
    unsold_value = summary.unsold_area_sqm * summary.asr_rate_per_sqm
    
    # Calculate total estimated receivables
    total_receivables = summary.total_balance_receivable_sold + unsold_value
    
    # Determine deposit percentage and amount (Form-5 logic)
    # Get project cost data
    cost = await db.project_costs.find_one(
        {"project_id": summary.project_id},
        {"_id": 0},
        sort=[("year", -1), ("quarter", -1)]
    )
    
    balance_cost = cost.get("balance_cost", 0) if cost else 0
    
    # If receivables > balance cost, deposit 70%, else 100%
    deposit_pct = 70 if total_receivables > balance_cost else 100
    amount_to_deposit = total_receivables * deposit_pct / 100
    
    summary_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    summary_doc = {
        "summary_id": summary_id,
        **summary.model_dump(),
        "unsold_inventory_value": unsold_value,
        "total_estimated_receivables": total_receivables,
        "deposit_percentage": deposit_pct,
        "amount_to_deposit": amount_to_deposit,
        "created_at": now
    }
    
    await db.financial_summaries.update_one(
        {"project_id": summary.project_id, "quarter": summary.quarter, "year": summary.year},
        {"$set": summary_doc},
        upsert=True
    )
    
    return FinancialSummaryResponse(**summary_doc)

@router.get("/financial-summary", response_model=List[FinancialSummaryResponse])
async def get_financial_summaries(
    project_id: str = Query(...),
    quarter: Optional[str] = None,
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"project_id": project_id}
    if quarter:
        query["quarter"] = quarter
    if year:
        query["year"] = year
    
    summaries = await db.financial_summaries.find(query, {"_id": 0}).to_list(100)
    return [FinancialSummaryResponse(**s) for s in summaries]

@router.get("/financial-summary/latest/{project_id}", response_model=FinancialSummaryResponse)
async def get_latest_financial_summary(project_id: str, current_user: dict = Depends(get_current_user)):
    summary = await db.financial_summaries.find_one(
        {"project_id": project_id},
        {"_id": 0},
        sort=[("year", -1), ("quarter", -1)]
    )
    if not summary:
        raise HTTPException(status_code=404, detail="No financial summary data found")
    return FinancialSummaryResponse(**summary)

@router.post("/unit-sales/bulk", response_model=Dict[str, Any])
async def bulk_create_unit_sales(
    project_id: str = Query(...),
    sales: List[UnitSaleBase] = None,
    current_user: dict = Depends(get_current_user)
):
    if not sales:
        raise HTTPException(status_code=400, detail="No sales data provided")
    
    created = 0
    errors = []
    for idx, sale in enumerate(sales):
        try:
            sale_id = str(uuid.uuid4())
            now = datetime.now(timezone.utc).isoformat()
            balance = sale.sale_value - sale.amount_received
            
            sale_doc = {
                "sale_id": sale_id,
                "project_id": project_id,
                **sale.model_dump(),
                "balance_receivable": balance,
                "created_at": now
            }
            await db.unit_sales.insert_one(sale_doc)
            created += 1
        except Exception as e:
            errors.append({"row": idx + 1, "error": str(e)})
    
    return {"created": created, "errors": errors}

# EXCEL IMPORT

@router.post("/import/sales-excel")
async def import_sales_excel(
    project_id: str = Query(...),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Import unit sales from Excel file - replaces all existing data for the project"""
    import openpyxl
    
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents))
    ws = wb.active
    
    # Find header row
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            headers[str(cell_value).lower().strip()] = col
    
    # Map columns
    column_map = {
        "unit_number": ["unit number", "unit no", "unit no.", "flat no", "flat number"],
        "building_name": ["building name", "building", "wing", "tower"],
        "carpet_area": ["carpet area", "area", "carpet area (sq ft)", "carpet area (sqm)"],
        "sale_value": ["sale value", "agreement value", "total value", "price"],
        "amount_received": ["amount received", "received", "collection", "amount collected"],
        "buyer_name": ["buyer name", "buyer", "customer name", "allottee"],
        "agreement_date": ["agreement date", "date", "booking date"],
        "apartment_classification": ["apartment classification", "apt classification", "apartment type", "flat type", "bhk type", "unit type", "type"]
    }
    
    def find_column(field):
        for alias in column_map.get(field, []):
            if alias in headers:
                return headers[alias]
        return None
    
    # DELETE all existing sales for this project before importing new data
    delete_result = await db.unit_sales.delete_many({"project_id": project_id})
    deleted_count = delete_result.deleted_count
    
    created = 0
    unsold_count = 0
    errors = []
    
    # Get buildings for this project to map building names to IDs
    buildings = await db.buildings.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    building_map = {b["building_name"].lower(): b["building_id"] for b in buildings}
    
    for row in range(2, ws.max_row + 1):
        try:
            unit_col = find_column("unit_number")
            building_col = find_column("building_name")
            area_col = find_column("carpet_area")
            value_col = find_column("sale_value")
            received_col = find_column("amount_received")
            buyer_col = find_column("buyer_name")
            date_col = find_column("agreement_date")
            apt_class_col = find_column("apartment_classification")
            
            unit_number = ws.cell(row=row, column=unit_col).value if unit_col else None
            building_name = ws.cell(row=row, column=building_col).value if building_col else None
            
            if not unit_number:
                continue
            
            carpet_area = float(ws.cell(row=row, column=area_col).value or 0) if area_col else 0
            sale_value = float(ws.cell(row=row, column=value_col).value or 0) if value_col else 0
            amount_received = float(ws.cell(row=row, column=received_col).value or 0) if received_col else 0
            buyer_name = str(ws.cell(row=row, column=buyer_col).value or "").strip() if buyer_col else ""
            agreement_date = ws.cell(row=row, column=date_col).value if date_col else None
            apartment_classification = str(ws.cell(row=row, column=apt_class_col).value or "NA").strip() if apt_class_col else "NA"
            
            if agreement_date and hasattr(agreement_date, 'isoformat'):
                agreement_date = agreement_date.isoformat()
            elif agreement_date:
                agreement_date = str(agreement_date)
            
            # Find building_id
            building_id = building_map.get(str(building_name).lower(), "") if building_name else ""
            
            sale_id = str(uuid.uuid4())
            now = datetime.now(timezone.utc).isoformat()
            balance = sale_value - amount_received
            
            # Determine sale status: if buyer_name is blank, it's unsold inventory
            is_sold = bool(buyer_name)
            status = "sold" if is_sold else "unsold"
            
            if not is_sold:
                unsold_count += 1
            
            sale_doc = {
                "sale_id": sale_id,
                "project_id": project_id,
                "unit_number": str(unit_number),
                "building_id": building_id,
                "building_name": str(building_name or ""),
                "carpet_area": carpet_area,
                "sale_value": sale_value,
                "amount_received": amount_received,
                "buyer_name": buyer_name,
                "agreement_date": agreement_date,
                "apartment_classification": apartment_classification,
                "balance_receivable": balance,
                "status": status,  # "sold" or "unsold"
                "created_at": now,
                "updated_at": now
            }
            await db.unit_sales.insert_one(sale_doc)
            created += 1
        except Exception as e:
            errors.append({"row": row, "error": str(e)})
    
    return {
        "deleted": deleted_count,
        "created": created,
        "sold_units": created - unsold_count,
        "unsold_units": unsold_count,
        "errors": errors,
        "total_rows": ws.max_row - 1,
        "message": f"Replaced {deleted_count} existing records with {created} new records ({unsold_count} unsold units)"
    }

@router.get("/import/sales-template")
async def get_sales_template():
    """Download Excel template for sales import"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import tempfile
    import os
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Data"
    
    headers = ["Unit Number", "Building Name", "Carpet Area", "Sale Value", "Amount Received", "Buyer Name", "Agreement Date", "Apartment Classification"]
    header_fill = PatternFill(start_color="172554", end_color="172554", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18
    
    # Add sample data
    sample_data = [
        ["A-101", "Tower A", 850, 7500000, 5000000, "John Doe", "2024-01-15", "2 BHK"],
        ["A-102", "Tower A", 920, 8200000, 4100000, "Jane Smith", "2024-02-20", "3 BHK"],
    ]
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to temp file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return FileResponse(
        path=temp_file.name,
        filename="CheckMate - Sales Data.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# DASHBOARD & SUMMARY


