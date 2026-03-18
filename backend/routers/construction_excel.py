"""
Construction Progress Excel Import / Export helpers.
Provides:
  - generate_construction_excel_template()   → BytesIO of .xlsx
  - parse_construction_excel()               → (tower_activities, infra_activities, meta)
"""

from io import BytesIO
from typing import Optional, Dict, Any, Tuple

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── colour palette (matches CheckMate blue theme) ──────────────────────────
BLUE_DARK   = "1E3A5F"
BLUE_MID    = "2563EB"
BLUE_LIGHT  = "DBEAFE"
BLUE_HEADER = "1D4ED8"
GREEN_FILL  = "DCFCE7"
YELLOW_FILL = "FEF9C3"
GREY_FILL   = "F1F5F9"
WHITE       = "FFFFFF"
ORANGE_FILL = "FFF7ED"

def _border(style="thin"):
    s = Side(style=style, color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _cell_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


# ── The full tower construction template definition ─────────────────────────
TOWER_TEMPLATE = [
    {
        "id": "basement_slab_completion",
        "name": "Completion of Basement Slabs (Below Plinth)",
        "activities": [
            {"id": "reinforcement_lintel_roof",  "name": "Reinforcement up to Lintel/Roof bottom", "weightage": 0},
            {"id": "shuttering_for_column",       "name": "Shuttering for Column",                  "weightage": 0},
            {"id": "concreting_for_column",        "name": "Concreting for Column",                   "weightage": 0},
            {"id": "shuttering_beams_roof",        "name": "Shuttering for Beams and Roof",           "weightage": 0},
            {"id": "reinforcement_beams_roof",     "name": "Reinforcement for Beams and Roof",        "weightage": 0},
            {"id": "concreting_beams_roof",        "name": "Concreting for Beams and Roof",           "weightage": 0},
            {"id": "dismantling_roof_shuttering",  "name": "Dismantling of Roof Shuttering",          "weightage": 0},
        ],
    },
    {
        "id": "plinth_completion",
        "name": "Completion of Plinth",
        "activities": [
            {"id": "excavation",                   "name": "Excavation",                              "weightage": 0},
            {"id": "pcc_below_footing",            "name": "PCC below footing",                       "weightage": 0},
            {"id": "shuttering_for_footing",       "name": "Shuttering for Footing",                  "weightage": 0},
            {"id": "reinforcement_footing_column", "name": "Reinforcement for Footing and Column",    "weightage": 0},
            {"id": "concreting_for_footing",       "name": "Concreting for Footing",                  "weightage": 0},
            {"id": "shuttering_column_to_plinth",  "name": "Shuttering for Column up to Plinth",      "weightage": 0},
            {"id": "concreting_for_column",        "name": "Concreting for Column",                   "weightage": 0},
            {"id": "shuttering_plinth_beam",       "name": "Shuttering for Plinth Beam",              "weightage": 0},
            {"id": "reinforcement_plinth_beam",    "name": "Reinforcement for Plinth Beam",           "weightage": 0},
            {"id": "concreting_plinth_beam",       "name": "Concreting for Plinth Beam",              "weightage": 0},
            {"id": "filling_earth_plinth_pcc",     "name": "Filling earth within Plinth and PCC",     "weightage": 0},
        ],
    },
    {
        "id": "slab_completion",
        "name": "Completion of Slabs at all levels",
        "activities": [
            {"id": "reinforcement_lintel_roof",  "name": "Reinforcement up to Lintel/Roof bottom", "weightage": 0},
            {"id": "shuttering_for_column",       "name": "Shuttering for Column",                  "weightage": 0},
            {"id": "concreting_for_column",        "name": "Concreting for Column",                   "weightage": 0},
            {"id": "shuttering_beams_roof",        "name": "Shuttering for Beams and Roof",           "weightage": 0},
            {"id": "reinforcement_beams_roof",     "name": "Reinforcement for Beams and Roof",        "weightage": 0},
            {"id": "concreting_beams_roof",        "name": "Concreting for Beams and Roof",           "weightage": 0},
            {"id": "dismantling_roof_shuttering",  "name": "Dismantling of Roof Shuttering",          "weightage": 0},
        ],
    },
    {
        "id": "brickwork_plastering",
        "name": "Completion of Brickwork and Plastering",
        "activities": [
            {"id": "brickwork_external_walls",      "name": "Brickwork External walls",                    "weightage": 0},
            {"id": "brickwork_internal_walls",      "name": "Brickwork Internal walls",                    "weightage": 0},
            {"id": "fixing_door_window_frames",     "name": "Fixing of Door/Window frames",                "weightage": 0},
            {"id": "fixing_concealed_pipes",        "name": "Fixing concealed Water & Electric pipes",     "weightage": 0},
            {"id": "plastering_external_walls",     "name": "Plastering External walls",                   "weightage": 0},
            {"id": "plastering_internal_walls",     "name": "Plastering Internal walls",                   "weightage": 0},
            {"id": "waterproof_plastering_toilets", "name": "Water-proof Plastering of Toilets",           "weightage": 0},
        ],
    },
    {
        "id": "plumbing",
        "name": "Plumbing",
        "activities": [
            {"id": "fixing_water_pipes",     "name": "Fixing External & Internal Water Pipes",           "weightage": 0},
            {"id": "fixing_wc_pipes_traps",  "name": "Fixing External & Internal WC Pipes & Traps",      "weightage": 0},
            {"id": "fixing_plumbing_fixtures","name": "Fixing of all Plumbing fixtures",                 "weightage": 0},
        ],
    },
    {
        "id": "electrical_works",
        "name": "Electrical Works",
        "activities": [
            {"id": "laying_all_cables",               "name": "Laying all Cables (Internal)",             "weightage": 0},
            {"id": "fixing_electrical_fixtures",      "name": "Fixing all Electrical fixtures (Internal)","weightage": 0},
            {"id": "electrical_breaker_box",          "name": "Electrical/Breaker Box (External)",         "weightage": 0},
            {"id": "electric_meter_box",              "name": "Electric meter box (External)",             "weightage": 0},
            {"id": "connecting_cable_electrical_box", "name": "Connecting cable to the Electrical box",   "weightage": 0},
        ],
    },
    {
        "id": "window_works",
        "name": "Aluminium/UPVC Window",
        "activities": [
            {"id": "fixing_frames", "name": "Fixing of frames", "weightage": 0},
            {"id": "fixing_glass",  "name": "Fixing of Glass",  "weightage": 0},
        ],
    },
    {
        "id": "tiling_flooring",
        "name": "Tiling/Flooring",
        "activities": [
            {"id": "laying_floor_tiles",                   "name": "Laying of Floor tiles",                              "weightage": 0},
            {"id": "laying_wall_tiles_kitchen_bathroom",   "name": "Laying of Wall tiles Kitchen & Bathroom",            "weightage": 0},
            {"id": "laying_granite_kitchen_counter",       "name": "Laying of Granite/Kadapa slab for Kitchen Counter",  "weightage": 0},
        ],
    },
    {
        "id": "door_shutter_fixing",
        "name": "Door Shutter Fixing",
        "activities": [
            {"id": "fixing_door_shutters",  "name": "Fixing of Door shutters",             "weightage": 0},
            {"id": "fixing_locks_handles",  "name": "Fixing of locks, handles & accessories","weightage": 0},
        ],
    },
    {
        "id": "water_proofing",
        "name": "Water Proofing",
        "activities": [
            {"id": "terrace_roof_waterproofing", "name": "Terrace roof water proofing", "weightage": 0},
        ],
    },
    {
        "id": "painting",
        "name": "Painting",
        "activities": [
            {"id": "painting_ceiling",       "name": "Ceiling",         "weightage": 0},
            {"id": "painting_walls",         "name": "Walls",           "weightage": 0},
            {"id": "painting_grills",        "name": "Grills",          "weightage": 0},
            {"id": "painting_doors_windows", "name": "Doors/Windows",   "weightage": 0},
        ],
    },
    {
        "id": "carpark",
        "name": "Carpark",
        "activities": [
            {"id": "levelling", "name": "Levelling", "weightage": 0},
            {"id": "paving",    "name": "Paving",    "weightage": 0},
        ],
    },
    {
        "id": "handover_intimation",
        "name": "Intimation of Handover",
        "activities": [
            {"id": "intimation_of_handover", "name": "Intimation of Handover", "weightage": 0},
        ],
    },
]

INFRA_TEMPLATE = [
    {"id": "road_footpath_storm_drain",       "name": "Road, Foot-path and storm water drain",                    "weightage": 0},
    {"id": "underground_sewage_network",      "name": "Underground sewage drainage network",                      "weightage": 0},
    {"id": "sewage_treatment_plant",          "name": "Sewage Treatment Plant",                                   "weightage": 0},
    {"id": "overhead_sump_reservoir",         "name": "Over-head and Sump water reservoir/Tank",                  "weightage": 0},
    {"id": "underground_water_distribution",  "name": "Under ground water distribution network",                  "weightage": 0},
    {"id": "electric_substation_cables",      "name": "Electric Substation & Under-ground electric cables",       "weightage": 0},
    {"id": "street_lights",                   "name": "Street Lights",                                            "weightage": 0},
    {"id": "entry_gate",                      "name": "Entry Gate",                                               "weightage": 0},
    {"id": "boundary_wall",                   "name": "Boundary wall",                                            "weightage": 0},
    {"id": "club_house",                      "name": "Club House",                                               "weightage": 0},
    {"id": "swimming_pool",                   "name": "Swimming Pool",                                            "weightage": 0},
    {"id": "amphitheatre",                    "name": "Amphitheatre",                                             "weightage": 0},
    {"id": "gardens_playground",              "name": "Gardens / Play Ground",                                    "weightage": 0},
]


# ───────────────────────────────────────────────────────────────────────────
#  GENERATOR
# ───────────────────────────────────────────────────────────────────────────

def generate_construction_excel_template(
    building_id: str,
    building_name: str,
    project_name: str,
    quarter: str,
    year: int,
    number_of_floors: int = 1,
    existing_tower_data: Optional[Dict] = None,
    existing_infra_data: Optional[Dict] = None,
) -> BytesIO:
    """
    Generate a pre-filled Excel template for construction progress entry.
    Returns a BytesIO object ready to stream.
    """
    wb = Workbook()

    # ── Sheet 1: Instructions / Meta ───────────────────────────────────────
    ws_meta = wb.active
    ws_meta.title = "Meta"
    _build_meta_sheet(ws_meta, building_id, building_name, project_name, quarter, year, number_of_floors)

    # ── Sheet 2: Tower Construction ────────────────────────────────────────
    ws_tower = wb.create_sheet("Tower Construction")
    _build_tower_sheet(ws_tower, existing_tower_data or {})

    # ── Sheet 3: Infrastructure Works ──────────────────────────────────────
    ws_infra = wb.create_sheet("Infrastructure Works")
    _build_infra_sheet(ws_infra, existing_infra_data or {})

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_meta_sheet(ws, building_id, building_name, project_name, quarter, year, number_of_floors):
    """Build the Meta/Instructions sheet."""
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40

    # Title
    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = "CheckMate — Construction Progress Import Template"
    c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill = _hdr_fill(BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Warning
    ws.merge_cells("A2:B2")
    c = ws["A2"]
    c.value = "⚠  Fill in COMPLETION % and APPLICABLE columns only. Do not rename sheets or delete rows."
    c.font = Font(name="Calibri", bold=True, size=10, color="92400E")
    c.fill = _hdr_fill(YELLOW_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 6  # spacer

    meta_rows = [
        ("Project",           project_name),
        ("Building",          building_name),
        ("Quarter",           quarter),
        ("Year",              year),
        ("Number of Floors",  number_of_floors),
        ("Building ID",       building_id),
    ]

    for i, (label, value) in enumerate(meta_rows, start=4):
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=value)
        lc.font = Font(name="Calibri", bold=True, size=11)
        lc.fill = _hdr_fill(GREY_FILL)
        lc.border = _border()
        lc.alignment = Alignment(vertical="center")
        vc.font = Font(name="Calibri", size=11)
        vc.border = _border()
        vc.alignment = Alignment(vertical="center")
        ws.row_dimensions[i].height = 18

    # Spacer
    ws.row_dimensions[10].height = 6

    # Instructions
    instructions = [
        ("Step 1:", "Select the correct Building, Quarter and Year in the app before importing."),
        ("Step 2:", "Go to the 'Tower Construction' sheet. Enter Completion % (0–100) and Applicable (Yes/No)."),
        ("Step 3:", "Go to the 'Infrastructure Works' sheet. Fill the same columns."),
        ("Step 4:", "Save the file and upload it via the Import Excel button in the app."),
        ("Note:",   "The system will automatically recalculate overall completion based on weightages."),
        ("Note:",   "Marking an activity as 'No' removes it from the calculation (weightage redistributed)."),
    ]

    for i, (step, text) in enumerate(instructions, start=11):
        sc = ws.cell(row=i, column=1, value=step)
        tc = ws.cell(row=i, column=2, value=text)
        sc.font = Font(name="Calibri", bold=True, size=10, color=BLUE_MID)
        sc.alignment = Alignment(vertical="center")
        tc.font = Font(name="Calibri", size=10)
        tc.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[i].height = 16


def _build_tower_sheet(ws, existing_data: Dict):
    """Build the Tower Construction sheet."""
    # Column widths
    col_widths = {"A": 38, "B": 42, "C": 14, "D": 18, "E": 20, "F": 22, "G": 20}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Title row
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "Tower Construction Progress"
    c.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c.fill = _hdr_fill(BLUE_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Note row
    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = "✏  Enter COMPLETION % (0-100) and APPLICABLE (Yes/No). Do not edit other columns."
    c.font = Font(name="Calibri", bold=True, size=10, color="92400E")
    c.fill = _hdr_fill(YELLOW_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Header row
    headers = ["Category", "Activity", "Weightage (%)", "COMPLETION %\n(0 - 100)", "APPLICABLE\n(Yes / No)", "Activity ID\n(do not edit)", "Category ID\n(do not edit)"]
    col_fills = [BLUE_LIGHT, BLUE_LIGHT, GREY_FILL, GREEN_FILL, GREEN_FILL, GREY_FILL, GREY_FILL]
    for col_idx, (header, fill_color) in enumerate(zip(headers, col_fills), start=1):
        c = ws.cell(row=3, column=col_idx, value=header)
        c.font = Font(name="Calibri", bold=True, size=10, color="1E293B")
        c.fill = _hdr_fill(fill_color)
        c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # Data validation for COMPLETION % column (D)
    dv_completion = DataValidation(
        type="decimal",
        operator="between",
        formula1=0,
        formula2=100,
        showErrorMessage=True,
        errorTitle="Invalid Value",
        error="Please enter a number between 0 and 100",
        prompt="Enter a number between 0 and 100",
        promptTitle="Completion %"
    )
    ws.add_data_validation(dv_completion)

    # Data validation for APPLICABLE column (E)
    dv_applicable = DataValidation(
        type="list",
        formula1='"Yes,No"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid Value",
        error='Please select "Yes" or "No"',
    )
    ws.add_data_validation(dv_applicable)

    row = 4
    for cat in TOWER_TEMPLATE:
        cat_id = cat["id"]
        cat_data = existing_data.get(cat_id, {})
        cat_first_row = row  # for potential merge

        for act_idx, activity in enumerate(cat["activities"]):
            act_id = activity["id"]
            act_data = cat_data.get(act_id, {})

            completion = act_data.get("completion", 0)
            is_applicable = act_data.get("is_applicable", True)
            applicable_str = "Yes" if is_applicable else "No"

            # Category cell (only first activity in category gets the label)
            cat_cell = ws.cell(row=row, column=1)
            if act_idx == 0:
                cat_cell.value = cat["name"]
                cat_cell.font = Font(name="Calibri", bold=True, size=10)
            cat_cell.fill = _hdr_fill(BLUE_LIGHT)
            cat_cell.border = _border()
            cat_cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Activity name
            c = ws.cell(row=row, column=2, value=activity["name"])
            c.font = Font(name="Calibri", size=10)
            c.border = _border()
            c.fill = _cell_fill(WHITE)
            c.alignment = Alignment(vertical="center", wrap_text=True)

            # Weightage (read-only visually)
            c = ws.cell(row=row, column=3, value=activity["weightage"])
            c.font = Font(name="Calibri", size=10, color="475569")
            c.border = _border()
            c.fill = _hdr_fill(GREY_FILL)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "0.00"

            # Completion % (editable)
            c = ws.cell(row=row, column=4, value=completion)
            c.font = Font(name="Calibri", bold=True, size=11, color="166534")
            c.border = _border()
            c.fill = _hdr_fill(GREEN_FILL)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "0.00"
            dv_completion.add(c)

            # Applicable (editable)
            c = ws.cell(row=row, column=5, value=applicable_str)
            c.font = Font(name="Calibri", bold=True, size=10, color="1E40AF")
            c.border = _border()
            c.fill = _hdr_fill(BLUE_LIGHT)
            c.alignment = Alignment(horizontal="center", vertical="center")
            dv_applicable.add(c)

            # Activity ID (reference)
            c = ws.cell(row=row, column=6, value=act_id)
            c.font = Font(name="Calibri", size=9, color="94A3B8")
            c.border = _border()
            c.fill = _hdr_fill(GREY_FILL)
            c.alignment = Alignment(horizontal="center", vertical="center")

            # Category ID (reference)
            c = ws.cell(row=row, column=7, value=cat_id)
            c.font = Font(name="Calibri", size=9, color="94A3B8")
            c.border = _border()
            c.fill = _hdr_fill(GREY_FILL)
            c.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[row].height = 16
            row += 1

    # Freeze panes at row 4
    ws.freeze_panes = "A4"


def _build_infra_sheet(ws, existing_data: Dict):
    """Build the Infrastructure Works sheet."""
    col_widths = {"A": 50, "B": 14, "C": 18, "D": 20, "E": 28}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Title
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Infrastructure Works Progress"
    c.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    c.fill = _hdr_fill(BLUE_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Note
    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = "✏  Enter COMPLETION % (0-100) and APPLICABLE (Yes/No). Do not edit other columns."
    c.font = Font(name="Calibri", bold=True, size=10, color="92400E")
    c.fill = _hdr_fill(YELLOW_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Headers
    headers = ["Activity", "Weightage (%)", "COMPLETION %\n(0 - 100)", "APPLICABLE\n(Yes / No)", "Activity ID\n(do not edit)"]
    fills   = [BLUE_LIGHT, GREY_FILL, GREEN_FILL, GREEN_FILL, GREY_FILL]
    for col_idx, (hdr, fill_color) in enumerate(zip(headers, fills), start=1):
        c = ws.cell(row=3, column=col_idx, value=hdr)
        c.font = Font(name="Calibri", bold=True, size=10, color="1E293B")
        c.fill = _hdr_fill(fill_color)
        c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    dv_completion = DataValidation(
        type="decimal", operator="between", formula1=0, formula2=100,
        showErrorMessage=True, errorTitle="Invalid Value",
        error="Please enter a number between 0 and 100",
    )
    ws.add_data_validation(dv_completion)

    dv_applicable = DataValidation(
        type="list", formula1='"Yes,No"', showDropDown=False,
        showErrorMessage=True, errorTitle="Invalid Value",
        error='Please select "Yes" or "No"',
    )
    ws.add_data_validation(dv_applicable)

    for row_idx, item in enumerate(INFRA_TEMPLATE, start=4):
        act_id = item["id"]
        act_data = existing_data.get(act_id, {})
        completion = act_data.get("completion", 0)
        is_applicable = act_data.get("is_applicable", True)
        applicable_str = "Yes" if is_applicable else "No"

        # Activity name
        c = ws.cell(row=row_idx, column=1, value=item["name"])
        c.font = Font(name="Calibri", size=10)
        c.border = _border()
        c.fill = _cell_fill(WHITE)
        c.alignment = Alignment(vertical="center", wrap_text=True)

        # Weightage
        c = ws.cell(row=row_idx, column=2, value=item["weightage"])
        c.font = Font(name="Calibri", size=10, color="475569")
        c.border = _border()
        c.fill = _hdr_fill(GREY_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0.0"

        # Completion %
        c = ws.cell(row=row_idx, column=3, value=completion)
        c.font = Font(name="Calibri", bold=True, size=11, color="166534")
        c.border = _border()
        c.fill = _hdr_fill(GREEN_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = "0.00"
        dv_completion.add(c)

        # Applicable
        c = ws.cell(row=row_idx, column=4, value=applicable_str)
        c.font = Font(name="Calibri", bold=True, size=10, color="1E40AF")
        c.border = _border()
        c.fill = _hdr_fill(BLUE_LIGHT)
        c.alignment = Alignment(horizontal="center", vertical="center")
        dv_applicable.add(c)

        # Activity ID
        c = ws.cell(row=row_idx, column=5, value=act_id)
        c.font = Font(name="Calibri", size=9, color="94A3B8")
        c.border = _border()
        c.fill = _hdr_fill(GREY_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = 18

    ws.freeze_panes = "A4"


# ───────────────────────────────────────────────────────────────────────────
#  PARSER
# ───────────────────────────────────────────────────────────────────────────

def parse_construction_excel(file_bytes: bytes) -> Tuple[Dict, Dict]:
    """
    Parse a single-building construction progress Excel file.
    Returns:
      tower_activities : dict  — same structure expected by /construction-progress/detailed
      infra_activities : dict  — same structure expected by /infrastructure-progress
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    # For single-building file the sheet is named "Tower Construction"
    tower_activities = _parse_tower_sheet(wb, sheet_name="Tower Construction")
    infra_activities = _parse_infra_sheet(wb)
    return tower_activities, infra_activities


def _parse_infra_sheet(wb: Workbook) -> Dict:
    """Parse Infrastructure Works sheet."""
    result = {}

    if "Infrastructure Works" not in wb.sheetnames:
        return result

    ws = wb["Infrastructure Works"]

    # A=Activity, B=Weightage, C=Completion%, D=Applicable, E=ActivityID
    COMPLETION_COL = 3   # C
    APPLICABLE_COL = 4   # D
    ACT_ID_COL     = 5   # E

    for row in ws.iter_rows(min_row=4):
        act_id_cell = row[ACT_ID_COL - 1]
        comp_cell   = row[COMPLETION_COL - 1]
        appl_cell   = row[APPLICABLE_COL - 1]

        act_id = str(act_id_cell.value).strip() if act_id_cell.value else ""
        if not act_id or act_id == "None":
            continue

        try:
            completion = float(comp_cell.value) if comp_cell.value is not None else 0.0
            completion = max(0.0, min(100.0, completion))
        except (TypeError, ValueError):
            completion = 0.0

        applicable_raw = str(appl_cell.value).strip().lower() if appl_cell.value else "yes"
        is_applicable  = applicable_raw not in ("no", "false", "0", "n")

        result[act_id] = {
            "completion":    completion,
            "is_applicable": is_applicable,
        }

    return result


# ═══════════════════════════════════════════════════════════════════════════
#  BULK  GENERATOR  &  PARSER
#  Generates one workbook covering ALL buildings in a project.
#  Layout:
#    Sheet "Meta"              — project info + building index with Copy-From column
#    Sheet "Infrastructure"    — project-level infra (one per project)
#    Sheet "{building_name}"   — one sheet per building (tower/villa/rowhouse/etc.)
# ═══════════════════════════════════════════════════════════════════════════

# Pastel palette for alternating building sheet tabs
_TAB_COLOURS = [
    "1D4ED8", "0F766E", "7C3AED", "B45309", "BE185D",
    "047857", "1E40AF", "6D28D9", "92400E", "9D174D",
]


def generate_bulk_construction_excel_template(
    project_id: str,
    project_name: str,
    quarter: str,
    year: int,
    buildings: list,           # list of dicts: {building_id, building_name, building_type, residential_floors}
    existing_tower_map: dict,  # {building_id: tower_activities_dict}
    existing_infra_data: dict, # infra activities dict (project-level)
) -> BytesIO:
    """
    Generate a bulk Excel template covering all buildings in the project.
    Each building gets its own sheet.  The Meta sheet has a 'Copy From Building'
    column — if filled, the importer copies that building's data to this one.
    """
    wb = Workbook()

    # ── Sheet 1: Meta ───────────────────────────────────────────────────────
    ws_meta = wb.active
    ws_meta.title = "Meta"
    _build_bulk_meta_sheet(ws_meta, project_id, project_name, quarter, year, buildings)

    # ── Sheet 2: Infrastructure ─────────────────────────────────────────────
    ws_infra = wb.create_sheet("Infrastructure Works")
    _build_infra_sheet(ws_infra, existing_infra_data or {})

    # ── One sheet per building ──────────────────────────────────────────────
    for idx, bldg in enumerate(buildings):
        bname     = bldg.get("building_name", f"Building {idx+1}")
        bid       = bldg.get("building_id", "")
        btype     = bldg.get("building_type", "residential_tower")
        existing  = existing_tower_map.get(bid, {})

        # Sanitise sheet name (Excel max 31 chars, no special chars)
        sheet_name = _safe_sheet_name(bname, idx)
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = _TAB_COLOURS[idx % len(_TAB_COLOURS)]
        _build_bulk_building_sheet(ws, bname, bid, btype, existing)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _safe_sheet_name(name: str, idx: int) -> str:
    """Sanitise a sheet name to be valid in Excel (max 31 chars, no : / \\ ? * [ ])."""
    import re
    clean = re.sub(r"[:/\\?*\[\]]", "", name).strip()
    if not clean:
        clean = f"Building_{idx+1}"
    return clean[:31]


def _build_bulk_meta_sheet(ws, project_id, project_name, quarter, year, buildings):
    """Meta sheet for the bulk workbook."""
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 30

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "CheckMate — Bulk Construction Progress Import Template"
    c.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    c.fill = _hdr_fill(BLUE_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Warning
    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = (
        "⚠  Fill COMPLETION % and APPLICABLE columns in each building sheet. "
        "Use 'Copy From Building' to replicate data from a similar building automatically."
    )
    c.font = Font(name="Calibri", bold=True, size=10, color="92400E")
    c.fill = _hdr_fill(YELLOW_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    # Project info rows
    info = [
        ("Project", project_name),
        ("Quarter", quarter),
        ("Year",    year),
    ]
    for i, (label, value) in enumerate(info, start=3):
        lc = ws.cell(row=i, column=2, value=label)
        vc = ws.cell(row=i, column=3, value=value)
        lc.font = Font(name="Calibri", bold=True, size=11)
        lc.fill = _hdr_fill(GREY_FILL)
        lc.border = _border()
        vc.font = Font(name="Calibri", size=11)
        vc.border = _border()
        ws.row_dimensions[i].height = 18

    # Spacer
    ws.row_dimensions[6].height = 8

    # Instructions about Copy From
    ws.merge_cells("A7:G7")
    c = ws["A7"]
    c.value = (
        "💡  Copy From Building:  Leave blank to enter data manually in that building's sheet. "
        "Or write the exact name of another building (e.g. 'Tower A') to auto-copy its data on import — "
        "useful when multiple towers/villas are identical."
    )
    c.font = Font(name="Calibri", size=10, color="1E40AF")
    c.fill = _hdr_fill(BLUE_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[7].height = 40

    ws.row_dimensions[8].height = 8

    # Building index header
    hdr_row = 9
    hdr_labels = ["#", "Building Name", "Building Type", "Floors", "Sheet Tab", "Building ID", "Copy From Building\n(leave blank or write building name)"]
    hdr_fills  = [BLUE_HEADER]*6 + [ORANGE_FILL]
    hdr_fonts  = [Font(name="Calibri", bold=True, size=10, color=WHITE)]*6 + [Font(name="Calibri", bold=True, size=10, color="92400E")]
    for col_idx, (lbl, fill, fnt) in enumerate(zip(hdr_labels, hdr_fills, hdr_fonts), start=1):
        c = ws.cell(row=hdr_row, column=col_idx, value=lbl)
        c.font = fnt
        c.fill = _hdr_fill(fill) if col_idx < 7 else _hdr_fill(ORANGE_FILL)
        c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[hdr_row].height = 32

    # Building rows
    building_names = [_safe_sheet_name(b.get("building_name", ""), i) for i, b in enumerate(buildings)]
    dv_copy = DataValidation(
        type="list",
        formula1='"' + ",".join(building_names) + '"',
        showDropDown=False,
        showErrorMessage=False,
    )
    ws.add_data_validation(dv_copy)

    for i, bldg in enumerate(buildings, start=1):
        row = hdr_row + i
        bname  = bldg.get("building_name", f"Building {i}")
        btype  = bldg.get("building_type", "residential_tower").replace("_", " ").title()
        floors = bldg.get("residential_floors", 0)
        bid    = bldg.get("building_id", "")
        sheet  = _safe_sheet_name(bname, i - 1)

        fill = GREEN_FILL if i % 2 == 0 else WHITE
        for col, val in enumerate([str(i), bname, btype, floors, sheet, bid], start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Calibri", size=10)
            c.fill = _hdr_fill(fill)
            c.border = _border()
            c.alignment = Alignment(vertical="center", wrap_text=(col == 2))

        # Copy From cell — editable, orange highlight
        cf_cell = ws.cell(row=row, column=7, value="")
        cf_cell.font = Font(name="Calibri", bold=True, size=10, color="92400E")
        cf_cell.fill = _hdr_fill(ORANGE_FILL)
        cf_cell.border = _border()
        cf_cell.alignment = Alignment(vertical="center")
        dv_copy.add(cf_cell)

        ws.row_dimensions[row].height = 18

    # Instructions at the bottom
    footer_row = hdr_row + len(buildings) + 2
    ws.merge_cells(f"A{footer_row}:G{footer_row}")
    c = ws.cell(row=footer_row, column=1,
                value="Steps: 1) Fill 'Copy From' if applicable  →  2) Open each building's sheet tab  →  3) Enter Completion % and Applicable  →  4) Save and upload via 'Bulk Import Excel'")
    c.font = Font(name="Calibri", size=10, color="475569", italic=True)
    c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[footer_row].height = 20


def _build_bulk_building_sheet(ws, building_name, building_id, building_type, existing_data):
    """
    Single building sheet in the bulk workbook.
    Same columns as the single-building template but with building name in the title.
    """
    btype_label = building_type.replace("_", " ").title() if building_type else "Building"

    # Column widths (same as single template)
    col_widths = {"A": 38, "B": 42, "C": 14, "D": 18, "E": 20, "F": 22, "G": 20}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"{building_name}  ·  {btype_label}  —  Tower Construction Progress"
    c.font = Font(name="Calibri", bold=True, size=12, color=WHITE)
    c.fill = _hdr_fill(BLUE_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    # Building ID reference
    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = f"Building ID: {building_id}  |  ✏  Edit COMPLETION % (col D) and APPLICABLE (col E) only. Do not rename or reorder rows."
    c.font = Font(name="Calibri", bold=True, size=10, color="92400E")
    c.fill = _hdr_fill(YELLOW_FILL)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 18

    # Column headers
    headers = ["Category", "Activity", "Weightage (%)", "COMPLETION %\n(0 - 100)", "APPLICABLE\n(Yes / No)", "Activity ID\n(do not edit)", "Category ID\n(do not edit)"]
    col_fills = [BLUE_LIGHT, BLUE_LIGHT, GREY_FILL, GREEN_FILL, GREEN_FILL, GREY_FILL, GREY_FILL]
    for col_idx, (header, fill_color) in enumerate(zip(headers, col_fills), start=1):
        c = ws.cell(row=3, column=col_idx, value=header)
        c.font = Font(name="Calibri", bold=True, size=10, color="1E293B")
        c.fill = _hdr_fill(fill_color)
        c.border = _border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 30

    # Validations
    dv_completion = DataValidation(
        type="decimal", operator="between", formula1=0, formula2=100,
        showErrorMessage=True, errorTitle="Invalid Value",
        error="Please enter a number between 0 and 100",
    )
    ws.add_data_validation(dv_completion)

    dv_applicable = DataValidation(
        type="list", formula1='"Yes,No"', showDropDown=False,
        showErrorMessage=True, errorTitle="Invalid Value",
        error='Please select "Yes" or "No"',
    )
    ws.add_data_validation(dv_applicable)

    # Activity rows (identical to single-building helper)
    row = 4
    for cat in TOWER_TEMPLATE:
        cat_id   = cat["id"]
        cat_data = existing_data.get(cat_id, {})

        for act_idx, activity in enumerate(cat["activities"]):
            act_id        = activity["id"]
            act_data      = cat_data.get(act_id, {})
            completion    = act_data.get("completion", 0)
            is_applicable = act_data.get("is_applicable", True)
            applicable_str = "Yes" if is_applicable else "No"

            cat_cell = ws.cell(row=row, column=1)
            if act_idx == 0:
                cat_cell.value = cat["name"]
                cat_cell.font = Font(name="Calibri", bold=True, size=10)
            cat_cell.fill = _hdr_fill(BLUE_LIGHT)
            cat_cell.border = _border()
            cat_cell.alignment = Alignment(vertical="center", wrap_text=True)

            c = ws.cell(row=row, column=2, value=activity["name"])
            c.font = Font(name="Calibri", size=10)
            c.border = _border()
            c.fill = _cell_fill(WHITE)
            c.alignment = Alignment(vertical="center", wrap_text=True)

            c = ws.cell(row=row, column=3, value=activity["weightage"])
            c.font = Font(name="Calibri", size=10, color="475569")
            c.border = _border()
            c.fill = _hdr_fill(GREY_FILL)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "0.00"

            c = ws.cell(row=row, column=4, value=completion)
            c.font = Font(name="Calibri", bold=True, size=11, color="166534")
            c.border = _border()
            c.fill = _hdr_fill(GREEN_FILL)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.number_format = "0.00"
            dv_completion.add(c)

            c = ws.cell(row=row, column=5, value=applicable_str)
            c.font = Font(name="Calibri", bold=True, size=10, color="1E40AF")
            c.border = _border()
            c.fill = _hdr_fill(BLUE_LIGHT)
            c.alignment = Alignment(horizontal="center", vertical="center")
            dv_applicable.add(c)

            c = ws.cell(row=row, column=6, value=act_id)
            c.font = Font(name="Calibri", size=9, color="94A3B8")
            c.border = _border()
            c.fill = _hdr_fill(GREY_FILL)
            c.alignment = Alignment(horizontal="center", vertical="center")

            c = ws.cell(row=row, column=7, value=cat_id)
            c.font = Font(name="Calibri", size=9, color="94A3B8")
            c.border = _border()
            c.fill = _hdr_fill(GREY_FILL)
            c.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[row].height = 16
            row += 1

    ws.freeze_panes = "A4"


# ───────────────────────────────────────────────────────────────────────────
#  BULK PARSER
# ───────────────────────────────────────────────────────────────────────────

def parse_bulk_construction_excel(file_bytes: bytes, buildings: list) -> Tuple[Dict, Dict]:
    """
    Parse a bulk construction progress Excel file.

    Args:
        file_bytes  : raw bytes of the uploaded .xlsx
        buildings   : list of dicts {building_id, building_name} from DB
                      (used to match sheet names → building IDs)

    Returns:
        tower_map   : {building_id: tower_activities_dict}
        infra_data  : infra activities dict (from "Infrastructure Works" sheet)
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)

    # Build a map: safe_sheet_name → building_id
    sheet_to_id = {}
    name_to_id  = {}   # building_name → building_id (for copy-from resolution)
    for i, bldg in enumerate(buildings):
        bid   = bldg.get("building_id", "")
        bname = bldg.get("building_name", "")
        sheet = _safe_sheet_name(bname, i)
        sheet_to_id[sheet] = bid
        name_to_id[bname]  = bid
        # Also map the safe name back to building_name for copy-from
        sheet_to_id[bname] = bid

    # ── Read Copy-From map from Meta sheet ──────────────────────────────────
    copy_from_map = {}   # building_id → source_building_id
    if "Meta" in wb.sheetnames:
        ws_meta = wb["Meta"]
        # Building table starts at row 10 (hdr row 9 + data from row 10)
        for row in ws_meta.iter_rows(min_row=10):
            if len(row) < 7:
                continue
            bname_cell   = row[1]  # col B
            copy_from_cell = row[6]  # col G
            if not bname_cell.value:
                continue
            bname_val    = str(bname_cell.value).strip()
            copy_from_val = str(copy_from_cell.value).strip() if copy_from_cell.value else ""
            if bname_val in name_to_id and copy_from_val:
                copy_from_map[name_to_id[bname_val]] = copy_from_val  # store name, resolve later

    # ── Parse each building sheet ────────────────────────────────────────────
    tower_map = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in ("Meta", "Infrastructure Works"):
            continue
        bid = sheet_to_id.get(sheet_name)
        if not bid:
            continue
        ws = wb[sheet_name]
        tower_map[bid] = _parse_tower_sheet(wb, sheet_name=sheet_name)

    # ── Apply Copy-From ───────────────────────────────────────────────────────
    # Resolve: copy_from_map has {target_bid: source_name}
    # Source name may be a building name or a sheet name
    for target_bid, source_name in copy_from_map.items():
        source_bid = name_to_id.get(source_name) or sheet_to_id.get(source_name)
        if source_bid and source_bid in tower_map:
            import copy as _copy
            tower_map[target_bid] = _copy.deepcopy(tower_map[source_bid])

    # ── Parse infrastructure sheet ───────────────────────────────────────────
    infra_data = _parse_infra_sheet(wb)

    return tower_map, infra_data


def _parse_tower_sheet(wb: Workbook, sheet_name: str = "Tower Construction") -> Dict:
    """
    Parse a tower construction sheet (works for both single and bulk workbooks).
    Column layout: A=Category, B=Activity, C=Weightage, D=Completion%, E=Applicable,
                   F=ActivityID, G=CategoryID
    """
    result = {}

    if sheet_name not in wb.sheetnames:
        return result

    ws = wb[sheet_name]

    COMPLETION_COL = 4
    APPLICABLE_COL = 5
    ACT_ID_COL     = 6
    CAT_ID_COL     = 7

    for row in ws.iter_rows(min_row=4):
        if len(row) < 7:
            continue
        cat_id_cell  = row[CAT_ID_COL - 1]
        act_id_cell  = row[ACT_ID_COL - 1]
        comp_cell    = row[COMPLETION_COL - 1]
        appl_cell    = row[APPLICABLE_COL - 1]

        cat_id  = str(cat_id_cell.value).strip()  if cat_id_cell.value  else ""
        act_id  = str(act_id_cell.value).strip()  if act_id_cell.value  else ""
        if not cat_id or not act_id or cat_id == "None" or act_id == "None":
            continue

        try:
            completion = float(comp_cell.value) if comp_cell.value is not None else 0.0
            completion = max(0.0, min(100.0, completion))
        except (TypeError, ValueError):
            completion = 0.0

        applicable_raw = str(appl_cell.value).strip().lower() if appl_cell.value else "yes"
        is_applicable  = applicable_raw not in ("no", "false", "0", "n")

        if cat_id not in result:
            result[cat_id] = {}
        result[cat_id][act_id] = {
            "completion":    completion,
            "is_applicable": is_applicable,
        }

    return result
