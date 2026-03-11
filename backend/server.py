from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Query, status, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import StreamingResponse, FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
import json
from io import BytesIO

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Settings
JWT_SECRET = os.environ.get('JWT_SECRET', 'rera-compliance-secret-key-2024')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Create the main app
app = FastAPI(title="CheckMate - RERA Manager API")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# =========================
# MODELS
# =========================

class UserBase(BaseModel):
    email: EmailStr
    name: str
    role: str = Field(..., description="One of: admin, developer, architect, engineer, ca, auditor")
    phone: Optional[str] = None
    license_number: Optional[str] = None
    firm_name: Optional[str] = None

class UserCreate(UserBase):
    password: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(UserBase):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    created_at: str
    is_active: bool = True

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

class ProjectBase(BaseModel):
    project_name: str
    state: str = "GOA"
    rera_number: str
    promoter_name: str
    promoter_address: str
    project_address: str
    survey_number: Optional[str] = None
    plot_number: Optional[str] = None
    chalta_number: Optional[str] = None  # PTS/Chalta No for RERA forms
    village: Optional[str] = None
    taluka: Optional[str] = None
    district: Optional[str] = None
    ward: Optional[str] = None  # Ward number
    municipality: Optional[str] = None  # Municipality name
    pin_code: Optional[str] = None
    plot_area: Optional[float] = None
    total_built_up_area: Optional[float] = None
    
    # Project boundaries (for RERA forms)
    boundary_north: Optional[str] = None
    boundary_south: Optional[str] = None
    boundary_east: Optional[str] = None
    boundary_west: Optional[str] = None
    
    # RERA Registration details
    rera_registration_date: Optional[str] = None
    rera_validity_date: Optional[str] = None
    project_phase: Optional[str] = None  # Phase 1, Phase 2, etc.
    
    # Project timeline
    project_start_date: Optional[str] = None
    expected_completion_date: Optional[str] = None
    
    # Designated Bank Account (for RERA compliance)
    designated_bank_name: Optional[str] = None
    designated_account_number: Optional[str] = None
    designated_ifsc_code: Optional[str] = None
    
    # Professional details - Architect
    architect_name: Optional[str] = None
    architect_license: Optional[str] = None
    architect_address: Optional[str] = None
    architect_contact: Optional[str] = None
    architect_email: Optional[str] = None
    
    # Professional details - Engineer
    engineer_name: Optional[str] = None
    engineer_license: Optional[str] = None
    engineer_address: Optional[str] = None
    engineer_contact: Optional[str] = None
    engineer_email: Optional[str] = None
    
    # Professional details - Structural Consultant
    structural_consultant_name: Optional[str] = None
    structural_consultant_license: Optional[str] = None
    
    # Professional details - MEP Consultant
    mep_consultant_name: Optional[str] = None
    mep_consultant_license: Optional[str] = None
    
    # Professional details - Site Supervisor
    site_supervisor_name: Optional[str] = None
    
    # Professional details - Quantity Surveyor
    quantity_surveyor_name: Optional[str] = None
    
    # Professional details - CA
    ca_name: Optional[str] = None
    ca_firm_name: Optional[str] = None
    ca_membership_number: Optional[str] = None
    ca_address: Optional[str] = None
    ca_contact: Optional[str] = None
    ca_email: Optional[str] = None
    
    # Professional details - Auditor
    auditor_name: Optional[str] = None
    auditor_firm_name: Optional[str] = None
    auditor_membership_number: Optional[str] = None
    auditor_address: Optional[str] = None
    auditor_contact: Optional[str] = None
    auditor_email: Optional[str] = None
    
    # Planning Authority
    planning_authority_name: Optional[str] = None

class ProjectCreate(ProjectBase):
    pass

class ProjectResponse(ProjectBase):
    model_config = ConfigDict(extra="ignore")
    project_id: str
    created_by: str
    created_at: str
    updated_at: str

class ParkingFloors(BaseModel):
    basement: int = 0
    stilt_ground: int = 0
    upper_level: int = 0

class BuildingBase(BaseModel):
    building_name: str
    building_type: str = "residential_tower"  # residential_tower, mixed_tower, row_house, bungalow
    
    # Parking floors configuration
    parking_basement: int = 0
    parking_stilt_ground: int = 0
    parking_upper_level: int = 0
    
    # Floor configuration for Towers
    commercial_floors: int = 0  # Only for mixed_tower
    residential_floors: int = 0
    apartments_per_floor: int = 0  # Only for towers
    
    # Legacy fields (calculated)
    floors: int = 0  # Total floors (auto-calculated)
    units: int = 0   # Total units (auto-calculated)
    
    estimated_cost: float = 0
    
    # Completion details (for Form-2)
    completion_cert_number: Optional[str] = None
    completion_cert_date: Optional[str] = None
    occupancy_cert_number: Optional[str] = None
    occupancy_cert_date: Optional[str] = None
    planning_authority: Optional[str] = None
    structural_consultant: Optional[str] = None
    mep_consultant: Optional[str] = None
    site_supervisor: Optional[str] = None

class BuildingCreate(BuildingBase):
    project_id: str

class BuildingBulkCreate(BaseModel):
    project_id: str
    building_names: List[str]  # List of building names to create
    template: BuildingBase  # Template with all configuration

class BuildingResponse(BuildingBase):
    model_config = ConfigDict(extra="ignore")
    building_id: str
    project_id: str
    created_at: str
    total_parking_floors: int = 0

# Comprehensive Construction Progress - Building Works (Based on detailed tracking sheet)
# Each activity has: completion (0-100), is_applicable (True/False for N/A), base_weightage

class ActivityItem(BaseModel):
    completion: float = 0  # 0-100%
    is_applicable: bool = True  # False = N/A, will recalibrate weightage
    base_weightage: float = 0  # Original weightage before recalibration

# a) Completion of Plinth - 10.89%
class PlinthCompletion(BaseModel):
    excavation: ActivityItem = ActivityItem(base_weightage=0.90)
    pcc_below_footing: ActivityItem = ActivityItem(base_weightage=0.80)
    shuttering_for_footing: ActivityItem = ActivityItem(base_weightage=0.90)
    reinforcement_footing_column: ActivityItem = ActivityItem(base_weightage=1.70)
    concreting_for_footing: ActivityItem = ActivityItem(base_weightage=1.50)
    shuttering_column_to_plinth: ActivityItem = ActivityItem(base_weightage=0.50)
    concreting_for_column: ActivityItem = ActivityItem(base_weightage=1.79)
    shuttering_plinth_beam: ActivityItem = ActivityItem(base_weightage=0.70)
    reinforcement_plinth_beam: ActivityItem = ActivityItem(base_weightage=1.00)
    concreting_plinth_beam: ActivityItem = ActivityItem(base_weightage=0.80)
    filling_earth_plinth_pcc: ActivityItem = ActivityItem(base_weightage=0.30)

# b) Completion of Slabs at all levels - 31.78% (divided by number of floors)
class SlabCompletion(BaseModel):
    reinforcement_lintel_roof: ActivityItem = ActivityItem(base_weightage=6.00)
    shuttering_for_column: ActivityItem = ActivityItem(base_weightage=4.40)
    concreting_for_column: ActivityItem = ActivityItem(base_weightage=5.00)
    shuttering_beams_roof: ActivityItem = ActivityItem(base_weightage=4.80)
    reinforcement_beams_roof: ActivityItem = ActivityItem(base_weightage=6.00)
    concreting_beams_roof: ActivityItem = ActivityItem(base_weightage=4.00)
    dismantling_roof_shuttering: ActivityItem = ActivityItem(base_weightage=1.58)

# c) Completion of Brickwork and Plastering - 12.71%
class BrickworkPlastering(BaseModel):
    brickwork_external_walls: ActivityItem = ActivityItem(base_weightage=2.50)
    brickwork_internal_walls: ActivityItem = ActivityItem(base_weightage=2.80)
    fixing_door_window_frames: ActivityItem = ActivityItem(base_weightage=1.50)
    fixing_concealed_pipes: ActivityItem = ActivityItem(base_weightage=1.40)
    plastering_external_walls: ActivityItem = ActivityItem(base_weightage=1.30)
    plastering_internal_walls: ActivityItem = ActivityItem(base_weightage=1.80)
    waterproof_plastering_toilets: ActivityItem = ActivityItem(base_weightage=1.41)

# d) Plumbing - 3.93%
class Plumbing(BaseModel):
    fixing_water_pipes: ActivityItem = ActivityItem(base_weightage=1.40)
    fixing_wc_pipes_traps: ActivityItem = ActivityItem(base_weightage=1.30)
    fixing_plumbing_fixtures: ActivityItem = ActivityItem(base_weightage=1.23)

# e) Electrical works - 9.08%
class ElectricalWorks(BaseModel):
    laying_all_cables: ActivityItem = ActivityItem(base_weightage=3.00)
    fixing_electrical_fixtures: ActivityItem = ActivityItem(base_weightage=2.00)
    electrical_breaker_box: ActivityItem = ActivityItem(base_weightage=2.00)
    electric_meter_box: ActivityItem = ActivityItem(base_weightage=1.00)
    connecting_cable_electrical_box: ActivityItem = ActivityItem(base_weightage=1.08)

# f) Aluminium/UPVC window - 8.02%
class WindowWorks(BaseModel):
    fixing_frames: ActivityItem = ActivityItem(base_weightage=4.40)
    fixing_glass: ActivityItem = ActivityItem(base_weightage=3.62)

# g) Tiling/flooring - 8.02%
class TilingFlooring(BaseModel):
    laying_floor_tiles: ActivityItem = ActivityItem(base_weightage=3.00)
    laying_wall_tiles_kitchen_bathroom: ActivityItem = ActivityItem(base_weightage=2.80)
    laying_granite_kitchen_counter: ActivityItem = ActivityItem(base_weightage=2.22)

# h) Door shutter fixing - 2.42%
class DoorShutterFixing(BaseModel):
    fixing_door_shutters: ActivityItem = ActivityItem(base_weightage=1.30)
    fixing_locks_handles: ActivityItem = ActivityItem(base_weightage=1.12)

# i) Water Proofing - 2.42%
class WaterProofing(BaseModel):
    terrace_roof_waterproofing: ActivityItem = ActivityItem(base_weightage=2.42)

# j) Painting - 8.32%
class Painting(BaseModel):
    painting_ceiling: ActivityItem = ActivityItem(base_weightage=2.30)
    painting_walls: ActivityItem = ActivityItem(base_weightage=2.80)
    painting_grills: ActivityItem = ActivityItem(base_weightage=1.80)
    painting_doors_windows: ActivityItem = ActivityItem(base_weightage=1.42)

# k) Carpark - 0.76%
class Carpark(BaseModel):
    levelling: ActivityItem = ActivityItem(base_weightage=0.20)
    paving: ActivityItem = ActivityItem(base_weightage=0.56)

# l) Intimation of Handover - 1.66%
class HandoverIntimation(BaseModel):
    intimation_of_handover: ActivityItem = ActivityItem(base_weightage=1.66)

# Complete Tower/Building Construction Progress
class TowerConstructionProgress(BaseModel):
    plinth_completion: PlinthCompletion = PlinthCompletion()
    slab_completion: SlabCompletion = SlabCompletion()
    brickwork_plastering: BrickworkPlastering = BrickworkPlastering()
    plumbing: Plumbing = Plumbing()
    electrical_works: ElectricalWorks = ElectricalWorks()
    window_works: WindowWorks = WindowWorks()
    tiling_flooring: TilingFlooring = TilingFlooring()
    door_shutter_fixing: DoorShutterFixing = DoorShutterFixing()
    water_proofing: WaterProofing = WaterProofing()
    painting: Painting = Painting()
    carpark: Carpark = Carpark()
    handover_intimation: HandoverIntimation = HandoverIntimation()

# Project Infrastructure Works - 100% total
class InfrastructureActivityItem(BaseModel):
    completion: float = 0
    is_applicable: bool = True
    base_weightage: float = 0

class ProjectInfrastructureWorks(BaseModel):
    road_footpath_storm_drain: InfrastructureActivityItem = InfrastructureActivityItem(base_weightage=21.5)
    underground_sewage_network: InfrastructureActivityItem = InfrastructureActivityItem(base_weightage=13.0)
    sewage_treatment_plant: InfrastructureActivityItem = InfrastructureActivityItem(base_weightage=8.5)
    overhead_sump_reservoir: InfrastructureActivityItem = InfrastructureActivityItem(base_weightage=8.5)
    underground_water_distribution: InfrastructureActivityItem = InfrastructureActivityItem(base_weightage=10.5)
    electric_substation_cables: InfrastructureActivityItem = InfrastructureActivityItem(base_weightage=10.5)
    street_lights: InfrastructureActivityItem = InfrastructureActivityItem(base_weightage=4.0)
    entry_gate: InfrastructureActivityItem = InfrastructureActivityItem(base_weightage=2.5)
    boundary_wall: InfrastructureActivityItem = InfrastructureActivityItem(base_weightage=6.0)
    club_house: InfrastructureActivityItem = InfrastructureActivityItem(base_weightage=7.0)
    swimming_pool: InfrastructureActivityItem = InfrastructureActivityItem(base_weightage=3.5)
    amphitheatre: InfrastructureActivityItem = InfrastructureActivityItem(base_weightage=2.5)
    gardens_playground: InfrastructureActivityItem = InfrastructureActivityItem(base_weightage=2.0)

# Legacy/Simple activity model for backward compatibility
class ConstructionActivityBase(BaseModel):
    activity_name: str
    weightage: float
    completion_percentage: float = 0
    is_applicable: bool = True  # For N/A support

class ConstructionProgressBase(BaseModel):
    building_id: str
    quarter: str
    year: int
    # Legacy activities list
    activities: List[ConstructionActivityBase] = []
    overall_completion: float = 0
    # NEW: Detailed tower construction progress
    tower_progress: Optional[TowerConstructionProgress] = None
    # Number of floors for slab weightage calculation
    number_of_floors: int = 1
    # Recalibrated weightages after N/A items removed
    recalibrated_total_weightage: float = 100.0

class ConstructionProgressCreate(ConstructionProgressBase):
    pass

class ConstructionProgressResponse(ConstructionProgressBase):
    model_config = ConfigDict(extra="ignore")
    progress_id: str
    project_id: str
    created_at: str
    # Calculated fields
    tower_completion_percentage: float = 0
    category_completions: Dict[str, float] = {}
    # Detailed tower activities data
    tower_activities: Optional[Dict] = None

# Project-level Infrastructure Works Progress
class InfrastructureProgressBase(BaseModel):
    project_id: str
    quarter: str
    year: int
    infrastructure_works: ProjectInfrastructureWorks = ProjectInfrastructureWorks()
    overall_completion: float = 0
    recalibrated_total_weightage: float = 100.0

class InfrastructureProgressCreate(InfrastructureProgressBase):
    pass

class InfrastructureProgressResponse(InfrastructureProgressBase):
    model_config = ConfigDict(extra="ignore")
    progress_id: str
    created_at: str

# Common Development Works (backward compatibility for existing code)
class CommonDevelopmentWorksBase(BaseModel):
    project_id: str
    quarter: str
    year: int
    works: Dict[str, Any] = {}
    overall_completion: float = 0

class CommonDevelopmentWorksCreate(CommonDevelopmentWorksBase):
    pass

class CommonDevelopmentWorksResponse(CommonDevelopmentWorksBase):
    model_config = ConfigDict(extra="ignore")
    works_id: str
    created_at: str

# Infrastructure Cost Estimates (for Buildings & Infrastructure page)
class InfrastructureCostItem(BaseModel):
    estimated_cost: float = 0
    is_applicable: bool = True

class InfrastructureCostBase(BaseModel):
    project_id: str
    road_footpath_storm_drain: InfrastructureCostItem = InfrastructureCostItem()
    underground_sewage_network: InfrastructureCostItem = InfrastructureCostItem()
    sewage_treatment_plant: InfrastructureCostItem = InfrastructureCostItem()
    overhead_sump_reservoir: InfrastructureCostItem = InfrastructureCostItem()
    underground_water_distribution: InfrastructureCostItem = InfrastructureCostItem()
    electric_substation_cables: InfrastructureCostItem = InfrastructureCostItem()
    street_lights: InfrastructureCostItem = InfrastructureCostItem()
    entry_gate: InfrastructureCostItem = InfrastructureCostItem()
    boundary_wall: InfrastructureCostItem = InfrastructureCostItem()
    club_house: InfrastructureCostItem = InfrastructureCostItem()
    swimming_pool: InfrastructureCostItem = InfrastructureCostItem()
    amphitheatre: InfrastructureCostItem = InfrastructureCostItem()
    gardens_playground: InfrastructureCostItem = InfrastructureCostItem()

class InfrastructureCostCreate(InfrastructureCostBase):
    pass

class InfrastructureCostResponse(InfrastructureCostBase):
    model_config = ConfigDict(extra="ignore")
    cost_id: str
    total_infrastructure_cost: float = 0
    created_at: str

# Estimated Development Cost (Fixed values set at project start)
class EstimatedDevelopmentCostBase(BaseModel):
    project_id: str
    # Buildings cost - auto-calculated from sum of all buildings
    buildings_cost: float = 0
    # Infrastructure cost - from infrastructure cost section
    infrastructure_cost: float = 0
    # Consultants fee - manual input
    consultants_fee: float = 0
    # Cost of Machineries - manual input
    machinery_cost: float = 0

class EstimatedDevelopmentCostCreate(EstimatedDevelopmentCostBase):
    pass

class EstimatedDevelopmentCostResponse(EstimatedDevelopmentCostBase):
    model_config = ConfigDict(extra="ignore")
    estimate_id: str
    total_estimated_development_cost: float = 0
    created_at: str
    updated_at: Optional[str] = None

# FORM-4: Detailed Project Cost structure matching CA Certificate format
class ProjectCostBase(BaseModel):
    project_id: str
    quarter: str
    year: int
    
    # LAND COST (Section 1.i in Form-4)
    # a. Land acquisition cost
    land_acquisition_cost: float = 0
    land_acquisition_estimated: float = 0
    land_legal_cost: float = 0
    land_interest_cost: float = 0
    
    # b. Development rights premium (FAR, additional FAR, fungible area)
    development_rights_premium: float = 0
    development_rights_estimated: float = 0
    
    # c. TDR cost
    tdr_cost: float = 0
    tdr_estimated: float = 0
    
    # d. Stamp duty, transfer charges, registration fees
    stamp_duty: float = 0
    stamp_duty_estimated: float = 0
    
    # e. Government charges (to any statutory authority)
    government_charges: float = 0
    government_charges_estimated: float = 0
    
    # f. Land premium (for redevelopment projects)
    land_premium_redevelopment: float = 0
    land_premium_estimated: float = 0
    
    # g. Rehabilitation scheme costs (if applicable)
    rehab_construction_cost: float = 0
    rehab_construction_estimated: float = 0
    rehab_transit_accommodation: float = 0
    rehab_clearance_cost: float = 0
    rehab_asr_premium: float = 0
    
    # DEVELOPMENT COST / CONSTRUCTION COST (Section 1.ii in Form-4)
    # a. Construction cost (as certified by Engineer)
    construction_cost_estimated: float = 0
    construction_cost_actual: float = 0
    
    # On-site expenditure
    onsite_salaries: float = 0
    onsite_consultants_fees: float = 0
    onsite_site_overheads: float = 0
    onsite_services_cost: float = 0  # Water, electricity, sewerage, drainage, roads
    onsite_machinery_equipment: float = 0
    onsite_consumables: float = 0
    
    # Off-site expenditure
    offsite_expenditure: float = 0
    
    # b. Taxes, cess, fees to statutory authority
    taxes_statutory: float = 0
    taxes_statutory_estimated: float = 0
    
    # c. Finance cost (loans, interest)
    finance_cost: float = 0
    finance_cost_estimated: float = 0
    
    # Extra/Additional items not in original estimate (Annexure A)
    extra_items_cost: float = 0
    extra_items_details: Optional[str] = None
    
    # Legacy fields for backward compatibility
    infrastructure_cost: float = 0
    equipment_cost: float = 0
    encumbrance_removal: float = 0
    estimated_land_cost: float = 0
    estimated_development_cost: float = 0

class ProjectCostCreate(ProjectCostBase):
    pass

class ProjectCostResponse(ProjectCostBase):
    model_config = ConfigDict(extra="ignore")
    cost_id: str
    total_land_cost: float = 0
    total_land_cost_estimated: float = 0
    total_development_cost: float = 0
    total_development_cost_estimated: float = 0
    total_estimated_cost: float = 0
    total_cost_incurred: float = 0
    balance_cost: float = 0
    cost_completion_percentage: float = 0
    created_at: str

# FORM-3: Building-wise cost tracking (Engineer's Certificate)
class BuildingCostBase(BaseModel):
    building_id: str
    quarter: str
    year: int
    estimated_cost: float = 0
    cost_incurred: float = 0
    extra_items_cost: float = 0  # Annexure A items
    extra_items_details: Optional[str] = None

class BuildingCostCreate(BuildingCostBase):
    pass

class BuildingCostResponse(BuildingCostBase):
    model_config = ConfigDict(extra="ignore")
    building_cost_id: str
    project_id: str
    completion_percentage: float = 0
    balance_cost: float = 0
    created_at: str

# FORM-5 & Annexure A: Sales and Receivables tracking
class UnitSaleBase(BaseModel):
    unit_number: str
    building_id: str
    building_name: str
    carpet_area: float
    sale_value: float
    amount_received: float = 0
    buyer_name: Optional[str] = None
    agreement_date: Optional[str] = None
    allotment_letter_date: Optional[str] = None  # For Annexure A
    is_sold: bool = True  # False for unsold inventory

class UnitSaleCreate(UnitSaleBase):
    project_id: str

class UnitSaleResponse(UnitSaleBase):
    model_config = ConfigDict(extra="ignore")
    sale_id: str
    project_id: str
    balance_receivable: float = 0
    created_at: str

# FORM-5: Financial Summary for ongoing projects
class FinancialSummaryBase(BaseModel):
    project_id: str
    quarter: str
    year: int
    
    # Designated Account tracking
    designated_account_opening_balance: float = 0
    amount_deposited_this_quarter: float = 0
    amount_withdrawn_this_quarter: float = 0
    designated_account_closing_balance: float = 0
    
    # Total withdrawals
    total_amount_withdrawn_till_date: float = 0
    
    # Receivables calculation
    total_balance_receivable_sold: float = 0  # From sold apartments
    unsold_area_sqm: float = 0
    asr_rate_per_sqm: float = 0  # Ready Reckoner / ASR rate
    unsold_inventory_value: float = 0  # Calculated: unsold_area * asr_rate
    total_estimated_receivables: float = 0  # sold receivables + unsold value
    
    # Deposit calculation (70% or 100%)
    deposit_percentage: int = 70  # 70 or 100
    amount_to_deposit: float = 0
    
    # For FORM-6 Annual report
    amount_collected_this_year: float = 0
    amount_collected_till_date: float = 0
    amount_withdrawn_this_year: float = 0

class FinancialSummaryCreate(FinancialSummaryBase):
    pass

class FinancialSummaryResponse(FinancialSummaryBase):
    model_config = ConfigDict(extra="ignore")
    summary_id: str
    created_at: str

class QuarterlyReportBase(BaseModel):
    project_id: str
    quarter: str
    year: int
    report_date: str
    report_status: str = "draft"

class QuarterlyReportCreate(QuarterlyReportBase):
    pass

class QuarterlyReportResponse(QuarterlyReportBase):
    model_config = ConfigDict(extra="ignore")
    report_id: str
    created_by: str
    created_at: str

class ReportTemplateBase(BaseModel):
    state: str
    report_name: str
    report_type: str
    template_html: str
    data_mapping: Dict[str, Any] = {}

class ReportTemplateCreate(ReportTemplateBase):
    pass

class ReportTemplateResponse(ReportTemplateBase):
    model_config = ConfigDict(extra="ignore")
    template_id: str
    created_at: str
    updated_at: str

class DashboardSummary(BaseModel):
    project_completion_percentage: float = 0
    total_estimated_cost: float = 0
    cost_incurred: float = 0
    balance_cost: float = 0
    total_sales_value: float = 0
    amount_collected: float = 0
    receivables: float = 0
    unsold_inventory_value: float = 0
    rera_deposit_required: float = 0
    total_units: int = 0
    units_sold: int = 0
    # Inventory reconciliation fields
    building_config_units: int = 0       # Units derived from building configuration
    sales_data_units: int = 0            # Units derived from sales data (sold + unsold)
    inventory_mismatch: bool = False     # True when the two totals differ
    inventory_mismatch_delta: int = 0    # sales_data_units − building_config_units

# =========================
# AUTHENTICATION HELPERS
# =========================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid token")
        user = await db.users.find_one({"user_id": user_id}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# =========================
# AUTH ROUTES
# =========================

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user: UserCreate):
    existing = await db.users.find_one({"email": user.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "user_id": user_id,
        "email": user.email,
        "name": user.name,
        "role": user.role,
        "phone": user.phone,
        "license_number": user.license_number,
        "firm_name": user.firm_name,
        "password_hash": hash_password(user.password),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    
    token = create_token(user_id, user.email, user.role)
    user_response = UserResponse(
        user_id=user_id,
        email=user.email,
        name=user.name,
        role=user.role,
        phone=user.phone,
        license_number=user.license_number,
        firm_name=user.firm_name,
        created_at=user_doc["created_at"],
        is_active=True
    )
    return TokenResponse(access_token=token, user=user_response)

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user["user_id"], user["email"], user["role"])
    user_response = UserResponse(
        user_id=user["user_id"],
        email=user["email"],
        name=user["name"],
        role=user["role"],
        phone=user.get("phone"),
        license_number=user.get("license_number"),
        firm_name=user.get("firm_name"),
        created_at=user["created_at"],
        is_active=user.get("is_active", True)
    )
    return TokenResponse(access_token=token, user=user_response)

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(
        user_id=current_user["user_id"],
        email=current_user["email"],
        name=current_user["name"],
        role=current_user["role"],
        phone=current_user.get("phone"),
        license_number=current_user.get("license_number"),
        firm_name=current_user.get("firm_name"),
        created_at=current_user["created_at"],
        is_active=current_user.get("is_active", True)
    )

# =========================
# PROJECT ROUTES
# =========================

@api_router.post("/projects", response_model=ProjectResponse)
async def create_project(project: ProjectCreate, current_user: dict = Depends(get_current_user)):
    project_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    project_doc = {
        "project_id": project_id,
        **project.model_dump(),
        "created_by": current_user["user_id"],
        "created_at": now,
        "updated_at": now
    }
    await db.projects.insert_one(project_doc)
    return ProjectResponse(**{k: v for k, v in project_doc.items() if k != "_id"})

@api_router.get("/projects", response_model=List[ProjectResponse])
async def get_projects(current_user: dict = Depends(get_current_user)):
    projects = await db.projects.find({}, {"_id": 0}).to_list(1000)
    return [ProjectResponse(**p) for p in projects]

@api_router.get("/projects/{project_id}", response_model=ProjectResponse)
async def get_project(project_id: str, current_user: dict = Depends(get_current_user)):
    project = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return ProjectResponse(**project)

@api_router.put("/projects/{project_id}", response_model=ProjectResponse)
async def update_project(project_id: str, project: ProjectCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.projects.find_one({"project_id": project_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Project not found")
    
    update_data = project.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.projects.update_one({"project_id": project_id}, {"$set": update_data})
    
    updated = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    return ProjectResponse(**updated)

@api_router.delete("/projects/{project_id}")
async def delete_project(project_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.projects.delete_one({"project_id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    # Delete related data
    await db.buildings.delete_many({"project_id": project_id})
    await db.construction_progress.delete_many({"project_id": project_id})
    await db.project_costs.delete_many({"project_id": project_id})
    await db.unit_sales.delete_many({"project_id": project_id})
    return {"message": "Project deleted"}

# =========================
# BUILDING ROUTES
# =========================

def calculate_building_totals(building_data: dict) -> dict:
    """Calculate total floors and units based on building type and configuration"""
    building_type = building_data.get("building_type", "residential_tower")
    
    # Calculate total parking floors
    total_parking = (
        building_data.get("parking_basement", 0) +
        building_data.get("parking_stilt_ground", 0) +
        building_data.get("parking_upper_level", 0)
    )
    
    # Calculate total floors based on building type
    if building_type in ["residential_tower", "mixed_tower"]:
        commercial = building_data.get("commercial_floors", 0) if building_type == "mixed_tower" else 0
        residential = building_data.get("residential_floors", 0)
        apartments_per_floor = building_data.get("apartments_per_floor", 0)
        
        total_floors = total_parking + commercial + residential
        total_units = residential * apartments_per_floor
    else:  # row_house or bungalow
        residential = building_data.get("residential_floors", 0)
        total_floors = total_parking + residential
        total_units = 1  # Each row house/bungalow is 1 unit
    
    building_data["floors"] = total_floors
    building_data["units"] = total_units
    building_data["total_parking_floors"] = total_parking
    
    return building_data

@api_router.post("/buildings", response_model=BuildingResponse)
async def create_building(building: BuildingCreate, current_user: dict = Depends(get_current_user)):
    building_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    building_data = building.model_dump()
    building_data = calculate_building_totals(building_data)
    
    building_doc = {
        "building_id": building_id,
        **building_data,
        "created_at": now
    }
    await db.buildings.insert_one(building_doc)
    return BuildingResponse(**{k: v for k, v in building_doc.items() if k != "_id"})

@api_router.post("/buildings/bulk", response_model=Dict[str, Any])
async def create_buildings_bulk(bulk_data: BuildingBulkCreate, current_user: dict = Depends(get_current_user)):
    """Create multiple buildings with the same configuration but different names"""
    created = []
    errors = []
    now = datetime.now(timezone.utc).isoformat()
    
    # Get template data and calculate totals
    template_data = bulk_data.template.model_dump()
    template_data = calculate_building_totals(template_data)
    
    for idx, name in enumerate(bulk_data.building_names):
        try:
            building_id = str(uuid.uuid4())
            building_doc = {
                "building_id": building_id,
                "project_id": bulk_data.project_id,
                **template_data,
                "building_name": name.strip(),
                "created_at": now
            }
            await db.buildings.insert_one(building_doc)
            created.append({"building_id": building_id, "building_name": name})
        except Exception as e:
            errors.append({"name": name, "error": str(e)})
    
    return {
        "created": len(created),
        "buildings": created,
        "errors": errors
    }

@api_router.get("/buildings/types")
async def get_building_types():
    """Get available building types with their configurations"""
    return {
        "types": [
            {
                "value": "residential_tower",
                "label": "Multi-storey Residential Apartment Tower",
                "has_commercial": False,
                "has_apartments_per_floor": True,
                "is_single_unit": False
            },
            {
                "value": "mixed_tower",
                "label": "Multi-storey Mixed (Commercial & Residential) Tower",
                "has_commercial": True,
                "has_apartments_per_floor": True,
                "is_single_unit": False
            },
            {
                "value": "row_house",
                "label": "Row House",
                "has_commercial": False,
                "has_apartments_per_floor": False,
                "is_single_unit": True
            },
            {
                "value": "bungalow",
                "label": "Bungalow",
                "has_commercial": False,
                "has_apartments_per_floor": False,
                "is_single_unit": True
            }
        ],
        "parking_options": [
            {"field": "parking_basement", "label": "Basement Parking Floors"},
            {"field": "parking_stilt_ground", "label": "Stilt/Ground Parking Floors"},
            {"field": "parking_upper_level", "label": "Upper Level Parking Floors"}
        ]
    }

@api_router.get("/buildings", response_model=List[BuildingResponse])
async def get_buildings(project_id: str = Query(...), current_user: dict = Depends(get_current_user)):
    buildings = await db.buildings.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    result = []
    for b in buildings:
        # Ensure backward compatibility - calculate totals if missing
        if "total_parking_floors" not in b:
            b = calculate_building_totals(b)
        result.append(BuildingResponse(**b))
    return result

@api_router.get("/buildings/{building_id}", response_model=BuildingResponse)
async def get_building(building_id: str, current_user: dict = Depends(get_current_user)):
    building = await db.buildings.find_one({"building_id": building_id}, {"_id": 0})
    if not building:
        raise HTTPException(status_code=404, detail="Building not found")
    if "total_parking_floors" not in building:
        building = calculate_building_totals(building)
    return BuildingResponse(**building)

@api_router.put("/buildings/{building_id}", response_model=BuildingResponse)
async def update_building(building_id: str, building: BuildingBase, current_user: dict = Depends(get_current_user)):
    existing = await db.buildings.find_one({"building_id": building_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Building not found")
    
    building_data = building.model_dump()
    building_data = calculate_building_totals(building_data)
    
    await db.buildings.update_one({"building_id": building_id}, {"$set": building_data})
    updated = await db.buildings.find_one({"building_id": building_id}, {"_id": 0})
    return BuildingResponse(**updated)

@api_router.delete("/buildings/{building_id}")
async def delete_building(building_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.buildings.delete_one({"building_id": building_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Building not found")
    return {"message": "Building deleted"}

# =========================
# INFRASTRUCTURE COST ROUTES
# =========================

INFRASTRUCTURE_ITEMS = [
    {"id": "road_footpath_storm_drain", "name": "Road, Foot-path and storm water drain", "weightage": 21.5},
    {"id": "underground_sewage_network", "name": "Underground sewage drainage network", "weightage": 13.0},
    {"id": "sewage_treatment_plant", "name": "Sewage Treatment Plant", "weightage": 8.5},
    {"id": "overhead_sump_reservoir", "name": "Over-head and Sump water reservoir/Tank", "weightage": 8.5},
    {"id": "underground_water_distribution", "name": "Under ground water distribution network", "weightage": 10.5},
    {"id": "electric_substation_cables", "name": "Electric Substation & Under-ground electric cables", "weightage": 10.5},
    {"id": "street_lights", "name": "Street Lights", "weightage": 4.0},
    {"id": "entry_gate", "name": "Entry Gate", "weightage": 2.5},
    {"id": "boundary_wall", "name": "Boundary wall", "weightage": 6.0},
    {"id": "club_house", "name": "Club House", "weightage": 7.0},
    {"id": "swimming_pool", "name": "Swimming Pool", "weightage": 3.5},
    {"id": "amphitheatre", "name": "Amphitheatre", "weightage": 2.5},
    {"id": "gardens_playground", "name": "Gardens / Play Ground", "weightage": 2.0}
]

@api_router.get("/infrastructure-costs/template")
async def get_infrastructure_cost_template():
    """Returns infrastructure cost items with weightages"""
    return {"items": INFRASTRUCTURE_ITEMS, "total_weightage": 100.0}

@api_router.post("/infrastructure-costs")
async def create_infrastructure_cost(
    project_id: str = Query(...),
    current_user: dict = Depends(get_current_user),
    request: Request = None
):
    """Save infrastructure cost estimates"""
    costs = await request.json() if request else {}
    
    # Calculate total infrastructure cost
    total = 0
    for item in INFRASTRUCTURE_ITEMS:
        item_data = costs.get(item["id"], {})
        if item_data.get("is_applicable", True):
            total += item_data.get("estimated_cost", 0)
    
    cost_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    cost_doc = {
        "cost_id": cost_id,
        "project_id": project_id,
        "costs": costs,
        "total_infrastructure_cost": total,
        "created_at": now
    }
    
    await db.infrastructure_costs.update_one(
        {"project_id": project_id},
        {"$set": cost_doc},
        upsert=True
    )
    
    return cost_doc

@api_router.get("/infrastructure-costs/{project_id}")
async def get_infrastructure_cost(project_id: str, current_user: dict = Depends(get_current_user)):
    cost = await db.infrastructure_costs.find_one({"project_id": project_id}, {"_id": 0})
    if not cost:
        # Return empty structure
        return {
            "project_id": project_id,
            "costs": {},
            "total_infrastructure_cost": 0
        }
    return cost

# =========================
# ESTIMATED DEVELOPMENT COST ROUTES
# =========================

@api_router.post("/estimated-development-cost")
async def create_estimated_development_cost(
    project_id: str = Query(...),
    current_user: dict = Depends(get_current_user),
    request: Request = None
):
    """Save estimated development cost (fixed values for project lifecycle)"""
    data = await request.json() if request else {}
    
    # Get total buildings cost from all buildings
    buildings = await db.buildings.find({"project_id": project_id}, {"_id": 0}).to_list(100)
    buildings_cost = sum(b.get("estimated_cost", 0) for b in buildings)
    
    # Get infrastructure cost
    infra_cost = await db.infrastructure_costs.find_one({"project_id": project_id}, {"_id": 0})
    infrastructure_cost = infra_cost.get("total_infrastructure_cost", 0) if infra_cost else data.get("infrastructure_cost", 0)
    
    # Get manual inputs
    consultants_fee = data.get("consultants_fee", 0)
    machinery_cost = data.get("machinery_cost", 0)
    site_development_cost = data.get("site_development_cost", 0)
    salaries = data.get("salaries", 0)
    site_overheads = data.get("site_overheads", 0)
    services_cost = data.get("services_cost", 0)
    taxes_premiums_fees = data.get("taxes_premiums_fees", 0)
    finance_cost = data.get("finance_cost", 0)
    
    # Calculate total
    total = (buildings_cost + infrastructure_cost + 
             site_development_cost + salaries + consultants_fee +
             site_overheads + services_cost + machinery_cost +
             taxes_premiums_fees + finance_cost)
    
    estimate_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    estimate_doc = {
        "estimate_id": estimate_id,
        "project_id": project_id,
        "buildings_cost": buildings_cost,
        "infrastructure_cost": infrastructure_cost,
        "site_development_cost": site_development_cost,
        "salaries": salaries,
        "consultants_fee": consultants_fee,
        "site_overheads": site_overheads,
        "services_cost": services_cost,
        "machinery_cost": machinery_cost,
        "taxes_premiums_fees": taxes_premiums_fees,
        "finance_cost": finance_cost,
        "total_estimated_development_cost": total,
        "created_at": now,
        "updated_at": now
    }
    
    await db.estimated_development_costs.update_one(
        {"project_id": project_id},
        {"$set": estimate_doc},
        upsert=True
    )
    
    return estimate_doc

@api_router.get("/estimated-development-cost/{project_id}")
async def get_estimated_development_cost(project_id: str, current_user: dict = Depends(get_current_user)):
    estimate = await db.estimated_development_costs.find_one({"project_id": project_id}, {"_id": 0})
    
    # Always get latest buildings and infrastructure costs
    buildings = await db.buildings.find({"project_id": project_id}, {"_id": 0}).to_list(100)
    buildings_cost = sum(b.get("estimated_cost", 0) for b in buildings)
    
    infra_cost = await db.infrastructure_costs.find_one({"project_id": project_id}, {"_id": 0})
    infrastructure_cost = infra_cost.get("total_infrastructure_cost", 0) if infra_cost else 0
    
    # Get site expenditure (estimated)
    site_exp = await db.site_expenditure.find_one({"project_id": project_id}, {"_id": 0})
    site_expenditure = site_exp if site_exp else {
        "site_development_cost": 0,
        "salaries": 0,
        "consultants_fee": 0,
        "site_overheads": 0,
        "services_cost": 0,
        "machinery_cost": 0
    }
    
    # Calculate total
    site_exp_total = (
        site_expenditure.get("site_development_cost", 0) +
        site_expenditure.get("salaries", 0) +
        site_expenditure.get("consultants_fee", 0) +
        site_expenditure.get("site_overheads", 0) +
        site_expenditure.get("services_cost", 0) +
        site_expenditure.get("machinery_cost", 0)
    )
    
    if not estimate:
        return {
            "project_id": project_id,
            "buildings_cost": buildings_cost,
            "infrastructure_cost": infrastructure_cost,
            "site_development_cost": site_expenditure.get("site_development_cost", 0),
            "salaries": site_expenditure.get("salaries", 0),
            "consultants_fee": site_expenditure.get("consultants_fee", 0),
            "site_overheads": site_expenditure.get("site_overheads", 0),
            "services_cost": site_expenditure.get("services_cost", 0),
            "machinery_cost": site_expenditure.get("machinery_cost", 0),
            "site_expenditure_total": site_exp_total,
            "taxes_premiums_fees": 0,
            "finance_cost": 0,
            "total_estimated_development_cost": buildings_cost + infrastructure_cost + site_exp_total,
            "is_draft": True
        }
    
    # Update with latest auto-calculated values from other sections
    estimate["buildings_cost"] = buildings_cost
    estimate["infrastructure_cost"] = infrastructure_cost
    estimate["site_development_cost"] = site_expenditure.get("site_development_cost", 0)
    estimate["salaries"] = site_expenditure.get("salaries", 0)
    estimate["consultants_fee"] = site_expenditure.get("consultants_fee", 0)
    estimate["site_overheads"] = site_expenditure.get("site_overheads", 0)
    estimate["services_cost"] = site_expenditure.get("services_cost", 0)
    estimate["machinery_cost"] = site_expenditure.get("machinery_cost", 0)
    estimate["site_expenditure_total"] = site_exp_total
    estimate["total_estimated_development_cost"] = (
        buildings_cost + infrastructure_cost + site_exp_total +
        estimate.get("taxes_premiums_fees", 0) + estimate.get("finance_cost", 0)
    )
    
    return estimate

@api_router.get("/estimated-development-cost/{project_id}/refresh-buildings")
async def refresh_buildings_cost(project_id: str, current_user: dict = Depends(get_current_user)):
    """Refresh buildings cost from current building data (before saving estimate)"""
    buildings = await db.buildings.find({"project_id": project_id}, {"_id": 0}).to_list(100)
    buildings_cost = sum(b.get("estimated_cost", 0) for b in buildings)
    
    infra_cost = await db.infrastructure_costs.find_one({"project_id": project_id}, {"_id": 0})
    infrastructure_cost = infra_cost.get("total_infrastructure_cost", 0) if infra_cost else 0
    
    return {
        "buildings_cost": buildings_cost,
        "infrastructure_cost": infrastructure_cost
    }

# =========================
# SITE EXPENDITURE ROUTES (Estimated)
# =========================

@api_router.get("/site-expenditure/{project_id}")
async def get_site_expenditure(project_id: str, current_user: dict = Depends(get_current_user)):
    """Get estimated on-site expenditure for a project"""
    expenditure = await db.site_expenditure.find_one({"project_id": project_id}, {"_id": 0})
    if not expenditure:
        return {
            "project_id": project_id,
            "site_development_cost": 0,
            "salaries": 0,
            "consultants_fee": 0,
            "site_overheads": 0,
            "services_cost": 0,
            "machinery_cost": 0,
            "total": 0
        }
    return expenditure

@api_router.post("/site-expenditure")
async def save_site_expenditure(
    project_id: str = Query(...),
    request: Request = None,
    current_user: dict = Depends(get_current_user)
):
    """Save estimated on-site expenditure for a project"""
    data = await request.json()
    
    site_development_cost = data.get("site_development_cost", 0)
    salaries = data.get("salaries", 0)
    consultants_fee = data.get("consultants_fee", 0)
    site_overheads = data.get("site_overheads", 0)
    services_cost = data.get("services_cost", 0)
    machinery_cost = data.get("machinery_cost", 0)
    
    total = site_development_cost + salaries + consultants_fee + site_overheads + services_cost + machinery_cost
    
    now = datetime.now(timezone.utc).isoformat()
    
    expenditure_doc = {
        "project_id": project_id,
        "site_development_cost": site_development_cost,
        "salaries": salaries,
        "consultants_fee": consultants_fee,
        "site_overheads": site_overheads,
        "services_cost": services_cost,
        "machinery_cost": machinery_cost,
        "total": total,
        "updated_at": now
    }
    
    await db.site_expenditure.update_one(
        {"project_id": project_id},
        {"$set": expenditure_doc},
        upsert=True
    )
    
    return expenditure_doc

# =========================
# ACTUAL SITE EXPENDITURE ROUTES (per quarter)
# =========================

@api_router.get("/actual-site-expenditure/{project_id}")
async def get_actual_site_expenditure(
    project_id: str, 
    quarter: str = Query(...),
    year: int = Query(...),
    current_user: dict = Depends(get_current_user)
):
    """Get actual on-site expenditure for a project for a specific quarter"""
    expenditure = await db.actual_site_expenditure.find_one(
        {"project_id": project_id, "quarter": quarter, "year": year}, 
        {"_id": 0}
    )
    if not expenditure:
        return {
            "project_id": project_id,
            "quarter": quarter,
            "year": year,
            "site_development_cost": 0,
            "salaries": 0,
            "consultants_fee": 0,
            "site_overheads": 0,
            "services_cost": 0,
            "machinery_cost": 0,
            "total": 0
        }
    return expenditure

@api_router.post("/actual-site-expenditure")
async def save_actual_site_expenditure(
    project_id: str = Query(...),
    quarter: str = Query(...),
    year: int = Query(...),
    request: Request = None,
    current_user: dict = Depends(get_current_user)
):
    """Save actual on-site expenditure for a project for a specific quarter"""
    data = await request.json()
    
    site_development_cost = data.get("site_development_cost", 0)
    salaries = data.get("salaries", 0)
    consultants_fee = data.get("consultants_fee", 0)
    site_overheads = data.get("site_overheads", 0)
    services_cost = data.get("services_cost", 0)
    machinery_cost = data.get("machinery_cost", 0)
    
    total = site_development_cost + salaries + consultants_fee + site_overheads + services_cost + machinery_cost
    
    now = datetime.now(timezone.utc).isoformat()
    
    expenditure_doc = {
        "project_id": project_id,
        "quarter": quarter,
        "year": year,
        "site_development_cost": site_development_cost,
        "salaries": salaries,
        "consultants_fee": consultants_fee,
        "site_overheads": site_overheads,
        "services_cost": services_cost,
        "machinery_cost": machinery_cost,
        "total": total,
        "updated_at": now
    }
    
    await db.actual_site_expenditure.update_one(
        {"project_id": project_id, "quarter": quarter, "year": year},
        {"$set": expenditure_doc},
        upsert=True
    )
    
    return expenditure_doc

# =========================
# LAND COST ROUTES
# =========================

@api_router.get("/land-cost/{project_id}")
async def get_land_cost(
    project_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get land cost data for a project (both estimated and actual)"""
    land_cost = await db.land_costs.find_one({"project_id": project_id}, {"_id": 0})
    
    default_fields = {
        "land_cost": 0,
        "premium_cost": 0,
        "tdr_cost": 0,
        "statutory_cost": 0,
        "land_premium": 0,
        "under_rehab_scheme": 0,
        "estimated_rehab_cost": 0,
        "actual_rehab_cost": 0,
        "land_clearance_cost": 0,
        "asr_linked_premium": 0
    }
    
    if not land_cost:
        return {
            "project_id": project_id,
            "estimated": {**default_fields, "total": 0},
            "actual": {**default_fields, "total": 0}
        }
    return land_cost

@api_router.post("/land-cost/{project_id}")
async def save_land_cost(
    project_id: str,
    request: Request,
    current_user: dict = Depends(get_current_user)
):
    """Save land cost data for a project (both estimated and actual)"""
    data = await request.json()
    
    def extract_costs(cost_data):
        fields = [
            "land_cost", "premium_cost", "tdr_cost", "statutory_cost",
            "land_premium", "under_rehab_scheme", "estimated_rehab_cost",
            "actual_rehab_cost", "land_clearance_cost", "asr_linked_premium"
        ]
        extracted = {f: cost_data.get(f, 0) or 0 for f in fields}
        extracted["total"] = sum(extracted.values())
        return extracted
    
    estimated = extract_costs(data.get("estimated", {}))
    actual = extract_costs(data.get("actual", {}))
    
    now = datetime.now(timezone.utc).isoformat()
    
    land_cost_doc = {
        "project_id": project_id,
        "estimated": estimated,
        "actual": actual,
        "updated_at": now
    }
    
    await db.land_costs.update_one(
        {"project_id": project_id},
        {"$set": land_cost_doc},
        upsert=True
    )
    
    return land_cost_doc

# =========================
# CONSTRUCTION PROGRESS ROUTES
# =========================

DEFAULT_ACTIVITIES = [
    {"activity_name": "Excavation", "weightage": 5, "completion_percentage": 0},
    {"activity_name": "Basement / Plinth", "weightage": 8, "completion_percentage": 0},
    {"activity_name": "Podium", "weightage": 5, "completion_percentage": 0},
    {"activity_name": "Stilt Floor", "weightage": 5, "completion_percentage": 0},
    {"activity_name": "Superstructure Slabs", "weightage": 20, "completion_percentage": 0},
    {"activity_name": "Internal Walls & Plaster", "weightage": 10, "completion_percentage": 0},
    {"activity_name": "Doors & Windows", "weightage": 7, "completion_percentage": 0},
    {"activity_name": "Sanitary Fittings", "weightage": 5, "completion_percentage": 0},
    {"activity_name": "Electrical Fittings", "weightage": 5, "completion_percentage": 0},
    {"activity_name": "Staircases and Lift Wells", "weightage": 5, "completion_percentage": 0},
    {"activity_name": "Water Tanks", "weightage": 3, "completion_percentage": 0},
    {"activity_name": "External Plumbing and Plaster", "weightage": 7, "completion_percentage": 0},
    {"activity_name": "Terrace Waterproofing", "weightage": 3, "completion_percentage": 0},
    {"activity_name": "Fire Fighting Systems", "weightage": 5, "completion_percentage": 0},
    {"activity_name": "Common Area Finishing", "weightage": 4, "completion_percentage": 0},
    {"activity_name": "Compound Wall and Final Works", "weightage": 3, "completion_percentage": 0},
]

@api_router.post("/construction-progress", response_model=ConstructionProgressResponse)
async def create_construction_progress(progress: ConstructionProgressCreate, current_user: dict = Depends(get_current_user)):
    # Get building to get project_id
    building = await db.buildings.find_one({"building_id": progress.building_id})
    if not building:
        raise HTTPException(status_code=404, detail="Building not found")
    
    # Calculate overall completion
    total_weighted = sum(a.weightage * a.completion_percentage / 100 for a in progress.activities)
    overall_completion = total_weighted
    
    progress_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    progress_doc = {
        "progress_id": progress_id,
        "project_id": building["project_id"],
        "building_id": progress.building_id,
        "quarter": progress.quarter,
        "year": progress.year,
        "activities": [a.model_dump() for a in progress.activities],
        "overall_completion": overall_completion,
        "created_at": now
    }
    
    # Upsert - update if exists for same building/quarter/year
    await db.construction_progress.update_one(
        {"building_id": progress.building_id, "quarter": progress.quarter, "year": progress.year},
        {"$set": progress_doc},
        upsert=True
    )
    return ConstructionProgressResponse(**progress_doc)

@api_router.get("/construction-progress", response_model=List[ConstructionProgressResponse])
async def get_construction_progress(
    project_id: str = Query(...),
    quarter: Optional[str] = None,
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"project_id": project_id}
    if quarter:
        query["quarter"] = quarter
    if year:
        query["year"] = year
    
    progress_list = await db.construction_progress.find(query, {"_id": 0}).to_list(1000)
    return [ConstructionProgressResponse(**p) for p in progress_list]

@api_router.get("/construction-progress/default-activities")
async def get_default_activities():
    return DEFAULT_ACTIVITIES

# Get comprehensive construction progress template with all activities
@api_router.get("/construction-progress/detailed-template")
async def get_detailed_construction_template():
    """Returns the comprehensive construction progress tracking template"""
    return {
        "tower_construction": {
            "total_weightage": 100,
            "categories": [
                {
                    "id": "plinth_completion",
                    "name": "Completion of Plinth",
                    "total_weightage": 10.89,
                    "activities": [
                        {"id": "excavation", "name": "Excavation", "weightage": 0.90},
                        {"id": "pcc_below_footing", "name": "PCC below footing", "weightage": 0.80},
                        {"id": "shuttering_for_footing", "name": "Shuttering for Footing", "weightage": 0.90},
                        {"id": "reinforcement_footing_column", "name": "Reinforcement for Footing and Column", "weightage": 1.70},
                        {"id": "concreting_for_footing", "name": "Concreting for Footing", "weightage": 1.50},
                        {"id": "shuttering_column_to_plinth", "name": "Shuttering for Column up to Plinth", "weightage": 0.50},
                        {"id": "concreting_for_column", "name": "Concreting for Column", "weightage": 1.79},
                        {"id": "shuttering_plinth_beam", "name": "Shuttering for Plinth Beam", "weightage": 0.70},
                        {"id": "reinforcement_plinth_beam", "name": "Reinforcement for Plinth Beam", "weightage": 1.00},
                        {"id": "concreting_plinth_beam", "name": "Concreting for Plinth Beam", "weightage": 0.80},
                        {"id": "filling_earth_plinth_pcc", "name": "Filling earth within Plinth and PCC", "weightage": 0.30}
                    ]
                },
                {
                    "id": "slab_completion",
                    "name": "Completion of Slabs at all levels",
                    "total_weightage": 31.78,
                    "note": "Divide by number of floors from Buildings section",
                    "activities": [
                        {"id": "reinforcement_lintel_roof", "name": "Reinforcement up to Lintel/Roof bottom", "weightage": 6.00},
                        {"id": "shuttering_for_column", "name": "Shuttering for Column", "weightage": 4.40},
                        {"id": "concreting_for_column", "name": "Concreting for Column", "weightage": 5.00},
                        {"id": "shuttering_beams_roof", "name": "Shuttering for Beams and Roof", "weightage": 4.80},
                        {"id": "reinforcement_beams_roof", "name": "Reinforcement for Beams and Roof", "weightage": 6.00},
                        {"id": "concreting_beams_roof", "name": "Concreting for Beams and Roof", "weightage": 4.00},
                        {"id": "dismantling_roof_shuttering", "name": "Dismantling of Roof Shuttering", "weightage": 1.58}
                    ]
                },
                {
                    "id": "brickwork_plastering",
                    "name": "Completion of Brickwork and Plastering",
                    "total_weightage": 12.71,
                    "activities": [
                        {"id": "brickwork_external_walls", "name": "Brickwork External walls", "weightage": 2.50},
                        {"id": "brickwork_internal_walls", "name": "Brickwork Internal walls", "weightage": 2.80},
                        {"id": "fixing_door_window_frames", "name": "Fixing of Door/Window frames", "weightage": 1.50},
                        {"id": "fixing_concealed_pipes", "name": "Fixing concealed Water & Electric pipes", "weightage": 1.40},
                        {"id": "plastering_external_walls", "name": "Plastering External walls", "weightage": 1.30},
                        {"id": "plastering_internal_walls", "name": "Plastering Internal walls", "weightage": 1.80},
                        {"id": "waterproof_plastering_toilets", "name": "Water-proof Plastering of Toilets", "weightage": 1.41}
                    ]
                },
                {
                    "id": "plumbing",
                    "name": "Plumbing",
                    "total_weightage": 3.93,
                    "activities": [
                        {"id": "fixing_water_pipes", "name": "Fixing External & Internal Water Pipes", "weightage": 1.40},
                        {"id": "fixing_wc_pipes_traps", "name": "Fixing External & Internal WC Pipes & Traps", "weightage": 1.30},
                        {"id": "fixing_plumbing_fixtures", "name": "Fixing of all Plumbing fixtures", "weightage": 1.23}
                    ]
                },
                {
                    "id": "electrical_works",
                    "name": "Electrical Works",
                    "total_weightage": 9.08,
                    "activities": [
                        {"id": "laying_all_cables", "name": "Laying all Cables (Internal)", "weightage": 3.00},
                        {"id": "fixing_electrical_fixtures", "name": "Fixing all Electrical fixtures (Internal)", "weightage": 2.00},
                        {"id": "electrical_breaker_box", "name": "Electrical/Breaker Box (External)", "weightage": 2.00},
                        {"id": "electric_meter_box", "name": "Electric meter box (External)", "weightage": 1.00},
                        {"id": "connecting_cable_electrical_box", "name": "Connecting cable to the Electrical box", "weightage": 1.08}
                    ]
                },
                {
                    "id": "window_works",
                    "name": "Aluminium/UPVC Window",
                    "total_weightage": 8.02,
                    "activities": [
                        {"id": "fixing_frames", "name": "Fixing of frames", "weightage": 4.40},
                        {"id": "fixing_glass", "name": "Fixing of Glass", "weightage": 3.62}
                    ]
                },
                {
                    "id": "tiling_flooring",
                    "name": "Tiling/Flooring",
                    "total_weightage": 8.02,
                    "activities": [
                        {"id": "laying_floor_tiles", "name": "Laying of Floor tiles", "weightage": 3.00},
                        {"id": "laying_wall_tiles_kitchen_bathroom", "name": "Laying of Wall tiles Kitchen & Bathroom", "weightage": 2.80},
                        {"id": "laying_granite_kitchen_counter", "name": "Laying of Granite/Kadapa slab for Kitchen Counter", "weightage": 2.22}
                    ]
                },
                {
                    "id": "door_shutter_fixing",
                    "name": "Door Shutter Fixing",
                    "total_weightage": 2.42,
                    "activities": [
                        {"id": "fixing_door_shutters", "name": "Fixing of Door shutters", "weightage": 1.30},
                        {"id": "fixing_locks_handles", "name": "Fixing of locks, handles & accessories", "weightage": 1.12}
                    ]
                },
                {
                    "id": "water_proofing",
                    "name": "Water Proofing",
                    "total_weightage": 2.42,
                    "activities": [
                        {"id": "terrace_roof_waterproofing", "name": "Terrace roof water proofing", "weightage": 2.42}
                    ]
                },
                {
                    "id": "painting",
                    "name": "Painting",
                    "total_weightage": 8.32,
                    "activities": [
                        {"id": "painting_ceiling", "name": "Ceiling", "weightage": 2.30},
                        {"id": "painting_walls", "name": "Walls", "weightage": 2.80},
                        {"id": "painting_grills", "name": "Grills", "weightage": 1.80},
                        {"id": "painting_doors_windows", "name": "Doors/Windows", "weightage": 1.42}
                    ]
                },
                {
                    "id": "carpark",
                    "name": "Carpark",
                    "total_weightage": 0.76,
                    "activities": [
                        {"id": "levelling", "name": "Levelling", "weightage": 0.20},
                        {"id": "paving", "name": "Paving", "weightage": 0.56}
                    ]
                },
                {
                    "id": "handover_intimation",
                    "name": "Intimation of Handover",
                    "total_weightage": 1.66,
                    "activities": [
                        {"id": "intimation_of_handover", "name": "Intimation of Handover", "weightage": 1.66}
                    ]
                }
            ]
        },
        "infrastructure_works": {
            "total_weightage": 100,
            "activities": [
                {"id": "road_footpath_storm_drain", "name": "Road, Foot-path and storm water drain", "weightage": 21.5},
                {"id": "underground_sewage_network", "name": "Underground sewage drainage network", "weightage": 13.0},
                {"id": "sewage_treatment_plant", "name": "Sewage Treatment Plant", "weightage": 8.5},
                {"id": "overhead_sump_reservoir", "name": "Over-head and Sump water reservoir/Tank", "weightage": 8.5},
                {"id": "underground_water_distribution", "name": "Under ground water distribution network", "weightage": 10.5},
                {"id": "electric_substation_cables", "name": "Electric Substation & Under-ground electric cables", "weightage": 10.5},
                {"id": "street_lights", "name": "Street Lights", "weightage": 4.0},
                {"id": "entry_gate", "name": "Entry Gate", "weightage": 2.5},
                {"id": "boundary_wall", "name": "Boundary wall", "weightage": 6.0},
                {"id": "club_house", "name": "Club House", "weightage": 7.0},
                {"id": "swimming_pool", "name": "Swimming Pool", "weightage": 3.5},
                {"id": "amphitheatre", "name": "Amphitheatre", "weightage": 2.5},
                {"id": "gardens_playground", "name": "Gardens / Play Ground", "weightage": 2.0}
            ]
        }
    }

# Calculate recalibrated weightage after N/A items
def calculate_recalibrated_completion(activities_data: dict, template_categories: list) -> tuple:
    """Calculate completion with recalibrated weightages for N/A items.
    Supports cost-based weightage mode per category (flag: _use_cost_weightage).
    When cost mode is ON, each sub-activity's effective weightage = (cost / total_cost) * cat_base_applicable_wt.
    """
    total_applicable_weightage = 0
    weighted_completion = 0
    category_completions = {}

    for category in template_categories:
        cat_id = category["id"]
        cat_data = activities_data.get(cat_id, {})
        cat_applicable_weightage = 0
        cat_weighted_completion = 0

        use_cost_weightage = cat_data.get("_use_cost_weightage", False)

        if use_cost_weightage:
            # --- Cost-based weightage mode ---
            # Sum cost and base-weightage for applicable activities only
            total_cost = 0
            cat_base_applicable = 0
            for activity in category["activities"]:
                act_id = activity["id"]
                act_data = cat_data.get(act_id, {})
                if act_data.get("is_applicable", True):
                    total_cost += float(act_data.get("cost", 0) or 0)
                    cat_base_applicable += activity["weightage"]

            for activity in category["activities"]:
                act_id = activity["id"]
                act_data = cat_data.get(act_id, {})
                is_applicable = act_data.get("is_applicable", True)
                if is_applicable:
                    completion = act_data.get("completion", 0)
                    cost = float(act_data.get("cost", 0) or 0)
                    # Effective weight = proportional to cost; if no costs yet, fall back to template
                    if total_cost > 0:
                        effective_wt = (cost / total_cost) * cat_base_applicable
                    else:
                        effective_wt = activity["weightage"]

                    total_applicable_weightage += effective_wt
                    cat_applicable_weightage += effective_wt
                    weighted_completion += effective_wt * completion / 100
                    cat_weighted_completion += effective_wt * completion / 100
        else:
            # --- Standard (template) weightage mode ---
            for activity in category["activities"]:
                act_id = activity["id"]
                act_data = cat_data.get(act_id, {})
                is_applicable = act_data.get("is_applicable", True)
                if is_applicable:
                    base_weightage = activity["weightage"]
                    completion = act_data.get("completion", 0)

                    total_applicable_weightage += base_weightage
                    cat_applicable_weightage += base_weightage
                    weighted_completion += base_weightage * completion / 100
                    cat_weighted_completion += base_weightage * completion / 100

        # Calculate category completion percentage
        if cat_applicable_weightage > 0:
            category_completions[cat_id] = round(cat_weighted_completion / cat_applicable_weightage * 100, 2)
        else:
            category_completions[cat_id] = 0  # All N/A

    # Calculate overall completion (recalibrated)
    overall_completion = 0
    if total_applicable_weightage > 0:
        overall_completion = weighted_completion / total_applicable_weightage * 100

    return round(overall_completion, 2), total_applicable_weightage, category_completions

# Save detailed construction progress with N/A support
@api_router.post("/construction-progress/detailed")
async def create_detailed_construction_progress(
    building_id: str = Query(...),
    quarter: str = Query(...),
    year: int = Query(...),
    number_of_floors: int = Query(1),
    current_user: dict = Depends(get_current_user),
    request: Request = None
):
    """Save detailed tower construction progress with N/A support and weightage recalibration"""
    # Get tower_activities from request body
    tower_activities = await request.json() if request else {}
    
    # Get building to get project_id
    building = await db.buildings.find_one({"building_id": building_id})
    if not building:
        raise HTTPException(status_code=404, detail="Building not found")
    
    # Get template for recalibration
    template = await get_detailed_construction_template()
    categories = template["tower_construction"]["categories"]
    
    # Calculate recalibrated completion
    overall_completion, total_applicable_weightage, category_completions = calculate_recalibrated_completion(
        tower_activities, categories
    )
    
    progress_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    progress_doc = {
        "progress_id": progress_id,
        "project_id": building["project_id"],
        "building_id": building_id,
        "quarter": quarter,
        "year": year,
        "tower_activities": tower_activities,
        "number_of_floors": number_of_floors,
        "overall_completion": overall_completion,
        "recalibrated_total_weightage": total_applicable_weightage,
        "category_completions": category_completions,
        "created_at": now
    }
    
    # Upsert
    await db.construction_progress.update_one(
        {"building_id": building_id, "quarter": quarter, "year": year},
        {"$set": progress_doc},
        upsert=True
    )
    
    return progress_doc

# =========================
# INFRASTRUCTURE PROGRESS ROUTES
# =========================

@api_router.post("/infrastructure-progress")
async def create_infrastructure_progress(
    project_id: str = Query(...),
    quarter: str = Query(...),
    year: int = Query(...),
    current_user: dict = Depends(get_current_user),
    request: Request = None
):
    """Save infrastructure works progress with N/A support and recalibration"""
    # Get activities from request body
    activities = await request.json() if request else {}
    
    # Get infrastructure template
    infrastructure_template = [
        {"id": "road_footpath_storm_drain", "name": "Road, Foot-path and storm water drain", "weightage": 21.5},
        {"id": "underground_sewage_network", "name": "Underground sewage drainage network", "weightage": 13.0},
        {"id": "sewage_treatment_plant", "name": "Sewage Treatment Plant", "weightage": 8.5},
        {"id": "overhead_sump_reservoir", "name": "Over-head and Sump water reservoir/Tank", "weightage": 8.5},
        {"id": "underground_water_distribution", "name": "Under ground water distribution network", "weightage": 10.5},
        {"id": "electric_substation_cables", "name": "Electric Substation & Under-ground electric cables", "weightage": 10.5},
        {"id": "street_lights", "name": "Street Lights", "weightage": 4.0},
        {"id": "entry_gate", "name": "Entry Gate", "weightage": 2.5},
        {"id": "boundary_wall", "name": "Boundary wall", "weightage": 6.0},
        {"id": "club_house", "name": "Club House", "weightage": 7.0},
        {"id": "swimming_pool", "name": "Swimming Pool", "weightage": 3.5},
        {"id": "amphitheatre", "name": "Amphitheatre", "weightage": 2.5},
        {"id": "gardens_playground", "name": "Gardens / Play Ground", "weightage": 2.0}
    ]
    
    # Calculate recalibrated completion
    total_applicable_weightage = 0
    weighted_completion = 0
    
    for item in infrastructure_template:
        item_id = item["id"]
        item_data = activities.get(item_id, {})
        is_applicable = item_data.get("is_applicable", True)
        
        if is_applicable:
            completion = item_data.get("completion", 0)
            weightage = item["weightage"]
            total_applicable_weightage += weightage
            weighted_completion += weightage * completion / 100
    
    overall_completion = 0
    if total_applicable_weightage > 0:
        overall_completion = weighted_completion / total_applicable_weightage * 100
    
    progress_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    progress_doc = {
        "progress_id": progress_id,
        "project_id": project_id,
        "quarter": quarter,
        "year": year,
        "activities": activities,
        "overall_completion": round(overall_completion, 2),
        "recalibrated_total_weightage": total_applicable_weightage,
        "created_at": now
    }
    
    await db.infrastructure_progress.update_one(
        {"project_id": project_id, "quarter": quarter, "year": year},
        {"$set": progress_doc},
        upsert=True
    )
    
    return progress_doc

@api_router.get("/infrastructure-progress")
async def get_infrastructure_progress(
    project_id: str = Query(...),
    quarter: Optional[str] = None,
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"project_id": project_id}
    if quarter:
        query["quarter"] = quarter
    if year:
        query["year"] = year
    
    progress_list = await db.infrastructure_progress.find(query, {"_id": 0}).to_list(100)
    return progress_list

@api_router.get("/infrastructure-progress/latest/{project_id}")
async def get_latest_infrastructure_progress(project_id: str, current_user: dict = Depends(get_current_user)):
    progress = await db.infrastructure_progress.find_one(
        {"project_id": project_id},
        {"_id": 0},
        sort=[("year", -1), ("quarter", -1)]
    )
    if not progress:
        raise HTTPException(status_code=404, detail="No infrastructure progress data found")
    return progress

# =========================
# PROJECT COST ROUTES (FORM-4 CA Certificate)
# =========================

async def _compute_cost_summary(project_id: str):
    """
    Compute Estimated Cost, Cost Incurred, and Balance Cost using exactly
    the same logic as the Project Costs page frontend, for the latest
    available quarter.  This is the single source of truth used by both
    the dashboard and the /project-costs/live-summary endpoint.
    """
    # ── Determine latest quarter ────────────────────────────────────────
    latest_cost_rec = await db.project_costs.find_one(
        {"project_id": project_id}, {"_id": 0}, sort=[("year", -1), ("quarter", -1)]
    )
    latest_progress_rec = await db.construction_progress.find_one(
        {"project_id": project_id}, {"_id": 0}, sort=[("year", -1), ("quarter", -1)]
    )
    if latest_cost_rec:
        quarter = latest_cost_rec.get("quarter", "Q1")
        year    = latest_cost_rec.get("year", datetime.now().year)
    elif latest_progress_rec:
        quarter = latest_progress_rec.get("quarter", "Q1")
        year    = latest_progress_rec.get("year", datetime.now().year)
    else:
        quarter = "Q1"
        year    = datetime.now().year

    # ── 1. Land costs ────────────────────────────────────────────────────
    land_doc = await db.land_costs.find_one({"project_id": project_id}, {"_id": 0})
    estimated_land = land_doc.get("estimated", {}).get("total", 0) if land_doc else 0
    actual_land    = land_doc.get("actual",    {}).get("total", 0) if land_doc else 0

    # ── 2. Estimated development cost ────────────────────────────────────
    # Replicate GET /estimated-development-cost/{project_id} exactly:
    # the stored total_estimated_development_cost may be stale, so we
    # recompute it live from its component collections.
    buildings = await db.buildings.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    buildings_cost_est = sum(b.get("estimated_cost", 0) for b in buildings)

    infra_cost_doc = await db.infrastructure_costs.find_one(
        {"project_id": project_id}, {"_id": 0}
    )
    total_infra_cost     = infra_cost_doc.get("total_infrastructure_cost", 0) if infra_cost_doc else 0

    # Estimated site expenditure (db.site_expenditure, NOT actual_site_expenditure)
    est_site_exp_doc = await db.site_expenditure.find_one({"project_id": project_id}, {"_id": 0})
    est_site_exp_total = (
        (est_site_exp_doc.get("site_development_cost", 0) or 0) +
        (est_site_exp_doc.get("salaries",              0) or 0) +
        (est_site_exp_doc.get("consultants_fee",       0) or 0) +
        (est_site_exp_doc.get("site_overheads",        0) or 0) +
        (est_site_exp_doc.get("services_cost",         0) or 0) +
        (est_site_exp_doc.get("machinery_cost",        0) or 0)
    ) if est_site_exp_doc else 0

    # taxes_premiums_fees and finance_cost are user-entered on the estimate
    est_dev_doc = await db.estimated_development_costs.find_one(
        {"project_id": project_id}, {"_id": 0}
    )
    est_taxes_premiums = est_dev_doc.get("taxes_premiums_fees", 0) if est_dev_doc else 0
    est_finance_cost   = est_dev_doc.get("finance_cost",        0) if est_dev_doc else 0

    estimated_dev = (
        buildings_cost_est + total_infra_cost + est_site_exp_total +
        est_taxes_premiums + est_finance_cost
    )

    # ── 3. Actual construction cost = Σ(building estimated_cost × completion%) ──
    construction_progress = await db.construction_progress.find(
        {"project_id": project_id, "quarter": quarter, "year": year}, {"_id": 0}
    ).to_list(1000)
    progress_map = {p.get("building_id"): p.get("overall_completion", 0) for p in construction_progress}
    actual_construction = sum(
        b.get("estimated_cost", 0) * progress_map.get(b.get("building_id"), 0) / 100
        for b in buildings
    )

    # ── 4. Actual infrastructure cost = infra_total × infra_completion% ──
    # infra_cost_doc and total_infra_cost already fetched above
    infra_progress_doc = await db.infrastructure_progress.find_one(
        {"project_id": project_id, "quarter": quarter, "year": year}, {"_id": 0}
    )
    infra_completion    = infra_progress_doc.get("overall_completion", 0) if infra_progress_doc else 0
    actual_infrastructure = total_infra_cost * infra_completion / 100

    # ── 5. Actual site expenditure for the latest quarter ────────────────
    site_exp_doc = await db.actual_site_expenditure.find_one(
        {"project_id": project_id, "quarter": quarter, "year": year}, {"_id": 0}
    )
    actual_site_exp = site_exp_doc.get("total", 0) if site_exp_doc else 0

    # ── 6. Taxes & finance (user-entered on Project Costs page) ──────────
    taxes_finance = 0
    if latest_cost_rec:
        taxes_finance = (
            latest_cost_rec.get("taxes_statutory", 0) +
            latest_cost_rec.get("finance_cost", 0)
        )

    # ── Totals (mirrors frontend Project Costs page exactly) ─────────────
    total_estimated = estimated_land + estimated_dev
    total_dev_cost  = actual_construction + actual_infrastructure + actual_site_exp + taxes_finance
    total_incurred  = actual_land + total_dev_cost
    balance_cost    = total_estimated - total_incurred

    return {
        "total_estimated_cost": round(total_estimated, 2),
        "total_cost_incurred":  round(total_incurred,  2),
        "balance_cost":         round(balance_cost,    2),
        "latest_quarter": quarter,
        "latest_year":    year,
        # Breakdown for transparency
        "estimated_land":         round(estimated_land,         2),
        "estimated_dev":          round(estimated_dev,          2),
        "actual_land":            round(actual_land,            2),
        "actual_construction":    round(actual_construction,    2),
        "actual_infrastructure":  round(actual_infrastructure,  2),
        "actual_site_exp":        round(actual_site_exp,        2),
        "taxes_finance":          round(taxes_finance,          2),
    }


async def _build_form4_data(
    project_id: str, quarter: str, year: int,
    buildings: list, construction_progress: list,
    project_cost: dict, sales: list
) -> dict:
    """
    Pre-compute every value needed by the Form-4 CA Certificate report.
    Fetches land_costs, site_expenditure, infrastructure collections in-place.
    Returns a flat dict consumed by all three report generators (Excel/PDF/DOCX).
    """
    pc = project_cost or {}

    # ── LAND COSTS (land_costs collection) ──────────────────────────────────
    land_doc    = await db.land_costs.find_one({"project_id": project_id}, {"_id": 0}) or {}
    land_est    = land_doc.get("estimated", {})
    land_inc    = land_doc.get("actual",    {})

    lc_a_est = land_est.get("land_cost", 0)
    lc_a_inc = land_inc.get("land_cost", 0)
    lc_b_est = land_est.get("premium_cost", 0)
    lc_b_inc = land_inc.get("premium_cost", 0)
    lc_c_est = land_est.get("tdr_cost", 0)
    lc_c_inc = land_inc.get("tdr_cost", 0)
    lc_d_est = land_est.get("statutory_cost", 0)
    lc_d_inc = land_inc.get("statutory_cost", 0)
    lc_e_est = land_est.get("land_premium", 0)
    lc_e_inc = land_inc.get("land_premium", 0)

    rehab_i_est   = land_est.get("estimated_rehab_cost", 0)
    rehab_i_inc   = land_inc.get("estimated_rehab_cost", 0)
    rehab_ii_inc  = land_inc.get("actual_rehab_cost", 0)
    rehab_iii_inc = land_inc.get("land_clearance_cost", 0)
    rehab_iv_inc  = land_inc.get("asr_linked_premium", 0)
    rehab_any     = any([rehab_i_inc, rehab_ii_inc, rehab_iii_inc, rehab_iv_inc])

    land_sub_est = lc_a_est + lc_b_est + lc_c_est + lc_d_est + lc_e_est + rehab_i_est
    land_sub_inc = lc_a_inc + lc_b_inc + lc_c_inc + lc_d_inc + lc_e_inc
    if rehab_any:
        land_sub_inc += min(rehab_i_inc, rehab_ii_inc) + rehab_iii_inc + rehab_iv_inc

    # ── DEVELOPMENT / CONSTRUCTION COSTS ────────────────────────────────────
    # a(i)  Estimated construction cost = sum of building estimated costs
    dev_a1_est = sum(b.get("estimated_cost", 0) for b in (buildings or []))

    # a(ii) Actual construction cost = Σ(building_est_cost × completion%)
    prog_map   = {p.get("building_id"): p.get("overall_completion", 0)
                  for p in (construction_progress or [])}
    dev_a2_inc = sum(
        b.get("estimated_cost", 0) * prog_map.get(b.get("building_id"), 0) / 100
        for b in (buildings or [])
    )

    # a(iii) On-site expenditure
    #   Estimated: infrastructure_costs + site_expenditure (salaries, consultants, etc.)
    infra_doc  = await db.infrastructure_costs.find_one({"project_id": project_id}, {"_id": 0}) or {}
    infra_est  = infra_doc.get("total_infrastructure_cost", 0)

    site_exp_doc = await db.site_expenditure.find_one({"project_id": project_id}, {"_id": 0}) or {}
    site_exp_est = (
        (site_exp_doc.get("site_development_cost", 0) or 0) +
        (site_exp_doc.get("salaries",              0) or 0) +
        (site_exp_doc.get("consultants_fee",       0) or 0) +
        (site_exp_doc.get("site_overheads",        0) or 0) +
        (site_exp_doc.get("services_cost",         0) or 0) +
        (site_exp_doc.get("machinery_cost",        0) or 0)
    )
    dev_a3_est = infra_est + site_exp_est

    #   Incurred: infrastructure (est × completion%) + actual_site_expenditure
    infra_prog_doc = await db.infrastructure_progress.find_one(
        {"project_id": project_id, "quarter": quarter, "year": year}, {"_id": 0}
    ) or {}
    infra_completion = infra_prog_doc.get("overall_completion", 0)
    infra_inc        = infra_est * infra_completion / 100

    act_site_doc = await db.actual_site_expenditure.find_one(
        {"project_id": project_id, "quarter": quarter, "year": year}, {"_id": 0}
    ) or {}
    dev_a3_inc = infra_inc + act_site_doc.get("total", 0)

    # b. Taxes, cess, fees to statutory authority
    est_dev_doc = await db.estimated_development_costs.find_one(
        {"project_id": project_id}, {"_id": 0}
    ) or {}
    dev_b_est = est_dev_doc.get("taxes_premiums_fees", 0)
    dev_b_inc = pc.get("taxes_statutory", 0)

    # c. Finance cost
    dev_c_est = est_dev_doc.get("finance_cost", 0)
    dev_c_inc = pc.get("finance_cost", 0)

    # Sub-totals
    dev_sub_est  = dev_a1_est + dev_a3_est + dev_b_est + dev_c_est
    dev_a_for_min = min(dev_a1_est, dev_a2_inc) if dev_a1_est and dev_a2_inc else (dev_a2_inc or 0)
    dev_sub_inc  = dev_a_for_min + dev_a3_inc + dev_b_inc + dev_c_inc

    # ── SUMMARY CALCULATIONS ─────────────────────────────────────────────────
    total_est      = land_sub_est + dev_sub_est
    total_inc      = land_sub_inc + dev_sub_inc

    # % Completion = Total Actual Dev Cost / Total Estimated Dev Cost
    arch_pct       = (dev_sub_inc / dev_sub_est) if dev_sub_est else 0

    proportion     = (total_inc / total_est) if total_est else 0
    withdraw_allow = total_inc   # same as total_est * proportion by definition
    withdrawn_td   = pc.get("total_amount_withdrawn_till_date", 0)
    net_withdraw   = withdraw_allow - withdrawn_td

    # ── ADDITIONAL INFORMATION (ongoing projects) ────────────────────────────
    bal_cost      = total_est - total_inc

    # Compute receivables directly from sales data
    sold_sales_list   = [s for s in (sales or []) if s.get("buyer_name")]
    unsold_sales_list = [s for s in (sales or []) if not s.get("buyer_name")]

    # Balance receivable from sold units = sum of (sale_value - amount_received)
    bal_recv_sold = sum(
        (s.get("sale_value", 0) or 0) - (s.get("amount_received", 0) or 0)
        for s in sold_sales_list
    )

    # Total area of unsold units
    unsold_area = sum((s.get("carpet_area", 0) or 0) for s in unsold_sales_list)

    # Average sale price per sq.m. from sold units
    total_sale_val_sold   = sum((s.get("sale_value", 0) or 0)  for s in sold_sales_list)
    total_carpet_area_sold = sum((s.get("carpet_area", 0) or 0) for s in sold_sales_list)
    avg_sale_price = (total_sale_val_sold / total_carpet_area_sold) if total_carpet_area_sold else 0

    # Estimated value of unsold units = avg sale price × unsold area
    unsold_val = avg_sale_price * unsold_area

    # Total estimated receivables = total sale value of sold units + unsold estimated value
    total_recv = total_sale_val_sold + unsold_val

    # Also keep ASR rate from financial_summaries if available (for display reference)
    fs = await db.financial_summaries.find_one(
        {"project_id": project_id, "quarter": quarter, "year": year}, {"_id": 0}
    )
    if not fs:
        fs = await db.financial_summaries.find_one(
            {"project_id": project_id}, {"_id": 0},
            sort=[("year", -1), ("quarter", -1)]
        )
    fs = fs or {}
    asr_rate = fs.get("asr_rate_per_sqm", 0)

    deposit_pct   = 0.70 if total_recv > bal_cost else 1.00
    deposit_amt   = total_recv * deposit_pct

    building_map  = {b.get("building_id"): b.get("building_name", "") for b in (buildings or [])}

    return {
        # Land cost items
        "lc_a_est": lc_a_est, "lc_a_inc": lc_a_inc,
        "lc_b_est": lc_b_est, "lc_b_inc": lc_b_inc,
        "lc_c_est": lc_c_est, "lc_c_inc": lc_c_inc,
        "lc_d_est": lc_d_est, "lc_d_inc": lc_d_inc,
        "lc_e_est": lc_e_est, "lc_e_inc": lc_e_inc,
        "rehab_i_est": rehab_i_est, "rehab_i_inc": rehab_i_inc,
        "rehab_ii_inc": rehab_ii_inc,
        "rehab_iii_inc": rehab_iii_inc,
        "rehab_iv_inc": rehab_iv_inc,
        "rehab_any": rehab_any,
        "land_sub_est": land_sub_est, "land_sub_inc": land_sub_inc,
        # Development cost items
        "dev_a1_est": dev_a1_est,
        "dev_a2_inc": dev_a2_inc,
        "dev_a3_est": dev_a3_est,
        "dev_a3_inc": dev_a3_inc,
        "dev_b_est": dev_b_est, "dev_b_inc": dev_b_inc,
        "dev_c_est": dev_c_est, "dev_c_inc": dev_c_inc,
        "dev_sub_est": dev_sub_est, "dev_sub_inc": dev_sub_inc,
        # Summary
        "total_est": round(total_est, 2),
        "total_inc": round(total_inc, 2),
        "arch_pct": arch_pct,
        "proportion": proportion,
        "withdraw_allow": round(withdraw_allow, 2),
        "withdrawn_td": round(withdrawn_td, 2),
        "net_withdraw": round(net_withdraw, 2),
        # Additional info
        "bal_cost": round(bal_cost, 2),
        "bal_recv_sold": round(bal_recv_sold, 2),
        "unsold_area": unsold_area,
        "asr_rate": asr_rate,
        "avg_sale_price": round(avg_sale_price, 2),
        "total_sale_val_sold": round(total_sale_val_sold, 2),
        "unsold_val": round(unsold_val, 2),
        "total_recv": round(total_recv, 2),
        "deposit_pct": deposit_pct,
        "deposit_amt": round(deposit_amt, 2),
        # For Annexure A
        "sales": sales or [],
        "buildings": buildings or [],
        "building_map": building_map,
    }


@api_router.get("/project-costs/live-summary/{project_id}")
async def get_project_cost_live_summary(
    project_id: str, current_user: dict = Depends(get_current_user)
):
    """Return cost totals computed the same way as the Project Costs page."""
    return await _compute_cost_summary(project_id)


@api_router.post("/project-costs", response_model=ProjectCostResponse)
async def create_project_cost(cost: ProjectCostCreate, current_user: dict = Depends(get_current_user)):
    # Calculate Land Cost totals (Section 1.i in Form-4)
    total_land = (
        cost.land_acquisition_cost + cost.land_legal_cost + cost.land_interest_cost +
        cost.development_rights_premium + cost.tdr_cost + cost.stamp_duty + 
        cost.government_charges + cost.land_premium_redevelopment +
        cost.rehab_construction_cost + cost.rehab_transit_accommodation + 
        cost.rehab_clearance_cost + cost.rehab_asr_premium
    )
    
    total_land_estimated = (
        cost.land_acquisition_estimated + cost.development_rights_estimated +
        cost.tdr_estimated + cost.stamp_duty_estimated + cost.government_charges_estimated +
        cost.land_premium_estimated + cost.rehab_construction_estimated
    )
    
    # Calculate Development/Construction Cost totals (Section 1.ii in Form-4)
    total_dev = (
        cost.construction_cost_actual + cost.onsite_salaries + cost.onsite_consultants_fees +
        cost.onsite_site_overheads + cost.onsite_services_cost + cost.onsite_machinery_equipment +
        cost.onsite_consumables + cost.offsite_expenditure + cost.taxes_statutory + cost.finance_cost +
        cost.extra_items_cost
    )
    
    total_dev_estimated = cost.construction_cost_estimated + cost.taxes_statutory_estimated + cost.finance_cost_estimated
    
    # Legacy totals for backward compatibility
    total_estimated = total_land_estimated + total_dev_estimated
    if total_estimated == 0:
        total_estimated = cost.estimated_land_cost + cost.estimated_development_cost
    
    total_incurred = total_land + total_dev
    balance = total_estimated - total_incurred
    
    # Cost completion percentage
    cost_completion = (total_incurred / total_estimated * 100) if total_estimated > 0 else 0
    
    cost_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    cost_doc = {
        "cost_id": cost_id,
        **cost.model_dump(),
        "total_land_cost": total_land,
        "total_land_cost_estimated": total_land_estimated,
        "total_development_cost": total_dev,
        "total_development_cost_estimated": total_dev_estimated,
        "total_estimated_cost": total_estimated,
        "total_cost_incurred": total_incurred,
        "balance_cost": balance,
        "cost_completion_percentage": round(cost_completion, 2),
        "created_at": now
    }
    
    # Upsert
    await db.project_costs.update_one(
        {"project_id": cost.project_id, "quarter": cost.quarter, "year": cost.year},
        {"$set": cost_doc},
        upsert=True
    )
    return ProjectCostResponse(**cost_doc)

@api_router.get("/project-costs", response_model=List[ProjectCostResponse])
async def get_project_costs(
    project_id: str = Query(...),
    quarter: Optional[str] = None,
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"project_id": project_id}
    if quarter:
        query["quarter"] = quarter
    if year:
        query["year"] = year
    
    costs = await db.project_costs.find(query, {"_id": 0}).to_list(1000)
    return [ProjectCostResponse(**c) for c in costs]

@api_router.get("/project-costs/latest/{project_id}", response_model=ProjectCostResponse)
async def get_latest_project_cost(project_id: str, current_user: dict = Depends(get_current_user)):
    cost = await db.project_costs.find_one(
        {"project_id": project_id},
        {"_id": 0},
        sort=[("year", -1), ("quarter", -1)]
    )
    if not cost:
        raise HTTPException(status_code=404, detail="No cost data found")
    return ProjectCostResponse(**cost)

# =========================
# BUILDING COST ROUTES
# =========================

@api_router.post("/building-costs", response_model=BuildingCostResponse)
async def create_building_cost(cost: BuildingCostCreate, current_user: dict = Depends(get_current_user)):
    building = await db.buildings.find_one({"building_id": cost.building_id})
    if not building:
        raise HTTPException(status_code=404, detail="Building not found")
    
    completion = (cost.cost_incurred / cost.estimated_cost * 100) if cost.estimated_cost > 0 else 0
    balance = cost.estimated_cost - cost.cost_incurred
    
    cost_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    cost_doc = {
        "building_cost_id": cost_id,
        "project_id": building["project_id"],
        **cost.model_dump(),
        "completion_percentage": round(completion, 2),
        "balance_cost": balance,
        "created_at": now
    }
    
    await db.building_costs.update_one(
        {"building_id": cost.building_id, "quarter": cost.quarter, "year": cost.year},
        {"$set": cost_doc},
        upsert=True
    )
    return BuildingCostResponse(**cost_doc)

@api_router.get("/building-costs", response_model=List[BuildingCostResponse])
async def get_building_costs(
    project_id: str = Query(...),
    quarter: Optional[str] = None,
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"project_id": project_id}
    if quarter:
        query["quarter"] = quarter
    if year:
        query["year"] = year
    
    costs = await db.building_costs.find(query, {"_id": 0}).to_list(1000)
    return [BuildingCostResponse(**c) for c in costs]

# =========================
# UNIT SALES ROUTES (Annexure-A)
# =========================

@api_router.post("/unit-sales", response_model=UnitSaleResponse)
async def create_unit_sale(sale: UnitSaleCreate, current_user: dict = Depends(get_current_user)):
    sale_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    balance = sale.sale_value - sale.amount_received
    
    sale_doc = {
        "sale_id": sale_id,
        **sale.model_dump(),
        "balance_receivable": balance,
        "created_at": now
    }
    await db.unit_sales.insert_one(sale_doc)
    return UnitSaleResponse(**{k: v for k, v in sale_doc.items() if k != "_id"})

@api_router.get("/unit-sales", response_model=List[UnitSaleResponse])
async def get_unit_sales(project_id: str = Query(...), current_user: dict = Depends(get_current_user)):
    sales = await db.unit_sales.find({"project_id": project_id}, {"_id": 0}).to_list(10000)
    return [UnitSaleResponse(**s) for s in sales]

@api_router.put("/unit-sales/{sale_id}", response_model=UnitSaleResponse)
async def update_unit_sale(sale_id: str, sale: UnitSaleBase, current_user: dict = Depends(get_current_user)):
    existing = await db.unit_sales.find_one({"sale_id": sale_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Unit sale not found")
    
    balance = sale.sale_value - sale.amount_received
    update_data = sale.model_dump()
    update_data["balance_receivable"] = balance
    
    await db.unit_sales.update_one({"sale_id": sale_id}, {"$set": update_data})
    updated = await db.unit_sales.find_one({"sale_id": sale_id}, {"_id": 0})
    return UnitSaleResponse(**updated)

@api_router.delete("/unit-sales/{sale_id}")
async def delete_unit_sale(sale_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.unit_sales.delete_one({"sale_id": sale_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Unit sale not found")
    return {"message": "Unit sale deleted"}

# =========================
# COMMON DEVELOPMENT WORKS (FORM-1 Table B)
# =========================

@api_router.post("/common-development-works", response_model=CommonDevelopmentWorksResponse)
async def create_common_development_works(works: CommonDevelopmentWorksCreate, current_user: dict = Depends(get_current_user)):
    works_data = works.works.model_dump()
    
    # Calculate overall completion based on proposed works
    total_weight = 0
    weighted_completion = 0
    
    work_items = [
        ("internal_roads_footpaths", "internal_roads_proposed"),
        ("water_supply", "water_supply_proposed"),
        ("sewerage", "sewerage_proposed"),
        ("storm_water_drains", "storm_water_proposed"),
        ("landscaping_trees", "landscaping_proposed"),
        ("street_lighting", "street_lighting_proposed"),
        ("community_buildings", "community_buildings_proposed"),
        ("sewage_treatment", "sewage_treatment_proposed"),
        ("solid_waste_management", "solid_waste_proposed"),
        ("rainwater_harvesting", "rainwater_proposed"),
        ("energy_management", "energy_management_proposed"),
        ("fire_safety", "fire_safety_proposed"),
        ("electrical_infrastructure", "electrical_proposed"),
    ]
    
    for completion_field, proposed_field in work_items:
        if works_data.get(proposed_field, False):
            total_weight += 1
            weighted_completion += works_data.get(completion_field, 0)
    
    overall = (weighted_completion / total_weight) if total_weight > 0 else 0
    
    works_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    works_doc = {
        "works_id": works_id,
        "project_id": works.project_id,
        "quarter": works.quarter,
        "year": works.year,
        "works": works_data,
        "overall_completion": round(overall, 2),
        "created_at": now
    }
    
    await db.common_development_works.update_one(
        {"project_id": works.project_id, "quarter": works.quarter, "year": works.year},
        {"$set": works_doc},
        upsert=True
    )
    
    return CommonDevelopmentWorksResponse(**works_doc)

@api_router.get("/common-development-works", response_model=List[CommonDevelopmentWorksResponse])
async def get_common_development_works(
    project_id: str = Query(...),
    quarter: Optional[str] = None,
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"project_id": project_id}
    if quarter:
        query["quarter"] = quarter
    if year:
        query["year"] = year
    
    works = await db.common_development_works.find(query, {"_id": 0}).to_list(100)
    return [CommonDevelopmentWorksResponse(**w) for w in works]

@api_router.get("/common-development-works/latest/{project_id}", response_model=CommonDevelopmentWorksResponse)
async def get_latest_common_development_works(project_id: str, current_user: dict = Depends(get_current_user)):
    works = await db.common_development_works.find_one(
        {"project_id": project_id},
        {"_id": 0},
        sort=[("year", -1), ("quarter", -1)]
    )
    if not works:
        raise HTTPException(status_code=404, detail="No common development works data found")
    return CommonDevelopmentWorksResponse(**works)

# =========================
# FINANCIAL SUMMARY (FORM-5)
# =========================

@api_router.post("/financial-summary", response_model=FinancialSummaryResponse)
async def create_financial_summary(summary: FinancialSummaryCreate, current_user: dict = Depends(get_current_user)):
    # Calculate unsold inventory value
    unsold_value = summary.unsold_area_sqm * summary.asr_rate_per_sqm
    
    # Calculate total estimated receivables
    total_receivables = summary.total_balance_receivable_sold + unsold_value
    
    # Determine deposit percentage and amount (Form-5 logic)
    # Get project cost data
    cost = await db.project_costs.find_one(
        {"project_id": summary.project_id},
        {"_id": 0},
        sort=[("year", -1), ("quarter", -1)]
    )
    
    balance_cost = cost.get("balance_cost", 0) if cost else 0
    
    # If receivables > balance cost, deposit 70%, else 100%
    deposit_pct = 70 if total_receivables > balance_cost else 100
    amount_to_deposit = total_receivables * deposit_pct / 100
    
    summary_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    
    summary_doc = {
        "summary_id": summary_id,
        **summary.model_dump(),
        "unsold_inventory_value": unsold_value,
        "total_estimated_receivables": total_receivables,
        "deposit_percentage": deposit_pct,
        "amount_to_deposit": amount_to_deposit,
        "created_at": now
    }
    
    await db.financial_summaries.update_one(
        {"project_id": summary.project_id, "quarter": summary.quarter, "year": summary.year},
        {"$set": summary_doc},
        upsert=True
    )
    
    return FinancialSummaryResponse(**summary_doc)

@api_router.get("/financial-summary", response_model=List[FinancialSummaryResponse])
async def get_financial_summaries(
    project_id: str = Query(...),
    quarter: Optional[str] = None,
    year: Optional[int] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"project_id": project_id}
    if quarter:
        query["quarter"] = quarter
    if year:
        query["year"] = year
    
    summaries = await db.financial_summaries.find(query, {"_id": 0}).to_list(100)
    return [FinancialSummaryResponse(**s) for s in summaries]

@api_router.get("/financial-summary/latest/{project_id}", response_model=FinancialSummaryResponse)
async def get_latest_financial_summary(project_id: str, current_user: dict = Depends(get_current_user)):
    summary = await db.financial_summaries.find_one(
        {"project_id": project_id},
        {"_id": 0},
        sort=[("year", -1), ("quarter", -1)]
    )
    if not summary:
        raise HTTPException(status_code=404, detail="No financial summary data found")
    return FinancialSummaryResponse(**summary)

@api_router.post("/unit-sales/bulk", response_model=Dict[str, Any])
async def bulk_create_unit_sales(
    project_id: str = Query(...),
    sales: List[UnitSaleBase] = None,
    current_user: dict = Depends(get_current_user)
):
    if not sales:
        raise HTTPException(status_code=400, detail="No sales data provided")
    
    created = 0
    errors = []
    for idx, sale in enumerate(sales):
        try:
            sale_id = str(uuid.uuid4())
            now = datetime.now(timezone.utc).isoformat()
            balance = sale.sale_value - sale.amount_received
            
            sale_doc = {
                "sale_id": sale_id,
                "project_id": project_id,
                **sale.model_dump(),
                "balance_receivable": balance,
                "created_at": now
            }
            await db.unit_sales.insert_one(sale_doc)
            created += 1
        except Exception as e:
            errors.append({"row": idx + 1, "error": str(e)})
    
    return {"created": created, "errors": errors}

# =========================
# EXCEL IMPORT
# =========================

@api_router.post("/import/sales-excel")
async def import_sales_excel(
    project_id: str = Query(...),
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Import unit sales from Excel file - replaces all existing data for the project"""
    import openpyxl
    
    contents = await file.read()
    wb = openpyxl.load_workbook(BytesIO(contents))
    ws = wb.active
    
    # Find header row
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            headers[str(cell_value).lower().strip()] = col
    
    # Map columns
    column_map = {
        "unit_number": ["unit number", "unit no", "unit no.", "flat no", "flat number"],
        "building_name": ["building name", "building", "wing", "tower"],
        "carpet_area": ["carpet area", "area", "carpet area (sq ft)", "carpet area (sqm)"],
        "sale_value": ["sale value", "agreement value", "total value", "price"],
        "amount_received": ["amount received", "received", "collection", "amount collected"],
        "buyer_name": ["buyer name", "buyer", "customer name", "allottee"],
        "agreement_date": ["agreement date", "date", "booking date"]
    }
    
    def find_column(field):
        for alias in column_map.get(field, []):
            if alias in headers:
                return headers[alias]
        return None
    
    # DELETE all existing sales for this project before importing new data
    delete_result = await db.unit_sales.delete_many({"project_id": project_id})
    deleted_count = delete_result.deleted_count
    
    created = 0
    unsold_count = 0
    errors = []
    
    # Get buildings for this project to map building names to IDs
    buildings = await db.buildings.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    building_map = {b["building_name"].lower(): b["building_id"] for b in buildings}
    
    for row in range(2, ws.max_row + 1):
        try:
            unit_col = find_column("unit_number")
            building_col = find_column("building_name")
            area_col = find_column("carpet_area")
            value_col = find_column("sale_value")
            received_col = find_column("amount_received")
            buyer_col = find_column("buyer_name")
            date_col = find_column("agreement_date")
            
            unit_number = ws.cell(row=row, column=unit_col).value if unit_col else None
            building_name = ws.cell(row=row, column=building_col).value if building_col else None
            
            if not unit_number:
                continue
            
            carpet_area = float(ws.cell(row=row, column=area_col).value or 0) if area_col else 0
            sale_value = float(ws.cell(row=row, column=value_col).value or 0) if value_col else 0
            amount_received = float(ws.cell(row=row, column=received_col).value or 0) if received_col else 0
            buyer_name = str(ws.cell(row=row, column=buyer_col).value or "").strip() if buyer_col else ""
            agreement_date = ws.cell(row=row, column=date_col).value if date_col else None
            
            if agreement_date and hasattr(agreement_date, 'isoformat'):
                agreement_date = agreement_date.isoformat()
            elif agreement_date:
                agreement_date = str(agreement_date)
            
            # Find building_id
            building_id = building_map.get(str(building_name).lower(), "") if building_name else ""
            
            sale_id = str(uuid.uuid4())
            now = datetime.now(timezone.utc).isoformat()
            balance = sale_value - amount_received
            
            # Determine sale status: if buyer_name is blank, it's unsold inventory
            is_sold = bool(buyer_name)
            status = "sold" if is_sold else "unsold"
            
            if not is_sold:
                unsold_count += 1
            
            sale_doc = {
                "sale_id": sale_id,
                "project_id": project_id,
                "unit_number": str(unit_number),
                "building_id": building_id,
                "building_name": str(building_name or ""),
                "carpet_area": carpet_area,
                "sale_value": sale_value,
                "amount_received": amount_received,
                "buyer_name": buyer_name,
                "agreement_date": agreement_date,
                "balance_receivable": balance,
                "status": status,  # "sold" or "unsold"
                "created_at": now,
                "updated_at": now
            }
            await db.unit_sales.insert_one(sale_doc)
            created += 1
        except Exception as e:
            errors.append({"row": row, "error": str(e)})
    
    return {
        "deleted": deleted_count,
        "created": created,
        "sold_units": created - unsold_count,
        "unsold_units": unsold_count,
        "errors": errors,
        "total_rows": ws.max_row - 1,
        "message": f"Replaced {deleted_count} existing records with {created} new records ({unsold_count} unsold units)"
    }

@api_router.get("/import/sales-template")
async def get_sales_template():
    """Download Excel template for sales import"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import tempfile
    import os
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Data"
    
    headers = ["Unit Number", "Building Name", "Carpet Area", "Sale Value", "Amount Received", "Buyer Name", "Agreement Date"]
    header_fill = PatternFill(start_color="172554", end_color="172554", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 18
    
    # Add sample data
    sample_data = [
        ["A-101", "Tower A", 850, 7500000, 5000000, "John Doe", "2024-01-15"],
        ["A-102", "Tower A", 920, 8200000, 4100000, "Jane Smith", "2024-02-20"],
    ]
    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save to temp file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    
    return FileResponse(
        path=temp_file.name,
        filename="CheckMate - Sales Data.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =========================
# DASHBOARD & SUMMARY
# =========================

@api_router.get("/dashboard/{project_id}", response_model=DashboardSummary)
async def get_dashboard(project_id: str, current_user: dict = Depends(get_current_user)):
    # Cost summary – computed via shared helper (same logic as Project Costs page)
    cost_summary = await _compute_cost_summary(project_id)

    # Get all unit sales
    sales = await db.unit_sales.find({"project_id": project_id}, {"_id": 0}).to_list(10000)

    # Get buildings
    buildings = await db.buildings.find({"project_id": project_id}, {"_id": 0}).to_list(1000)

    # Get construction progress
    progress_list = await db.construction_progress.find(
        {"project_id": project_id},
        {"_id": 0},
        sort=[("year", -1), ("quarter", -1)]
    ).to_list(1000)
    
    # Calculate totals
    # Only include sold units in financial calculations
    sold_sales = [s for s in sales if s.get("status") == "sold" or s.get("is_sold", False)]
    total_sales_value = sum(s.get("sale_value", 0) for s in sold_sales)
    amount_collected = sum(s.get("amount_received", 0) for s in sold_sales)
    receivables = sum(s.get("balance_receivable", 0) for s in sold_sales)
    units_sold = len(sold_sales)

    # Inventory reconciliation: compare building-config total vs sales-data total
    building_config_units = sum(b.get("units", 0) for b in buildings)
    sales_data_units = len(sales)  # all rows: sold + unsold
    inventory_mismatch = (building_config_units != sales_data_units) and (building_config_units > 0 or sales_data_units > 0)
    inventory_mismatch_delta = sales_data_units - building_config_units

    # total_units: use sales_data_units (most complete source) when available,
    # else fall back to building configuration
    total_units = sales_data_units if sales else building_config_units
    
    # Calculate unsold inventory
    unsold_units = total_units - units_sold
    avg_unit_value = total_sales_value / units_sold if units_sold > 0 else 0
    unsold_inventory_value = unsold_units * avg_unit_value
    
    # Calculate overall completion
    if progress_list:
        # Get latest progress for each building
        building_progress = {}
        for p in progress_list:
            bid = p.get("building_id")
            if bid not in building_progress:
                building_progress[bid] = p.get("overall_completion", 0)
        overall_completion = sum(building_progress.values()) / len(building_progress) if building_progress else 0
    else:
        overall_completion = 0
    
    # ── COST DATA (from shared helper – same logic as Project Costs page) ───
    total_estimated = cost_summary["total_estimated_cost"]
    cost_incurred   = cost_summary["total_cost_incurred"]
    balance_cost    = cost_summary["balance_cost"]
    
    # RERA deposit calculation (Form-5 logic)
    if balance_cost > 0 and receivables > 0:
        ratio = receivables / balance_cost
        if ratio > 1:
            rera_deposit = receivables * 0.7
        else:
            rera_deposit = receivables
    else:
        rera_deposit = 0
    
    return DashboardSummary(
        project_completion_percentage=round(overall_completion, 2),
        total_estimated_cost=total_estimated,
        cost_incurred=cost_incurred,
        balance_cost=balance_cost,
        total_sales_value=total_sales_value,
        amount_collected=amount_collected,
        receivables=receivables,
        unsold_inventory_value=unsold_inventory_value,
        rera_deposit_required=rera_deposit,
        total_units=total_units,
        units_sold=units_sold,
        building_config_units=building_config_units,
        sales_data_units=sales_data_units,
        inventory_mismatch=inventory_mismatch,
        inventory_mismatch_delta=inventory_mismatch_delta
    )

# =========================
# QUARTERLY REPORTS
# =========================

@api_router.post("/quarterly-reports", response_model=QuarterlyReportResponse)
async def create_quarterly_report(report: QuarterlyReportCreate, current_user: dict = Depends(get_current_user)):
    report_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    report_doc = {
        "report_id": report_id,
        **report.model_dump(),
        "created_by": current_user["user_id"],
        "created_at": now
    }
    await db.quarterly_reports.insert_one(report_doc)
    return QuarterlyReportResponse(**{k: v for k, v in report_doc.items() if k != "_id"})

@api_router.get("/quarterly-reports", response_model=List[QuarterlyReportResponse])
async def get_quarterly_reports(project_id: str = Query(...), current_user: dict = Depends(get_current_user)):
    reports = await db.quarterly_reports.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    return [QuarterlyReportResponse(**r) for r in reports]

# =========================
# REPORT TEMPLATES
# =========================

@api_router.post("/report-templates", response_model=ReportTemplateResponse)
async def create_report_template(template: ReportTemplateCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    template_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    template_doc = {
        "template_id": template_id,
        **template.model_dump(),
        "created_at": now,
        "updated_at": now
    }
    await db.report_templates.insert_one(template_doc)
    return ReportTemplateResponse(**{k: v for k, v in template_doc.items() if k != "_id"})

@api_router.get("/report-templates", response_model=List[ReportTemplateResponse])
async def get_report_templates(state: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if state:
        query["state"] = state
    templates = await db.report_templates.find(query, {"_id": 0}).to_list(1000)
    return [ReportTemplateResponse(**t) for t in templates]

@api_router.get("/report-templates/{template_id}", response_model=ReportTemplateResponse)
async def get_report_template(template_id: str, current_user: dict = Depends(get_current_user)):
    template = await db.report_templates.find_one({"template_id": template_id}, {"_id": 0})
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    return ReportTemplateResponse(**template)

@api_router.put("/report-templates/{template_id}", response_model=ReportTemplateResponse)
async def update_report_template(template_id: str, template: ReportTemplateCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    existing = await db.report_templates.find_one({"template_id": template_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Template not found")
    
    update_data = template.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.report_templates.update_one({"template_id": template_id}, {"$set": update_data})
    
    updated = await db.report_templates.find_one({"template_id": template_id}, {"_id": 0})
    return ReportTemplateResponse(**updated)

# =========================
# REPORT GENERATION
# =========================

@api_router.get("/generate-report/{project_id}/{report_type}")
async def generate_report(
    project_id: str,
    report_type: str,
    quarter: str = Query(...),
    year: int = Query(...),
    current_user: dict = Depends(get_current_user)
):
    """Generate RERA report as HTML (for PDF conversion on frontend)"""
    
    # Get project
    project = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Get template
    template = await db.report_templates.find_one(
        {"state": project["state"], "report_type": report_type},
        {"_id": 0}
    )
    
    # Get related data
    buildings = await db.buildings.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    progress_list = await db.construction_progress.find(
        {"project_id": project_id, "quarter": quarter, "year": year},
        {"_id": 0}
    ).to_list(1000)
    
    cost = await db.project_costs.find_one(
        {"project_id": project_id, "quarter": quarter, "year": year},
        {"_id": 0}
    )
    
    building_costs = await db.building_costs.find(
        {"project_id": project_id, "quarter": quarter, "year": year},
        {"_id": 0}
    ).to_list(1000)
    
    sales = await db.unit_sales.find({"project_id": project_id}, {"_id": 0}).to_list(10000)
    
    # Calculate summary
    total_sales_value = sum(s.get("sale_value", 0) for s in sales)
    amount_collected = sum(s.get("amount_received", 0) for s in sales)
    receivables = sum(s.get("balance_receivable", 0) for s in sales)
    
    # Build data context
    data = {
        "project": project,
        "buildings": buildings,
        "construction_progress": progress_list,
        "project_cost": cost or {},
        "building_costs": building_costs,
        "sales": sales,
        "quarter": quarter,
        "year": year,
        "report_date": datetime.now().strftime("%d/%m/%Y"),
        "total_sales_value": total_sales_value,
        "amount_collected": amount_collected,
        "receivables": receivables,
        "balance_cost": cost.get("balance_cost", 0) if cost else 0
    }
    
    # If template exists, use it
    if template:
        html = template["template_html"]
        # Replace placeholders
        for key, value in flatten_dict(data).items():
            placeholder = "{{" + key + "}}"
            html = html.replace(placeholder, str(value) if value else "")
        return {"html": html, "data": data}
    
    # Return raw data for frontend rendering
    return {"html": None, "data": data}


# =========================
# PDF GENERATION ENDPOINT
# =========================

@api_router.get("/generate-pdf/{project_id}/{report_type}")
async def generate_pdf_report(
    project_id: str,
    report_type: str,
    quarter: str = Query(...),
    year: int = Query(...),
    current_user: dict = Depends(get_current_user)
):
    """Generate RERA report as downloadable PDF"""
    from pdf_generator import (
        generate_form1_pdf, 
        generate_form3_pdf, 
        generate_form4_pdf, 
        generate_annexure_a_pdf
    )
    
    # Get project
    project = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    # Get related data
    buildings = await db.buildings.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    
    construction_progress = await db.construction_progress.find(
        {"project_id": project_id, "quarter": quarter, "year": year},
        {"_id": 0}
    ).to_list(1000)
    
    infrastructure_progress = await db.infrastructure_progress.find_one(
        {"project_id": project_id, "quarter": quarter, "year": year},
        {"_id": 0}
    )
    
    project_cost = await db.project_costs.find_one(
        {"project_id": project_id, "quarter": quarter, "year": year},
        {"_id": 0}
    )
    
    building_costs = await db.building_costs.find(
        {"project_id": project_id, "quarter": quarter, "year": year},
        {"_id": 0}
    ).to_list(1000)
    
    estimated_dev_cost = await db.estimated_development_costs.find_one(
        {"project_id": project_id},
        {"_id": 0}
    )
    
    sales = await db.unit_sales.find({"project_id": project_id}, {"_id": 0}).to_list(10000)
    
    # Generate PDF based on report type
    try:
        if report_type == "form-1":
            pdf_buffer = generate_form1_pdf(
                project, buildings, construction_progress, 
                infrastructure_progress, quarter, year
            )
            filename = f"Form1_Construction_Progress_{project.get('project_name', 'Project')}_{quarter}_{year}.pdf"
            
        elif report_type == "form-3":
            pdf_buffer = generate_form3_pdf(
                project, buildings, construction_progress,
                infrastructure_progress, estimated_dev_cost, quarter, year
            )
            filename = f"Form3_Cost_Incurred_{project.get('project_name', 'Project')}_{quarter}_{year}.pdf"
            
        elif report_type == "form-4":
            form4_data = await _build_form4_data(
                project_id, quarter, year,
                buildings, construction_progress, project_cost, sales
            )
            pdf_buffer = generate_form4_pdf(project, form4_data, quarter, year)
            filename = f"Form4_CA_Certificate_{project.get('project_name', 'Project')}_{quarter}_{year}.pdf"
            
        elif report_type == "annexure-a":
            pdf_buffer = generate_annexure_a_pdf(
                project, sales, buildings, quarter, year
            )
            filename = f"AnnexureA_Sales_{project.get('project_name', 'Project')}_{quarter}_{year}.pdf"
            
        else:
            raise HTTPException(
                status_code=400, 
                detail=f"PDF generation not yet available for {report_type}. Use HTML preview instead."
            )
        
        # Clean filename
        filename = filename.replace(" ", "_").replace("/", "-")
        
        return StreamingResponse(
            pdf_buffer,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Type": "application/pdf"
            }
        )
        
    except HTTPException:
        raise  # Re-raise HTTP exceptions as-is
    except Exception as e:
        logger.error(f"PDF generation error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate PDF: {str(e)}")


# =========================
# EXCEL GENERATION ENDPOINT
# =========================

@api_router.get("/generate-excel/{project_id}/{report_type}")
async def generate_excel_report(
    project_id: str,
    report_type: str,
    quarter: str = Query(...),
    year: int = Query(...),
    current_user: dict = Depends(get_current_user)
):
    """Generate RERA report as downloadable Excel (.xlsx)"""
    from excel_generator import (
        generate_form1_excel,
        generate_form3_excel,
        generate_form4_excel,
        generate_annexure_a_excel,
    )

    project = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    buildings = await db.buildings.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    construction_progress = await db.construction_progress.find(
        {"project_id": project_id, "quarter": quarter, "year": year}, {"_id": 0}
    ).to_list(1000)
    infrastructure_progress = await db.infrastructure_progress.find_one(
        {"project_id": project_id, "quarter": quarter, "year": year}, {"_id": 0}
    )
    project_cost = await db.project_costs.find_one(
        {"project_id": project_id, "quarter": quarter, "year": year}, {"_id": 0}
    )
    building_costs = await db.building_costs.find(
        {"project_id": project_id, "quarter": quarter, "year": year}, {"_id": 0}
    ).to_list(1000)
    estimated_dev_cost = await db.estimated_development_costs.find_one(
        {"project_id": project_id}, {"_id": 0}
    )
    sales = await db.unit_sales.find({"project_id": project_id}, {"_id": 0}).to_list(10000)

    try:
        if report_type == "form-1":
            buf = generate_form1_excel(project, buildings, construction_progress, infrastructure_progress, quarter, year)
            filename = f"Form1_Construction_Progress_{project.get('project_name', 'Project')}_{quarter}_{year}.xlsx"
        elif report_type == "form-3":
            buf = generate_form3_excel(project, buildings, construction_progress, infrastructure_progress, estimated_dev_cost, quarter, year)
            filename = f"Form3_Cost_Incurred_{project.get('project_name', 'Project')}_{quarter}_{year}.xlsx"
        elif report_type == "form-4":
            form4_data = await _build_form4_data(
                project_id, quarter, year,
                buildings, construction_progress, project_cost, sales
            )
            buf = generate_form4_excel(project, form4_data, quarter, year)
            filename = f"Form4_CA_Certificate_{project.get('project_name', 'Project')}_{quarter}_{year}.xlsx"
        elif report_type == "annexure-a":
            buf = generate_annexure_a_excel(project, sales, buildings, quarter, year)
            filename = f"AnnexureA_Sales_{project.get('project_name', 'Project')}_{quarter}_{year}.xlsx"
        else:
            raise HTTPException(status_code=400, detail=f"Excel generation not available for {report_type}.")

        filename = filename.replace(" ", "_").replace("/", "-")
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Excel generation error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate Excel: {str(e)}")


# =========================
# WORD GENERATION ENDPOINT
# =========================

@api_router.get("/generate-docx/{project_id}/{report_type}")
async def generate_docx_report(
    project_id: str,
    report_type: str,
    quarter: str = Query(...),
    year: int = Query(...),
    current_user: dict = Depends(get_current_user)
):
    """Generate RERA report as downloadable Word document (.docx)"""
    from docx_generator import (
        generate_form1_docx,
        generate_form3_docx,
        generate_form4_docx,
        generate_annexure_a_docx,
    )

    project = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    buildings = await db.buildings.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    construction_progress = await db.construction_progress.find(
        {"project_id": project_id, "quarter": quarter, "year": year}, {"_id": 0}
    ).to_list(1000)
    infrastructure_progress = await db.infrastructure_progress.find_one(
        {"project_id": project_id, "quarter": quarter, "year": year}, {"_id": 0}
    )
    project_cost = await db.project_costs.find_one(
        {"project_id": project_id, "quarter": quarter, "year": year}, {"_id": 0}
    )
    building_costs = await db.building_costs.find(
        {"project_id": project_id, "quarter": quarter, "year": year}, {"_id": 0}
    ).to_list(1000)
    estimated_dev_cost = await db.estimated_development_costs.find_one(
        {"project_id": project_id}, {"_id": 0}
    )
    sales = await db.unit_sales.find({"project_id": project_id}, {"_id": 0}).to_list(10000)

    try:
        if report_type == "form-1":
            buf = generate_form1_docx(project, buildings, construction_progress, infrastructure_progress, quarter, year)
            filename = f"Form1_Construction_Progress_{project.get('project_name', 'Project')}_{quarter}_{year}.docx"
        elif report_type == "form-3":
            buf = generate_form3_docx(project, buildings, construction_progress, infrastructure_progress, estimated_dev_cost, quarter, year)
            filename = f"Form3_Cost_Incurred_{project.get('project_name', 'Project')}_{quarter}_{year}.docx"
        elif report_type == "form-4":
            form4_data = await _build_form4_data(
                project_id, quarter, year,
                buildings, construction_progress, project_cost, sales
            )
            buf = generate_form4_docx(project, form4_data, quarter, year)
            filename = f"Form4_CA_Certificate_{project.get('project_name', 'Project')}_{quarter}_{year}.docx"
        elif report_type == "annexure-a":
            buf = generate_annexure_a_docx(project, sales, buildings, quarter, year)
            filename = f"AnnexureA_Sales_{project.get('project_name', 'Project')}_{quarter}_{year}.docx"
        else:
            raise HTTPException(status_code=400, detail=f"Word generation not available for {report_type}.")

        filename = filename.replace(" ", "_").replace("/", "-")
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Word generation error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to generate Word doc: {str(e)}")


def flatten_dict(d, parent_key='', sep='.'):
    """Flatten nested dict"""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep).items())
        else:
            items.append((new_key, v))
    return dict(items)

# =========================
# DATA VALIDATION
# =========================

@api_router.get("/validate/{project_id}")
async def validate_project_data(project_id: str, current_user: dict = Depends(get_current_user)):
    """Validate project data for RERA compliance"""
    warnings = []
    errors = []
    
    # Get all data
    project = await db.projects.find_one({"project_id": project_id}, {"_id": 0})
    if not project:
        return {"errors": [{"type": "critical", "message": "Project not found"}], "warnings": []}
    
    buildings = await db.buildings.find({"project_id": project_id}, {"_id": 0}).to_list(1000)
    sales = await db.unit_sales.find({"project_id": project_id}, {"_id": 0}).to_list(10000)
    cost = await db.project_costs.find_one({"project_id": project_id}, {"_id": 0}, sort=[("year", -1), ("quarter", -1)])
    
    # Check for missing required fields
    required_fields = ["project_name", "rera_number", "promoter_name", "architect_name", "engineer_name", "ca_name"]
    for field in required_fields:
        if not project.get(field):
            warnings.append({"type": "missing_field", "message": f"Missing required field: {field}"})
    
    # Check for duplicate units
    unit_numbers = [s["unit_number"] for s in sales]
    duplicates = [u for u in set(unit_numbers) if unit_numbers.count(u) > 1]
    if duplicates:
        errors.append({"type": "duplicate", "message": f"Duplicate unit numbers found: {', '.join(duplicates)}"})
    
    # Check for negative receivables
    negative_receivables = [s for s in sales if s.get("balance_receivable", 0) < 0]
    if negative_receivables:
        warnings.append({
            "type": "negative_receivable",
            "message": f"{len(negative_receivables)} units have negative receivables (over-collection)"
        })
    
    # Check completion vs cost mismatch
    if cost:
        total_estimated = cost.get("total_estimated_cost", 0)
        cost_incurred = cost.get("total_cost_incurred", 0)
        
        if cost_incurred > total_estimated:
            warnings.append({
                "type": "cost_overrun",
                "message": f"Cost incurred ({cost_incurred:,.0f}) exceeds estimated cost ({total_estimated:,.0f})"
            })
    
    # Check receivables vs balance cost
    total_receivables = sum(s.get("balance_receivable", 0) for s in sales)
    balance_cost = cost.get("balance_cost", 0) if cost else 0

    if total_receivables > 0 and balance_cost > 0:
        ratio = total_receivables / balance_cost
        if ratio > 1.5:
            warnings.append({
                "type": "receivable_mismatch",
                "message": f"Receivables ({total_receivables:,.0f}) significantly exceed balance cost ({balance_cost:,.0f})"
            })

    # Inventory reconciliation check: building config total vs sales data total
    building_config_units = sum(b.get("units", 0) for b in buildings)
    sales_data_units = len(sales)  # all rows: sold + unsold
    units_sold_count = len([s for s in sales if s.get("status") == "sold" or s.get("is_sold", False)])
    units_unsold_count = sales_data_units - units_sold_count

    if building_config_units != sales_data_units and (building_config_units > 0 or sales_data_units > 0):
        delta = sales_data_units - building_config_units
        direction = "more" if delta > 0 else "fewer"
        errors.append({
            "type": "inventory_mismatch",
            "message": (
                f"Inventory count mismatch: Buildings configuration shows {building_config_units} units, "
                f"but Sales data contains {sales_data_units} units "
                f"({units_sold_count} sold + {units_unsold_count} unsold). "
                f"Sales data has {abs(delta)} {direction} units than the building configuration. "
                f"Please reconcile before generating RERA reports."
            ),
            "building_config_units": building_config_units,
            "sales_data_units": sales_data_units,
            "units_sold": units_sold_count,
            "units_unsold": units_unsold_count,
            "delta": delta
        })

    return {
        "is_valid": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
        "summary": {
            "buildings": len(buildings),
            "building_config_units": building_config_units,
            "sales_data_units": sales_data_units,
            "units_sold": units_sold_count,
            "units_unsold": units_unsold_count,
            "inventory_mismatch": building_config_units != sales_data_units,
            "total_receivables": total_receivables,
            "balance_cost": balance_cost
        }
    }

# =========================
# HEALTH CHECK
# =========================

@api_router.get("/")
async def root():
    return {"message": "CheckMate - RERA Manager API", "version": "1.0.0"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# Include router
app.include_router(api_router)

# CORS
_cors_origins_env = os.environ.get('CORS_ORIGINS', '')
_cors_origins = [o.strip() for o in _cors_origins_env.split(',') if o.strip()] or [
    'https://checkmate-frontend-ei62.onrender.com',
    'http://localhost:3000',
]
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=_cors_origins,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup():
    # Create indexes
    await db.users.create_index("email", unique=True)
    await db.users.create_index("user_id", unique=True)
    await db.projects.create_index("project_id", unique=True)
    await db.buildings.create_index("building_id", unique=True)
    await db.unit_sales.create_index("sale_id", unique=True)
    await db.report_templates.create_index([("state", 1), ("report_type", 1)])
    
    # Initialize default report templates for GOA
    existing = await db.report_templates.find_one({"state": "GOA"})
    if not existing:
        await initialize_goa_templates()
    
    logger.info("CheckMate - RERA Manager API started")

async def initialize_goa_templates():
    """Initialize default Goa RERA report templates"""
    templates = [
        {
            "template_id": str(uuid.uuid4()),
            "state": "GOA",
            "report_name": "Form-1: Architect Certificate",
            "report_type": "form-1",
            "template_html": GOA_FORM_1_TEMPLATE,
            "data_mapping": {},
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "template_id": str(uuid.uuid4()),
            "state": "GOA",
            "report_name": "Form-2: Architect Completion Certificate",
            "report_type": "form-2",
            "template_html": GOA_FORM_2_TEMPLATE,
            "data_mapping": {},
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "template_id": str(uuid.uuid4()),
            "state": "GOA",
            "report_name": "Form-3: Engineer Certificate",
            "report_type": "form-3",
            "template_html": GOA_FORM_3_TEMPLATE,
            "data_mapping": {},
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "template_id": str(uuid.uuid4()),
            "state": "GOA",
            "report_name": "Form-4: CA Certificate",
            "report_type": "form-4",
            "template_html": GOA_FORM_4_TEMPLATE,
            "data_mapping": {},
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "template_id": str(uuid.uuid4()),
            "state": "GOA",
            "report_name": "Form-5: CA Compliance Certificate",
            "report_type": "form-5",
            "template_html": GOA_FORM_5_TEMPLATE,
            "data_mapping": {},
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "template_id": str(uuid.uuid4()),
            "state": "GOA",
            "report_name": "Form-6: Auditor Certificate",
            "report_type": "form-6",
            "template_html": GOA_FORM_6_TEMPLATE,
            "data_mapping": {},
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        },
        {
            "template_id": str(uuid.uuid4()),
            "state": "GOA",
            "report_name": "Annexure-A: Statement of Receivables",
            "report_type": "annexure-a",
            "template_html": GOA_ANNEXURE_A_TEMPLATE,
            "data_mapping": {},
            "created_at": datetime.now(timezone.utc).isoformat(),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
    ]
    
    for template in templates:
        await db.report_templates.insert_one(template)
    
    logger.info("Initialized Goa RERA templates")

# =========================
# GOA RERA TEMPLATES
# =========================

GOA_FORM_1_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
<style>
body { font-family: 'Times New Roman', serif; font-size: 12pt; line-height: 1.6; padding: 40px; }
.header { text-align: center; margin-bottom: 30px; }
.title { font-size: 16pt; font-weight: bold; margin-bottom: 10px; }
.subtitle { font-size: 14pt; margin-bottom: 20px; }
table { width: 100%; border-collapse: collapse; margin: 20px 0; }
th, td { border: 1px solid #000; padding: 8px; text-align: left; }
th { background-color: #f0f0f0; }
.signature { margin-top: 50px; }
.form-section { margin: 20px 0; }
</style>
</head>
<body>
<div class="header">
<div class="title">GOA REAL ESTATE REGULATORY AUTHORITY</div>
<div class="subtitle">FORM - 1</div>
<div class="subtitle">ARCHITECT'S CERTIFICATE</div>
<div>(Percentage of Completion)</div>
</div>

<div class="form-section">
<p><strong>Project Name:</strong> {{project.project_name}}</p>
<p><strong>RERA Registration No:</strong> {{project.rera_number}}</p>
<p><strong>Quarter:</strong> {{quarter}} {{year}}</p>
<p><strong>Report Date:</strong> {{report_date}}</p>
</div>

<div class="form-section">
<p>I, {{project.architect_name}}, Architect, having License No. {{project.architect_license}}, 
hereby certify that the construction of the project mentioned above is progressing as per the sanctioned plans 
and the percentage of completion is as follows:</p>
</div>

<table>
<tr>
<th>Sr. No.</th>
<th>Building/Wing</th>
<th>Activity</th>
<th>Weightage (%)</th>
<th>Completion (%)</th>
<th>Weighted Completion (%)</th>
</tr>
<!-- Building data will be populated here -->
</table>

<div class="signature">
<p>Date: {{report_date}}</p>
<p>Place: {{project.district}}, {{project.state}}</p>
<br><br>
<p>_____________________________</p>
<p>{{project.architect_name}}</p>
<p>Architect</p>
<p>License No: {{project.architect_license}}</p>
</div>
</body>
</html>
"""

GOA_FORM_2_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
<style>
body { font-family: 'Times New Roman', serif; font-size: 12pt; line-height: 1.6; padding: 40px; }
.header { text-align: center; margin-bottom: 30px; }
.title { font-size: 16pt; font-weight: bold; margin-bottom: 10px; }
.subtitle { font-size: 14pt; margin-bottom: 20px; }
table { width: 100%; border-collapse: collapse; margin: 20px 0; }
th, td { border: 1px solid #000; padding: 8px; text-align: left; }
th { background-color: #f0f0f0; }
.signature { margin-top: 50px; }
</style>
</head>
<body>
<div class="header">
<div class="title">GOA REAL ESTATE REGULATORY AUTHORITY</div>
<div class="subtitle">FORM - 2</div>
<div class="subtitle">ARCHITECT'S COMPLETION CERTIFICATE</div>
</div>

<div class="form-section">
<p><strong>Project Name:</strong> {{project.project_name}}</p>
<p><strong>RERA Registration No:</strong> {{project.rera_number}}</p>
</div>

<table>
<tr>
<th>Sr. No.</th>
<th>Building/Wing Name</th>
<th>Completion Cert. No.</th>
<th>Completion Cert. Date</th>
<th>Occupancy Cert. No.</th>
<th>Occupancy Cert. Date</th>
</tr>
<!-- Building completion data -->
</table>

<div class="signature">
<p>Date: {{report_date}}</p>
<p>Place: {{project.district}}, {{project.state}}</p>
<br><br>
<p>_____________________________</p>
<p>{{project.architect_name}}</p>
<p>Architect</p>
</div>
</body>
</html>
"""

GOA_FORM_3_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
<style>
body { font-family: 'Times New Roman', serif; font-size: 12pt; line-height: 1.6; padding: 40px; }
.header { text-align: center; margin-bottom: 30px; }
.title { font-size: 16pt; font-weight: bold; margin-bottom: 10px; }
.subtitle { font-size: 14pt; margin-bottom: 20px; }
table { width: 100%; border-collapse: collapse; margin: 20px 0; }
th, td { border: 1px solid #000; padding: 8px; text-align: left; }
th { background-color: #f0f0f0; }
.amount { text-align: right; }
.signature { margin-top: 50px; }
</style>
</head>
<body>
<div class="header">
<div class="title">GOA REAL ESTATE REGULATORY AUTHORITY</div>
<div class="subtitle">FORM - 3</div>
<div class="subtitle">ENGINEER'S CERTIFICATE</div>
<div>(Cost Incurred)</div>
</div>

<div class="form-section">
<p><strong>Project Name:</strong> {{project.project_name}}</p>
<p><strong>RERA Registration No:</strong> {{project.rera_number}}</p>
<p><strong>Quarter:</strong> {{quarter}} {{year}}</p>
</div>

<table>
<tr>
<th>Sr. No.</th>
<th>Building/Wing Name</th>
<th>Estimated Cost (Rs.)</th>
<th>Cost Incurred (Rs.)</th>
<th>Completion %</th>
<th>Balance Cost (Rs.)</th>
</tr>
<!-- Building cost data -->
</table>

<div class="signature">
<p>Date: {{report_date}}</p>
<p>Place: {{project.district}}, {{project.state}}</p>
<br><br>
<p>_____________________________</p>
<p>{{project.engineer_name}}</p>
<p>Engineer</p>
<p>License No: {{project.engineer_license}}</p>
</div>
</body>
</html>
"""

GOA_FORM_4_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
<style>
body { font-family: 'Times New Roman', serif; font-size: 12pt; line-height: 1.6; padding: 40px; }
.header { text-align: center; margin-bottom: 30px; }
.title { font-size: 16pt; font-weight: bold; margin-bottom: 10px; }
.subtitle { font-size: 14pt; margin-bottom: 20px; }
table { width: 100%; border-collapse: collapse; margin: 20px 0; }
th, td { border: 1px solid #000; padding: 8px; text-align: left; }
th { background-color: #f0f0f0; }
.amount { text-align: right; }
.signature { margin-top: 50px; }
</style>
</head>
<body>
<div class="header">
<div class="title">GOA REAL ESTATE REGULATORY AUTHORITY</div>
<div class="subtitle">FORM - 4</div>
<div class="subtitle">CHARTERED ACCOUNTANT'S CERTIFICATE</div>
<div>(Project Cost)</div>
</div>

<div class="form-section">
<p><strong>Project Name:</strong> {{project.project_name}}</p>
<p><strong>RERA Registration No:</strong> {{project.rera_number}}</p>
<p><strong>Quarter:</strong> {{quarter}} {{year}}</p>
</div>

<h4>A. Land Cost</h4>
<table>
<tr><th>Particulars</th><th>Estimated Amount (Rs.)</th><th>Actual Amount (Rs.)</th></tr>
<tr><td>Land Acquisition Cost</td><td class="amount">{{project_cost.estimated_land_cost}}</td><td class="amount">{{project_cost.land_acquisition_cost}}</td></tr>
<tr><td>Development Rights Premium</td><td class="amount">-</td><td class="amount">{{project_cost.development_rights_premium}}</td></tr>
<tr><td>TDR Cost</td><td class="amount">-</td><td class="amount">{{project_cost.tdr_cost}}</td></tr>
<tr><td>Stamp Duty</td><td class="amount">-</td><td class="amount">{{project_cost.stamp_duty}}</td></tr>
<tr><td>Government Charges</td><td class="amount">-</td><td class="amount">{{project_cost.government_charges}}</td></tr>
<tr><td><strong>Total Land Cost</strong></td><td class="amount"><strong>{{project_cost.estimated_land_cost}}</strong></td><td class="amount"><strong>{{project_cost.total_land_cost}}</strong></td></tr>
</table>

<h4>B. Development Cost</h4>
<table>
<tr><th>Particulars</th><th>Estimated Amount (Rs.)</th><th>Actual Amount (Rs.)</th></tr>
<tr><td>Construction Cost</td><td class="amount">{{project_cost.estimated_development_cost}}</td><td class="amount">{{project_cost.construction_cost}}</td></tr>
<tr><td>Infrastructure Cost</td><td class="amount">-</td><td class="amount">{{project_cost.infrastructure_cost}}</td></tr>
<tr><td>Equipment Cost</td><td class="amount">-</td><td class="amount">{{project_cost.equipment_cost}}</td></tr>
<tr><td>Taxes & Statutory Charges</td><td class="amount">-</td><td class="amount">{{project_cost.taxes_statutory}}</td></tr>
<tr><td>Finance Cost</td><td class="amount">-</td><td class="amount">{{project_cost.finance_cost}}</td></tr>
<tr><td><strong>Total Development Cost</strong></td><td class="amount"><strong>{{project_cost.estimated_development_cost}}</strong></td><td class="amount"><strong>{{project_cost.total_development_cost}}</strong></td></tr>
</table>

<table>
<tr><td><strong>Total Estimated Cost</strong></td><td class="amount"><strong>{{project_cost.total_estimated_cost}}</strong></td></tr>
<tr><td><strong>Total Cost Incurred</strong></td><td class="amount"><strong>{{project_cost.total_cost_incurred}}</strong></td></tr>
<tr><td><strong>Balance Cost</strong></td><td class="amount"><strong>{{project_cost.balance_cost}}</strong></td></tr>
</table>

<div class="signature">
<p>Date: {{report_date}}</p>
<p>Place: {{project.district}}, {{project.state}}</p>
<br><br>
<p>_____________________________</p>
<p>{{project.ca_name}}</p>
<p>Chartered Accountant</p>
<p>Firm: {{project.ca_firm_name}}</p>
</div>
</body>
</html>
"""

GOA_FORM_5_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
<style>
body { font-family: 'Times New Roman', serif; font-size: 12pt; line-height: 1.6; padding: 40px; }
.header { text-align: center; margin-bottom: 30px; }
.title { font-size: 16pt; font-weight: bold; margin-bottom: 10px; }
.subtitle { font-size: 14pt; margin-bottom: 20px; }
table { width: 100%; border-collapse: collapse; margin: 20px 0; }
th, td { border: 1px solid #000; padding: 8px; text-align: left; }
th { background-color: #f0f0f0; }
.amount { text-align: right; }
.highlight { background-color: #ffffcc; }
.signature { margin-top: 50px; }
</style>
</head>
<body>
<div class="header">
<div class="title">GOA REAL ESTATE REGULATORY AUTHORITY</div>
<div class="subtitle">FORM - 5</div>
<div class="subtitle">CA CERTIFICATE - RECEIVABLE COMPLIANCE</div>
</div>

<div class="form-section">
<p><strong>Project Name:</strong> {{project.project_name}}</p>
<p><strong>RERA Registration No:</strong> {{project.rera_number}}</p>
<p><strong>Quarter:</strong> {{quarter}} {{year}}</p>
</div>

<h4>Receivable Compliance Calculation</h4>
<table>
<tr><td>Total Receivables (A)</td><td class="amount">Rs. {{receivables}}</td></tr>
<tr><td>Balance Cost (B)</td><td class="amount">Rs. {{balance_cost}}</td></tr>
<tr><td>Ratio (A/B)</td><td class="amount">{{ratio}}</td></tr>
</table>

<h4>RERA Deposit Requirement</h4>
<table>
<tr class="highlight">
<td><strong>If Ratio > 1:</strong> Deposit 70% of Receivables</td>
<td class="amount">Rs. {{deposit_70}}</td>
</tr>
<tr class="highlight">
<td><strong>If Ratio ≤ 1:</strong> Deposit 100% of Receivables</td>
<td class="amount">Rs. {{deposit_100}}</td>
</tr>
<tr>
<td><strong>Required RERA Deposit</strong></td>
<td class="amount"><strong>Rs. {{rera_deposit}}</strong></td>
</tr>
</table>

<div class="signature">
<p>Date: {{report_date}}</p>
<p>Place: {{project.district}}, {{project.state}}</p>
<br><br>
<p>_____________________________</p>
<p>{{project.ca_name}}</p>
<p>Chartered Accountant</p>
<p>Firm: {{project.ca_firm_name}}</p>
</div>
</body>
</html>
"""

GOA_FORM_6_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
<style>
body { font-family: 'Times New Roman', serif; font-size: 12pt; line-height: 1.6; padding: 40px; }
.header { text-align: center; margin-bottom: 30px; }
.title { font-size: 16pt; font-weight: bold; margin-bottom: 10px; }
.subtitle { font-size: 14pt; margin-bottom: 20px; }
table { width: 100%; border-collapse: collapse; margin: 20px 0; }
th, td { border: 1px solid #000; padding: 8px; text-align: left; }
th { background-color: #f0f0f0; }
.amount { text-align: right; }
.signature { margin-top: 50px; }
</style>
</head>
<body>
<div class="header">
<div class="title">GOA REAL ESTATE REGULATORY AUTHORITY</div>
<div class="subtitle">FORM - 6</div>
<div class="subtitle">AUDITOR'S CERTIFICATE</div>
<div>(Statement of Accounts)</div>
</div>

<div class="form-section">
<p><strong>Project Name:</strong> {{project.project_name}}</p>
<p><strong>RERA Registration No:</strong> {{project.rera_number}}</p>
<p><strong>Financial Year:</strong> {{year}}</p>
</div>

<h4>Statement of Accounts</h4>
<table>
<tr><td>Project Completion Percentage</td><td class="amount">{{completion_percentage}}%</td></tr>
<tr><td>Total Collections</td><td class="amount">Rs. {{amount_collected}}</td></tr>
<tr><td>Total Withdrawals/Utilization</td><td class="amount">Rs. {{cost_incurred}}</td></tr>
<tr><td>Balance in RERA Account</td><td class="amount">Rs. {{balance}}</td></tr>
</table>

<p>I hereby certify that I have audited the accounts of the above project and confirm that 
the funds collected from allottees have been utilized for the purposes of the project 
in accordance with the Real Estate (Regulation and Development) Act, 2016.</p>

<div class="signature">
<p>Date: {{report_date}}</p>
<p>Place: {{project.district}}, {{project.state}}</p>
<br><br>
<p>_____________________________</p>
<p>{{project.auditor_name}}</p>
<p>Auditor</p>
<p>Firm: {{project.auditor_firm_name}}</p>
</div>
</body>
</html>
"""

GOA_ANNEXURE_A_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
<style>
body { font-family: 'Times New Roman', serif; font-size: 10pt; line-height: 1.4; padding: 20px; }
.header { text-align: center; margin-bottom: 20px; }
.title { font-size: 14pt; font-weight: bold; margin-bottom: 10px; }
.subtitle { font-size: 12pt; margin-bottom: 15px; }
table { width: 100%; border-collapse: collapse; margin: 15px 0; font-size: 9pt; }
th, td { border: 1px solid #000; padding: 5px; text-align: left; }
th { background-color: #f0f0f0; text-align: center; }
.amount { text-align: right; }
.center { text-align: center; }
.total-row { font-weight: bold; background-color: #f5f5f5; }
</style>
</head>
<body>
<div class="header">
<div class="title">GOA REAL ESTATE REGULATORY AUTHORITY</div>
<div class="subtitle">ANNEXURE - A</div>
<div class="subtitle">STATEMENT OF RECEIVABLES</div>
</div>

<div class="form-section">
<p><strong>Project Name:</strong> {{project.project_name}}</p>
<p><strong>RERA Registration No:</strong> {{project.rera_number}}</p>
<p><strong>Quarter:</strong> {{quarter}} {{year}}</p>
</div>

<table>
<tr>
<th>Sr. No.</th>
<th>Unit No.</th>
<th>Building</th>
<th>Carpet Area (sq.ft.)</th>
<th>Buyer Name</th>
<th>Agreement Date</th>
<th>Sale Value (Rs.)</th>
<th>Amount Received (Rs.)</th>
<th>Balance Receivable (Rs.)</th>
</tr>
<!-- Sales data rows will be populated here -->
<tr class="total-row">
<td colspan="6" class="center">TOTAL</td>
<td class="amount">{{total_sales_value}}</td>
<td class="amount">{{amount_collected}}</td>
<td class="amount">{{receivables}}</td>
</tr>
</table>

<div class="signature">
<p>Date: {{report_date}}</p>
<p>Place: {{project.district}}, {{project.state}}</p>
<br><br>
<p>_____________________________</p>
<p>{{project.promoter_name}}</p>
<p>Promoter/Developer</p>
</div>
</body>
</html>
"""

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
